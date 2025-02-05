from starlette.responses import JSONResponse, HTMLResponse, StreamingResponse
from starlette.background import BackgroundTask
from ..settings import DB, list_status, factories
import time, datetime
from bson import ObjectId
from ..modules.users import authenticate
from ..modules import func, google_storage
from ..resources import templates, make_page_navi, get_page_data
from resources.db import convert_to_mongodb_time, convert_to_time
from copy import deepcopy
import asyncio
import base64
import traceback, json
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image, ImageDraw
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import code128
import base64
import random, copy
import urllib
import csv



def render_content(request, context):
    # Rend with context
    template = templates.get_template('api_render/index_chart.html')
    html_content = template.render(request=request, context=context)
    return html_content

def render_content_orders(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/api_orders.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way

def render_content_orders_other_team(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/api_orders_other_team.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way

def render_ff_ngoai_product(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/get_products_ff_ngoai.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way

def render_ff_ngoai_sizes(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/get_sizes_ff_ngoai.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way
    
def render_ff_ngoai_colors(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/get_colors_ff_ngoai.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way    

def render_analytics(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/analytics.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way
def render_private_analytics(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/private_analytics.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {traceback.format_exc()}")
        return None  # or handle the error in an appropriate way
def render_other_analytics(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/other_team_analytics.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way
def render_designer_analytics(request, context):
    try:
        # Rend with context
        template = templates.get_template('api_render/designer_analytics.html')
        html_content = template.render(request=request, context=context)
        return html_content
    except Exception as e:
        print(f"Error rendering template: {e}")
        return None  # or handle the error in an appropriate way

# 
# Api
#
@authenticate
async def filter_order(request):
    db_client = request.state.db
    user = request.state.user_data
    user_team = user['team']
    groups = user['groups']
    # # check permission
    # if user['groups'] == ['designer']:
    #     return JSONResponse({'status': 400, 'message': f"You don't have permission to access this page!"})
    try:
        if request.method == "POST":
            data = await request.json()
            date = data['date']
            status = data['status']
            try: factory = data['factory']
            except: factory = []
            try: designer = data['designer']
            except: designer = []
            try: shop = data['shop']
            except: shop = []
            try: filter_team = data['team']
            except: filter_team = [user_team]
            product_type = data['type']
            try: seller = data['seller']
            except: seller = []
            try: search = data['order_id']
            except: search = ''
            try: shipping_day = data['shipping_day']
            except: shipping_day = ''
            try: ff_provider_status = data['ff_provider_status']
            except: ff_provider_status = ''
            try: page = data['page']
            except: page = 1
            try: route = data['route']
            except: route = ''
            try: have_tracking = data['have_tracking']
            except: have_tracking = ''
            try: tracking_synced = data['tracking_synced']
            except: tracking_synced = ''
            

            seller = [id.lower() for id in seller]

            if 'xuong_bn' in user['username']:
                factory = ['Thủy']
            if 'xuong_thao' in user['username']:
                factory = ['Thảo']
            if 'xuong_bach' in user['username']:
                factory = ['Bách']

            if user_team == ['']:
                user_team = request.state.user_data['team']

            if 'vmm_team' in user_team:
                filter_team = ['vmm_team']

            try:
                start_date = convert_to_mongodb_time(date.split(' - ')[0], 'start')
                end_date = convert_to_mongodb_time(date.split(' - ')[1], 'end')
            except:
                start_date = None
                end_date = None

            if factory == [''] or factory == ['None']:
                factory = ['', None]
            
            if status == [''] or status == ['None']:
                status = ['', None]

            try:
                table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
            except:
                table_column = {}

            params = {
                'user': user,
                'groups': groups,
                'start_date': start_date,
                'end_date': end_date,
                'filter_team': filter_team,
                'shop': shop,
                'product_type': product_type,
                'status': status,
                'factory': factory,
                'designer': designer,
                'seller': seller,
                'search': search,
                'shipping_day': shipping_day,
                'ff_provider_status': ff_provider_status,
                'have_tracking': have_tracking,
                'tracking_synced': tracking_synced,
            }
            query = func.build_query_filter_orders(params, is_api=True)

            is_ff_ngoai = False

            if route == '/ff_ngoai':
                query.update({
                    # "_original_data.line_items.wrong": {"$ne": "sku_wrong"},
                    "_original_data.line_items.fulfillment_status": {"$in": ['FF Ngoai']},
                    "_team": {"$in": filter_team + ['nb_team']}
                })  
                is_ff_ngoai = True
            elif route == '/wrong_sku':
                query.update({
                    "_original_data.line_items.wrong": "sku_wrong"
                })
            elif route == '/wrong_product_type':
                product_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
                product_types = [product_type['product_type'] for product_type in product_types if product_type['product_type'] != '']
                query.update({
                    "_original_data.line_items": {
                        "$elemMatch": {
                            "product_type": {
                                "$nin": product_types
                            }
                        }
                    }
                })
            elif route == '/order_assigned':
                query.update({
                    "_original_data.line_items.assign_designer.designer": user['username'],
                    "_original_data.line_items.assign_designer.design_checked": False,
                    "_original_data.line_items.wrong": {"$ne": "sku_wrong"},
                })
            elif route == '/orders':
                query.update({
                    "_original_data.line_items.wrong": {"$ne": "sku_wrong"},
                })
            elif route == '/check_design':
                try:
                    seller_id = user['special_id']
                except:
                    seller_id = ''
                query.update({
                    "_original_data.line_items.wrong": {"$ne": "sku_wrong"},
                    "_original_data.line_items.fulfillment_status": {"$in": ["Design uploaded", "Sai design"]},
                    "_original_data.seller_id": seller_id
                })
            
            # Get page
            page_number, limit_, skip_ = await get_page_data(request, page=page)

            # make page navi
            count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
            page_navi = make_page_navi(request, count_data, limit_, page_number, 'orders')

            if 'mod' in user['groups'] and query.get('_original_data.line_items.fulfillment_status') == {'$ne': 'FF Ngoai'}:
                query.update({'_original_data.line_items.fulfillment_status': {"$in": ['', 'Design uploaded'], '$ne': 'FF Ngoai'}})
            if user_team in ['xuong_bn', 'xuong_thao', 'xuong_bach']:
                query.pop('_team', None)
            try:
                print('query_api: ', query)
                orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None, limit_=0, sort_="create_at", skip_=skip_)
                formatted_orders = await func.format_show_orders(orders)

                p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
                nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
                nb_product_types = sorted(set(nb_product_types))

                if user_team in ['xuong_bn', 'da_team', 'xuong_thao', 'xuong_bach']:
                    status_list = ["Ready", "Pending", "Half Done", "Doing", "Done recently", "In", "Out", "Sai design"]
                elif 'designer' in user['groups'] or 'mod' in user['groups']:
                    status_list = ["None", "Pending", "Ready", "Designing", "Design uploaded", "Rejected", "Sai design", "Support check", "Cancel", "Refund"]
                elif 'admin' in user['groups'] or 'dev' in user['groups']:
                    status_list = list_status + ['Cancel', 'Refund']
                else:
                    status_list = list_status

                if params['factory'] == ['', None] or params['factory'] == ['None']:
                    factory = ['None']
                else:
                    factory = params['factory']
                if params['status'] == ['', None ] or params['status'] == ['None']:
                    status = ['None']
                else:
                    status = params['status']   
                if  'start_date' in params and params['start_date'] != None:
                    date = params['start_date'].strftime("%Y/%m/%d") + ' - ' + params['end_date'].strftime("%Y/%m/%d")
                else:
                    date = ''      
                paramurl = {
                    'date' : date,
                    'status': ','.join(status),
                    'factory': ','.join(factory) ,
                    'designer':  ','.join(params.get('designer',[])),
                    'type':   ','.join(params.get('product_type',[])),
                    'shop':   ','.join(params.get('shop',[])),
                    'team': ','.join(params.get('filter_team',[])),
                    'seller': ','.join(params.get('seller',[])) ,
                    'shipping_day': '',

                }
                param_url = urllib.parse.urlencode(paramurl).replace('%5B%27%27%5D','').replace('%5B%5D','')
                context = {
                    "orders": formatted_orders,
                    "nb_product_types": nb_product_types,
                    "table_column": table_column.get('table_column', []),
                    "status": {"options": status_list, "selected": status},
                    "factories": {"options": factories, "selected": factory},
                    "page_navi": page_navi,
                    "is_ff_ngoai": is_ff_ngoai,
                    "have_tracking": have_tracking,
                    "tracking_synced": tracking_synced,
                    'param_url':param_url,
                }

                if user_team in ['nb_team', 'vmm_team']:
                    html = render_content_orders(request, context)
                else:
                    html = render_content_orders_other_team(request, context)
                return HTMLResponse(html)
            except Exception as e:
                return JSONResponse({'status': 400, 'message': f"Get orders failed - {traceback.format_exc()}!"})
    except:
        print(f"Get orders failed! - {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"Get orders failed! - {traceback.format_exc()}!"})

async def change_product_quantity(request):
    # Config
    db_client = request.state.db
    try:
        if request.method == "POST":
            data = await request.json()
            _id = data['id']
            quantity = data['quantity']
            type_ = data['type']
            if type_ == 'warehouse':
                product = await db_client.find_one({'_id': ObjectId(_id)}, collection_name=DB['COL_WAREHOUSE'])
            else:
                product = await db_client.find_one({'_id': ObjectId(_id)}, collection_name=DB['COL_PRODUCTS'])
            if product is None:
                return JSONResponse({'status': 400, 'message': f"Product with id {_id} not found!"})
            else:
                old_quantity = product.get('quantity')
                new_quantity = int(quantity)
                product['quantity'] = new_quantity
                product['user_logs'].append(f"{request.state.user_data['username']} changed quantity from {old_quantity} -> {new_quantity} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                if type_ == 'warehouse':
                    today = convert_to_mongodb_time(time.time()).strftime("%Y-%m-%d")
                    ok = False
                    for i in range(len(product['data_date'])):
                        if product['data_date'][i]['date'].strftime("%Y-%m-%d") == today:
                            product['data_date'][i]['quantity'] = new_quantity
                            ok = True
                            break
                    if not ok:
                        product['data_date'].append({'date': convert_to_mongodb_time(time.time()), 'quantity': new_quantity})
                    result = await db_client.update_one(_id, product, collection_name=DB['COL_WAREHOUSE'])
                else:
                    result = await db_client.update_one(_id, product, collection_name=DB['COL_PRODUCTS'])
                return JSONResponse({'status': 200, 'message': f"Product {product['name']} updated quantity {old_quantity} -> {new_quantity} successfully!", 'new_quantity': new_quantity})
    except:
        print(f"Change product quantity failed! - {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"Change product quantity failed! - {traceback.format_exc()}!"})

async def change_order_status(request):
    # Config
    db_client = request.state.db
    try:
        if request.method == "POST":
            try:
                data = await request.json()
            except:
                return JSONResponse({'status': 400, 'message': f"Invalid request!"})    
            order_id = data['orderId']
            product_id = data['productId']
            status = data['status']
            note = data['note']
            query = {"_original_data.order_id": str(order_id)}
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])


            if order is None:
                return JSONResponse({'status': 400, 'message': f"Order with id {order_id} not found!"})
            
            
            for line_item in order['_original_data']['line_items']:
                if line_item['line_item_id'] == str(product_id):
                    if 'user_logs' in line_item:
                        user_log = f"{request.state.user_data['username']} changed status from {line_item['fulfillment_status']} -> {status} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        line_item['user_logs'].append(user_log)
                    else:
                        user_log = [f"{request.state.user_data['username']} changed status from {line_item['fulfillment_status']} -> {status} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                        line_item['user_logs'] = user_log
                    line_item['user_logs'] = [user_log] if 'user_logs' not in line_item else line_item['user_logs'] + [user_log]
                    line_item['fulfillment_status'] = status

                    if status == 'Sai design':
                        line_item['design_note'] = note
                        try:
                            line_item['assign_designer']['design_checked'] = False
                        except:
                            pass
                    # update line_item basecost
                    if status in ['Ready']:
                        # get product basecost
                        filter_ = {
                            'base_cost': 1,
                            'us_shipping_1st': 1,
                            'us_shipping_additional': 1,
                            'ww_shipping_1st': 1,
                            'ww_shipping_additional': 1,
                        }
                        products = await db_client.find_many({'product_type': line_item['product_type']}, collection_name=DB['COL_PRODUCTS'], filter_=filter_)
                        # update line_item basecost
                        if line_item['base_cost'] == 0:
                            try:
                                if len(products) == 1:
                                    product = products[0]
                                else:
                                    product = next((p for p in products if p in line_item['size']), None)
                                if product is not None:
                                    basecost = round(float(product['base_cost']) * int(line_item['quantity']), 2)
                                    line_item['base_cost'] = basecost
                            except:
                                return JSONResponse({'status': 400, 'message': f"Product {line_item['product_type']} doesn't have basecost!"})
                        else:
                            basecost = 0

                        # update line_item ship_cost
                        if 'ship_cost' not in line_item or line_item['ship_cost'] == 0:
                            # Provisional shipping fee
                            ship_cost = 0
                            if order['_original_data']['delivery_info']['country'].lower() in ['us', 'united states']:
                                product = products[0]
                                ship_cost = product['us_shipping_1st']
                                if int(line_item['quantity']) > 1:
                                    ship_cost += round(product['us_shipping_additional'] * (int(line_item['quantity']) - 1), 2)
                            else:
                                ship_cost = product['ww_shipping_1st']
                                if int(line_item['quantity']) > 1:
                                    ship_cost += round(product['ww_shipping_additional'] * (int(line_item['quantity']) - 1), 2)
                            line_item['ship_cost'] = ship_cost
                        else:
                            ship_cost = 0

                        # update order basecost
                        if order['_team'] != 'nb_team':
                            query = {
                                'name': order['_team'],
                            }
                            team = await db_client.find_one(query, collection_name=DB['COL_TEAMS'])
                            if team is None:
                                return JSONResponse({'status': 400, 'message': f"Team with name {order['_team']} not found!"})
                            else:
                                if 'balance' not in team:
                                    team['balance'] = 0
                                current_balance = team['balance']
                                balance = round(float(current_balance) - basecost - ship_cost, 2)
                                # if balance <= 0:
                                #     return JSONResponse({'status': 400, 'message': f"Team {order['_team']} doesn't have enough money to pay for this order!"})
                                # else:
                                team['balance'] = balance
                                await db_client.update_one(team['_id'], team, collection_name=DB['COL_TEAMS'])

                                if 'user_logs' in order['_original_data']:
                                    order['_original_data']['user_logs'].append(f"Basecost: {basecost} - {convert_to_mongodb_time(time.time())}")
                                else:
                                    order['_original_data']['user_logs'] = [f"Basecost: {basecost} - {convert_to_mongodb_time(time.time())}"]
                    
                    if status in ['Ready'] and line_item['link_des'] == '':
                        if order['_team'] != 'nb_team':
                            return JSONResponse({'status': 400, 'message': f"This order doesn't have link des!"})
                        else:
                            try:
                                # find link des by sku in COL_REUSE_DES
                                query = {
                                    'sku': line_item['sku'],
                                }
                                reuse_des = await db_client.find_one(query, collection_name=DB['COL_REUSE_DES'])
                                if reuse_des is not None:
                                    line_item['link_des'] = reuse_des['link_des']
                                    line_item['assign_designer'] = {
                                        'designer': 'reuse_design',
                                        'design_checked': True,
                                        'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                    }
                            except:
                                pass

                    await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])


            return JSONResponse({'status': 200, 'message': f"Order {product_id} updated status {status} successfully!"})
    except Exception as e:
        data = await request.json()
        print(f"Change order status failed! - {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"{data}Change order status failed - {traceback.format_exc()}!"})

async def change_order_factory(request):
    # Config
    db_client = request.state.db
    try:
        if request.method == "POST":
            data = await request.json()
            order_id = data['orderId']
            product_id = data['productId']
            factory = data['factory']
            query = {
                "_original_data.order_id": str(order_id),
            }
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])
            if order is None:
                return JSONResponse({'status': 400, 'message': f"Orded with id {order_id} not found!"})
            else:
                for product in order['_original_data']['line_items']:
                    if product['line_item_id'] == product_id:
                        if 'user_logs' in product:
                            product['user_logs'].append(f"{request.state.user_data['username']} changed factory from {product['factory']} -> {factory} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                        else:
                            product['user_logs'] = [f"{request.state.user_data['username']} changed status from {product['factory']} -> {factory} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                        product['factory'] = factory
                await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
                
                return JSONResponse({'status': 200, 'message': f"Order {product_id} updated factory {factory} successfully!"})
    except:
        print(f"Change order factory failed! - {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"Change order factory failed! - {traceback.format_exc()}!"})

async def change_product_field(request):
    # Config
    db_client = request.state.db
    try:
        if request.method == "POST":
            data = await request.json()
            order_id = data['orderId']
            line_item_id = data['lineItemId']
            field = data['field']
            new_value = data['value']

            if field in ['shipping_day', 'dispatch_day']:
                # replace dd/mm/yyyy -> yyyy-mm-dd
                new_value = datetime.datetime.strptime(new_value, '%d/%m/%Y').strftime('%Y-%m-%d')
                new_value = convert_to_mongodb_time(new_value, 'start')
            
            if field == 'quantity':
                new_value = int(new_value)
            if field == 'base_cost':
                new_value = float(new_value)

            query = {"_original_data.line_items.fulfillment_order_id": str(order_id)}
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])

            if order is None:
                return JSONResponse({'status': 400, 'message': f"Order with id {order_id} not found!"})
            
            # find products to get product_type
            products = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={'product_type': 1, 'xuong': 1})
            
            nambe = [product['product_type'] for product in products if 'xuong' in product and product['xuong'] == 'Nambe']
            thuy = [product['product_type'] for product in products if 'xuong' in product and product['xuong'] == 'Thủy']
            tung = [product['product_type'] for product in products if 'xuong' in product and product['xuong'] == 'Tùng']
            thao = [product['product_type'] for product in products if 'xuong' in product and product['xuong'] == 'Thảo']

            digital = ['DGT']
            status = None
            if field == 'product_type':
                if new_value in nambe:
                    xuong = 'Nambe'
                elif new_value in thuy:
                    xuong = 'Thủy'
                elif new_value in tung:
                    xuong = 'Tùng'
                elif new_value in thao:
                    xuong = 'Thảo'
                elif new_value in digital:
                    xuong = ''
                    status = 'Done'
                elif new_value == '':
                    xuong = ''
                    status = ''
                else:
                    xuong = None
                    status = None
            else:
                xuong = None
            
            # Update the desired field in the line item
            for line_item in order['_original_data']['line_items']:
                if line_item['line_item_id'] == str(line_item_id):
                    try: old_value = line_item[field]
                    except: old_value = ''
                    line_item[field] = new_value
                    if xuong != None:
                        line_item['factory'] = xuong
                    if status != None:
                        line_item['fulfillment_status'] = status
                    if field == 'product_type' and new_value in ['AF']:
                        line_item['wrong'] = ''
                    if 'user_logs' in line_item:
                        line_item['user_logs'].append(f"{request.state.user_data['username']} changed {field} from {old_value} -> {new_value} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    else:
                        line_item['user_logs'] = [f"{request.state.user_data['username']} changed {field} from {old_value} -> {new_value} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

            return JSONResponse({'status': 200, 'message': f"Order {line_item_id} updated {field} successfully!", 'xuong': xuong, 'ff_status': status})
    except:
        print(f"Change product field failed! - {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"Change product field failed! - {traceback.format_exc()}!"})



@authenticate
async def delete_product(request):
    db_client = request.state.db

    if request.method == "POST":
        data = await request.json()
        try: 
            product_id = ObjectId(data['id'])
        except:
            product_id = data['id']
        type_ = data['type']

        # Find the product to be deleted
        if type_ == 'warehouse':
            product = await db_client.find_one({'_id': product_id}, collection_name=DB['COL_WAREHOUSE'])
        elif type_ == 'fulfillment':
            product = await db_client.find_one({'_original_data.order_id': product_id}, collection_name=DB['COL_FULFILLMENTS'])
        else:
            product = await db_client.find_one({'_id': product_id}, collection_name=DB['COL_PRODUCTS'])
        if not product:
            return JSONResponse({'status': 400, 'message': f"Product not found!"})

        # Backup the product and add logs
        product_backup = deepcopy(product)
        product_backup.pop('_id')
        product_backup['deleted_at'] = convert_to_mongodb_time(time.time())
        if 'user_logs' in product_backup:
            product_backup['user_logs'].append(f"{request.state.user_data['username']} deleted product at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        else:
            product_backup['user_logs'] = [f"{request.state.user_data['username']} deleted product at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
        if type_ == 'warehouse':
            await db_client.insert_one(product_backup, DB['COL_WAREHOUSE_BACKUPS'])
            await db_client.delete({'_id': product_id}, collection_name=DB['COL_WAREHOUSE'])
        elif type_ == 'fulfillment':
            await db_client.insert_one(product_backup, DB['COL_ORDERS_BACKUPS'])
            order = await db_client.find_one({'_original_data.order_id': product_id}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
            await db_client.delete({'_id': order['_id']}, collection_name=DB['COL_FULFILLMENTS'])
        else:
            await db_client.insert_one(product_backup, DB['COL_PRODUCT_BACKUPS'])
            await db_client.delete({'_id': product_id}, collection_name=DB['COL_PRODUCTS'])
            
        return JSONResponse({'status': 200, 'message': f"Product deleted successfully!"})
    return JSONResponse({'status': 400, 'message': f"Invalid request method!"})

@authenticate
async def delete_design_type(request):
    db_client = request.state.db

    if request.method == "POST":
        data = await request.json()
        try: 
            _id = ObjectId(data['id'])
        except:
            _id = data['id']

        # Find the product to be deleted
        design_type = await db_client.find_one({'_id': _id}, collection_name=DB['COL_DESIGN'])
        if not design_type:
            return JSONResponse({'status': 400, 'message': f"Product not found!"})

        await db_client.delete({'_id': _id}, collection_name=DB['COL_DESIGN'])
            
        return JSONResponse({'status': 200, 'message': f"Product type deleted successfully!"})
    return JSONResponse({'status': 400, 'message': f"Invalid request method!"})

@authenticate
async def delete_team(request):
    db_client = request.state.db

    if request.method == "POST":
        data = await request.json()
        team_id = ObjectId(data['id'])

        # Find the product to be deleted
        team = await db_client.find_one({'_id': team_id}, collection_name=DB['COL_TEAMS'])
        
        if not team:
            return JSONResponse({'status': 400, 'message': f"Team not found!"})
        for user in team['users']:
            await db_client.update_one(user, {"team": None}, collection_name=DB['COL_USERS']) 
        # delete team
        await db_client.delete({'_id': team_id}, collection_name=DB['COL_TEAMS'])
            
        return JSONResponse({'status': 200, 'message': f"Product deleted successfully!"})
    return JSONResponse({'status': 400, 'message': f"Invalid request method!"})

@authenticate
async def delete_shop(request):
    db_client = request.state.db

    if request.method == "POST":
        data = await request.json()
        shop_id = ObjectId(data['id'])

        # delete shop
        await db_client.delete({'_id': shop_id}, collection_name=DB['COL_SHOP'])    

        return JSONResponse({'status': 200, 'message': f"Shop deleted successfully!"})
    

async def get_range_date_products(request):
    db_client = request.state.db

    if request.method == "POST":
        data = await request.json()
        start = data['start']
        end = data['end']
        search = [data['search']]


        start_datetime = convert_to_mongodb_time(start, "start")
        end_datetime = convert_to_mongodb_time(end, "end")
        query = [
            {"$match": {"data_date.date": {"$gte": start_datetime, "$lte": end_datetime}, "name": {"$in": search}}},
            {"$project": {"name": 1, "data_date.date": 1, "data_date.quantity": 1}},
            {"$unwind": "$data_date"},
            {"$match": {"data_date.date": {"$gte": start_datetime, "$lte": end_datetime}, "name": {"$in": search}}},
            {"$group": {"_id": {"name": "$name", "date": "$data_date.date"}, "quantity": {"$sum": "$data_date.quantity"}}},
            {"$project": {"name": "$_id.name", "date": "$_id.date", "quantity": 1, "_id": 0}}
        ]
        # Find the products 
        products = await db_client.query_aggregate(query, collection_name=DB['COL_WAREHOUSE'])

        result = func.format_data(products)

        html = render_content(request, {'products': result})
        return JSONResponse({'status': 200, 'message': f"Get products successfully!", 'html': html})
    
async def get_range_date_orders(request):
    db_client = request.state.db

    if request.method == "POST":
        data = await request.json()
        start = data['start']
        end = data['end']

        start_datetime = convert_to_mongodb_time(start, "start")
        end_datetime = convert_to_mongodb_time(end, "end")
        query = {
            "_update_time": {"$gte": start_datetime, "$lte": end_datetime},
        }
        # Find the products 
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None, limit_=100, sort_="_update_time", skip_=0)
        orders = await func.format_show_orders(db_client, orders)
        html = render_content_orders(request, {'orders': orders})
        
        return JSONResponse({'status': 200, 'message': f"Get products successfully!", 'html': html})

async def table_column(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        column = data['column']
        user = data['user']

        try:
            user_found = await db_client.find_one({'username': user}, collection_name=DB['COL_USERS'])

            await db_client.update_one(user_found['_id'], {'table_column': column}, collection_name=DB['COL_USERS'])

            return JSONResponse({'status': 200, 'message': f"Get columns successfully!"})
        except Exception as e:
            return JSONResponse({'status': 400, 'message': f"Get columns failed - {e}!"})

async def update_merge_orders(selectedValue, db_client):
     # order = await db_client.find_one({'_original_data.products.id': selectedValue}, collection_name=DB['COL_FULFILLMENTS'], filter_={'products': 1})
    order = await db_client.find_one({'_original_data.line_items.product_id': selectedValue['productId']}, collection_name=DB['COL_FULFILLMENTS'], filter_={})
    if not order:
        return JSONResponse({'status': 400, 'message': f"Order not found!"})
    items_seclected = [ product for product in order['_original_data']['line_items'] if product['product_id'] == selectedValue['productId']  ]
    copy_item = items_seclected[0].copy()    
    copy_item['quantity'] = 0     
    for product in items_seclected:
        copy_item['quantity'] += product['quantity']
        order['_original_data']['line_items'].remove(product)
    order['_original_data']['line_items'].append(copy_item)
    await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

async def merge_orders(request):
    db_client = request.state.db
    data = await request.json()
    selectedValues = data['selectedValues']
    setSelectedValues = [dict(s) for s in set(frozenset(d.items()) for d in selectedValues)]
    await asyncio.gather(*(update_merge_orders(selectedValue, db_client) for selectedValue in setSelectedValues))       
    return JSONResponse({'status': 200, 'message': f"Get products successfully!"})    

def evenly_distribute_number(number, num_elements):
    element = number // num_elements  # Integer division
    remainder = number % num_elements
    result = [element] * num_elements  # Create an array with elements equal to the integer division
    if number == 2:
        result = [1, 1] 
    else:    
        for i in range(remainder):  # Distribute the remainder among the first few elements
            result[i] += 1
      
    return result

async def update_split_orders(selectedValue, db_client):
    order = await db_client.find_one({'_original_data.line_items.product_id': selectedValue['productId']}, collection_name=DB['COL_FULFILLMENTS'], filter_={})
    if not order:
        return JSONResponse({'status': 400, 'message': f"Order not found!"})
    for item in order['_original_data']['line_items']:
        a = 1
        products = []
        if item['product_id'] == selectedValue['productId']:
            list_split = evenly_distribute_number(item['quantity'],1+ int(selectedValue['split']))
            for split in list_split:
                copy_item = item.copy()         
                copy_item['quantity'] = split
                copy_item['line_item_id'] = item['line_item_id'] + f'-{a+1}'
                if item['num_color_kit'] != 0:
                    copy_item['num_color_kit'] = split
                if item['num_holder'] != 0:    
                    copy_item['num_holder'] = split
                if item['base_cost'] != 0:
                    copy_item['basecost'] = item['basecost'] / item['quantity'] * split
                products.append(copy_item)
                a += 1
            order['_original_data']['line_items'].remove(item)
            if products != []: 
                order['_original_data']['line_items'].extend(products)
    await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
    
async def split_orders(request):
    db_client = request.state.db
    data = await request.json()
    selectedValues = data['selectedValues']
    await asyncio.gather(*(update_split_orders(selectedValue, db_client) for selectedValue in selectedValues))
    return JSONResponse({'status': 200, 'message': f"Get products successfully!"})

async def forward_shipping(request):
    db_client = request.state.db
    tasks = []
    data = await request.json()
    selectedValues = data['selectedValues']
    for item in selectedValues:
        lineItemId = str(item['lineItemId'])
        order_id = str(item['orderId'])
        tasks.append(asyncio.create_task(task_get_shipping(db_client, order_id, lineItemId)))
    shippings = await asyncio.gather(*tasks)
    task = BackgroundTask(send_to_shippings, shippings=shippings)

    return JSONResponse({'status': 200, 'message': f"Forward shipping success"}, background=task)

async def send_to_shippings(shippings):
    async_rqs = func.async_rq(rqs=7, second=1, time_out=90, max_rq_tasks=77, rate_limit=True)
    tasks = []

    login = await async_rqs.async_request_session(method='POST', url='https://api.hanoinhanh.com/auth/login', payload={'username': '3102222666', 'password': 'nambe2611'}, type="JSON")

    token = login['html']['access_token']

    for shipping in shippings:
        if shipping != {}:
            tasks.append(asyncio.create_task(task_send_to_shipping(async_rqs, shipping, token)))
    await asyncio.gather(*tasks)

async def task_send_to_shipping(async_rqs, shipping, token):
    url = 'https://api.hanoinhanh.com/customer/order'
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'multipart/form-data'
    }
    await async_rqs.async_request_session(method='POST', url=url, payload=shipping, headers=headers, type="FORM")

async def task_get_shipping(db_client, order_id, lineItemId):
    order = await db_client.find_one({"_original_data.order_id": str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
    mongo_id = order['_id']
    if not order:
        return {}
    order = order['_original_data']
    product = next((product for product in order['line_items'] if str(product['line_item_id']) == str(lineItemId)), None)    
    if not product:
        return {}
    index_ = order['line_items'].index(product)
    order['line_items'][index_]['fulfillment_status'] = 'Shipping'
    country_name = order['delivery_info']['country']
    country = await db_client.find_one({"country_name": country_name}, collection_name=DB['COL_COUNTRY'])
    if not country:
        return {}
    delivery_country_id = country['id']

    shipping = {
        "customer_id": 62,
        "description": order_id,
        "order_no": order_id,
        "is_international": 1,
        "customer_phone": "3102222666",
        "customer_address": "Ha Noi, Viet Nam",
        "customer_contact_name": "ms Yen",
        "receiver_name": order['delivery_info']['name'],
        "delivery_country_id": delivery_country_id,
        "delivery_county": order['delivery_info']['state'],
        "delivery_city": order['delivery_info']['city'],
        "delivery_address": order['delivery_info']['address1'],
        "delivery_phone": order['delivery_info']['phone'],
        "postcode": order['delivery_info']['zip'],
    }
    all_fulfillment_status_shiped = True
    for item in order['line_items']:
        if item['fulfillment_status'] != 'Shipping':
            all_fulfillment_status_shiped = False
    update_ = {}
    if all_fulfillment_status_shiped:
        update_ = {
            f'_original_data.line_items.{index_}.fulfillment_status': 'Shipping',
            '_original_data.status': 'Shipping'
        }
            
    else:
        update_ = {
                f'_original_data.line_items.{index_}.fulfillment_status': 'Shipping'
            }  
           
    await db_client.update_one(mongo_id,update_ , collection_name=DB['COL_FULFILLMENTS'])
    return shipping

async def archive_order(request):
    db_client = request.state.db
    data = await request.json()
    selectedValues = data['selectedValues']
    for selectedValue in selectedValues:
        order = await db_client.find_one({'_original_data.line_items.line_item_id': str(selectedValue['lineItemId'])}, collection_name=DB['COL_FULFILLMENTS'], filter_={})
        if not order:
            return JSONResponse({'status': 400, 'message': f"Order not found!"})
        productss = []
        for item in order['_original_data']['line_items']:
            if item['line_item_id'] == str(selectedValue['lineItemId']):
                item['is_archived'] = True
                if 'user_logs' in item:
                    item['user_logs'].append(f"Archived at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                else:
                    item['user_logs'] = [f"Archived at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
            productss.append(item)
        order['_original_data']['line_items'] = productss
        await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
    return JSONResponse({'status': 200, 'message': f"Archive Product Done!"})

@authenticate
async def topup(request):
    db_client = request.state.db
    if request.method == "POST":
        form_data = await request.form()
        total = form_data['total']
        team = form_data['team']
        transaction = form_data['transaction']
        platform = form_data['platform']
        currency = form_data['currency']
        img = form_data['img'].file
        encoded_string = base64.b64encode(img.read()).decode('utf-8')

        # check last id invoice
        last_invoice = await db_client.find_one_sort_lasted({}, collection_name=DB['COL_INVOICE'])
        if last_invoice is None:
            id = 1
        else:
            id = last_invoice['id'] + 1
        
        #  insert topup
        topup = {
            "id": id,
            "total": round(float(total), 2),
            "team": team,
            "status": "pending",
            "transaction": transaction,
            "platform": platform,
            "img": encoded_string,
            "currency": currency,
            "created_at": convert_to_mongodb_time(time.time())
        }
        try:
            await db_client.insert_one(topup, DB['COL_INVOICE'])
            return JSONResponse({'status': 200, 'message': f"Topup successfully!"})
        except:
            return JSONResponse({'status': 400, 'message': f"Topup failed!"})

@authenticate
async def upload_order_design(request):
    db_client = request.state.db
    user_name = request.state.user_data['username']
    if request.method == "POST":
        try:
            data = await request.form()
        except:
            return JSONResponse({'message': f"Uploaded failed! - {traceback.format_exc()}"}, status_code=400)
        order_id = data.get('order_id', '')
        product_id = data.get('product_id', '')
        line_item_id = data.get('line_item_id', '')
        shop = data.get('shop', '')
        seller = data.get('seller', '')
        order_date = data.get('order_date', '')
        files = data.getlist('files')
        if order_id == "" or product_id == "":
            return JSONResponse({'status': 400, 'message': f"Missing order_id or product!"})

        # Handle the uploaded files and convert them to Base64
        files = []
        for file_field in data.keys():
            file = data[file_field]
            if hasattr(file, 'filename'):
                content = await file.read()
                file_base64 = base64.b64encode(content).decode('utf-8')
                files.append({
                    'filename': file.filename,
                    'content_base64': file_base64,
                })

        # Perform any other processing you need with the form data and uploaded files
        # ...
        try:
            link_folder = google_storage.upload_to_bucket(seller, shop, order_date, order_id, product_id, files)

            # update link_folder to product in order
            order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
            if not order:
                return JSONResponse({'status': 400, 'message': f"Order not found!"})
            for item in order['_original_data']['line_items']:
                if item['line_item_id'] == product_id:
                    item['link_des'] = link_folder

                    if item['fulfillment_status'] in ['', 'Designing', 'Rejected', 'Sai design']:
                        if 'assign_designer' in item:
                            updated_at = convert_to_time(time.time())
                            item['fulfillment_status'] = 'Design uploaded'
                            item['assign_designer']['updated_at'] = updated_at
                        else:
                            item['fulfillment_status'] = 'Design uploaded'
                            item['assign_designer'] = {
                                'designer': user_name,
                                'created_at': convert_to_time(time.time()),
                                'updated_at': convert_to_time(time.time()),
                                'design_checked': False,
                            }
                    if user_name == 'vmm_team':
                        item['fulfillment_status'] = 'Ready'
                        item['assign_designer'] = {
                            'designer': 'vmm_team',
                            'created_at': convert_to_mongodb_time(time.time()),
                            'design_checked': True,
                            'checker': '',
                            'total_rejected': 0
                        }

                    
                    if 'user_logs' in item and item['user_logs'] != None:
                        user_logs = item.get('user_logs', [])
                        user_logs.append(f"{user_name} Uploaded design at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                        item['user_logs'] = user_logs
                    else:
                        user_logs = [f"{user_name} Uploaded design at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                        item['user_logs'] = user_logs

            await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
            return JSONResponse({'message': "Uploaded successfully!", 'link_folder': link_folder})
        except Exception as e:
            print(f"Uploaded failed! - {traceback.format_exc()}")
            return JSONResponse({'message': "Uploaded failed!"}, status_code=400)

async def upload_order_design_reuse(request):
    db_client = request.state.db
    user_name = request.state.user_data['username']
    if request.method == "POST":
        try:
            data = await request.form()
            # 
            order_id = data.get('order_id', '')
            product_id = data.get('product_id', '')
            sku = data.get('sku', '')
            files = data.getlist('files')
            if order_id == "" or product_id == "":
                return JSONResponse({'status': 400, 'message': f"Missing order_id or product!"})

            # Handle the uploaded files and convert them to Base64
            files = []
            for file_field in data.keys():
                file = data[file_field]
                if hasattr(file, 'filename'):
                    content = await file.read()
                    file_base64 = base64.b64encode(content).decode('utf-8')
                    files.append({
                        'filename': file.filename,
                        'content_base64': file_base64,
                    })
        except:
            return JSONResponse({'message': f"Uploaded failed! - {traceback.format_exc()}"}, status_code=400)
        # Perform any other processing you need with the form data and uploaded files
        # ...
        try:
            link_folder = google_storage.upload_to_bucket_reuse(sku, files)

            # update link_folder to product in order
            order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
            if not order:
                return JSONResponse({'status': 400, 'message': f"Order not found!"})
            for item in order['_original_data']['line_items']:
                if item['line_item_id'] == product_id:
                    item['link_des'] = link_folder

                    # save in COL_REUSE_DES 
                    document_ = {
                        'sku': sku,
                        'link_des': link_folder,
                        'uploaded_by': user_name,
                    }

                    await db_client.insert_one(document_, DB['COL_REUSE_DES'])

                    if item['fulfillment_status'] in ['']:
                        item['fulfillment_status'] = 'Ready'
                        item['assign_designer'] = {
                            'designer': 'reuse_design',
                            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'design_checked': True,
                        }
                    if user_name == 'vmm_team':
                        item['fulfillment_status'] = 'Ready'
                        item['assign_designer'] = {
                            'designer': 'vmm_team',
                            'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'design_checked': True,
                            'checker': '',
                            'total_rejected': 0
                        }

                    
                    if 'user_logs' in item and item['user_logs'] != None:
                        user_logs = item.get('user_logs', [])
                        user_logs.append(f"{user_name} Uploaded design at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                        item['user_logs'] = user_logs
                    else:
                        user_logs = [f"{user_name} Uploaded design at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                        item['user_logs'] = user_logs

            await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])


            return JSONResponse({'message': "Uploaded successfully!", 'link_folder': link_folder})
        except:
            return JSONResponse({'message': f"Uploaded failed! - {traceback.format_exc()}"}, status_code=400)

async def upload_order_design_ff_ngoai(request):
    db_client = request.state.db
    user_name = request.state.user_data['username']
    if request.method == "POST":
        data = await request.form()
        # 
        order_id = data.get('order_id', '')
        product_id = data.get('product_id', '')
        line_item_id = data.get('line_item_id', '')
        shop = data.get('shop', '')
        seller = data.get('seller', '')
        order_date = data.get('order_date', '')
        files = data.getlist('files')
        if order_id == "" or product_id == "":
            return JSONResponse({'status': 400, 'message': f"Missing order_id or product!"})

        # Handle the uploaded files and convert them to Base64
        files = []
        for file_field in data.keys():
            file = data[file_field]
            if hasattr(file, 'filename'):
                content = await file.read()
                file_base64 = base64.b64encode(content).decode('utf-8')
                files.append({
                    'filename': file.filename,
                    'content_base64': file_base64,
                })

        # Perform any other processing you need with the form data and uploaded files
        # ...
        try:
            link_folder = google_storage.upload_to_bucket_no_delete(seller, shop, order_date, order_id, product_id, files)
            # update link_folder to product in order
            order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
            if not order:
                return JSONResponse({'status': 400, 'message': f"Order not found!"})
            for item in order['_original_data']['line_items']:
                if item['line_item_id'] == line_item_id:
                    if  item['link_des'] != '':
                        item['link_des']+= link_folder
                    else:
                        item['link_des'] = link_folder
                    if user_name == 'vmm_team':
                        item['fulfillment_status'] = 'Ready'
                        item['assign_designer'] = {
                            'designer': 'vmm_team',
                            'created_at': convert_to_mongodb_time(time.time()),
                            'design_checked': True,
                            'checker': '',
                            'total_rejected': 0
                        }

                    
                    if 'user_logs' in item and item['user_logs'] != None:
                        user_logs = item.get('user_logs', [])
                        user_logs.append(f"{user_name} Uploaded design ff_ngoai at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                        item['user_logs'] = user_logs
                    else:
                        user_logs = [f"{user_name} Uploaded design ff_ngoai at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                        item['user_logs'] = user_logs

            await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
            return JSONResponse({'message': "Uploaded successfully!", 'link_folder': link_folder})
        except Exception as e:
            print(e)
            return JSONResponse({'message': "Uploaded failed!"}, status_code=400)

async def upload_gen_design_pdf(request):
    if request.method == "POST":
        data = await request.form()
        # 
        file = data.get('file', {})
        # Handle the uploaded files and convert them to Base64
        content = await file.read()
        file_base64 = base64.b64encode(content).decode('utf-8')
        time_stamp = time.time()
        random_number = random.randint(0, 100000)
        files = [{
            'filename':str(random_number)+  str(time_stamp)+'_'+file.filename,
            'content_base64': file_base64,
        }]
        date = datetime.datetime.fromtimestamp(time_stamp).strftime('%Y-%m-%d')
        # Perform any other processing you need with the form data and uploaded files
        # ...
        try:
            link_folder = google_storage.upload_to_bucket_pdf_layer_gen('pdf_gen_img', date,  files)
            # update link_folder to product in order
            return JSONResponse({'message': "Uploaded successfully!", 'link_folder': link_folder})
        except Exception as e:
            print(e)
            return JSONResponse({'message': "Uploaded failed!"}, status_code=400)

async def upload_gen_design_pdf1(request):
    if request.method == "POST":
        data = await request.form()
        # 
        files = data.getlist('file')
        # Handle the uploaded files and convert them to Base64
        
        formfile = []
        for file in files:
            content = await file.read()
            file_base64 = base64.b64encode(content).decode('utf-8')
            formfile.append({
                'filename':file.filename,
                'content_base64': file_base64,
            })
        time_stamp = time.time()
        date = datetime.datetime.fromtimestamp(time_stamp).strftime('%Y-%m-%d')
        # Perform any other processing you need with the form data and uploaded files
        # ...
        try:
            link_folder = google_storage.upload_to_bucket_pdf_layer_gen('pdf_gen_img', date,  formfile)
            # update link_folder to product in order
            return JSONResponse({'message': "Uploaded successfully!", 'link_folder': link_folder})
        except Exception as e:
            print(e)
            return JSONResponse({'message': "Uploaded failed!"}, status_code=400)

async def upload_order_design_ff_ngoai_edited(request):
    db_client = request.state.db
    user_name = request.state.user_data['username']
    if request.method == "POST":
        data = await request.form()
        image_file = data.get('image_file', {})
        seller = data.get('seller', '')
        shop = data.get('shop', '')
        order_date = data.get('order_date', '')
        order_id = data.get('order_id', '')
        product_id = data.get('product_id', '')
        files = []
        files.append(json.loads(image_file))
        link_folder = google_storage.upload_to_bucket_no_delete(seller, shop, order_date, order_id, product_id, files)
        return JSONResponse({'message': "Uploaded successfully!", 'link_folder': link_folder[0]})

async def image_fulfill_edit(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.form()
        img_link = data.get('img_link', '')
        x = data.get('x', '0')
        y = data.get('y', '0')
        width = data.get('width', '0')
        height = data.get('height', '0')
        flipping = data.get('flipping', 'false')
        img_flipping = False
        if flipping == 'true':
            img_flipping = True
        result = await func.eidt_image_fulfill(img_link, float(x),  float(y),  float(width),  float(height), img_flipping)
        return JSONResponse({'message': "Uploaded successfully!", 'base64': result})


async def get_image_info(request):
    db_client = request.state.db
    try:
        if request.method == "POST":
            data = await request.json()
            link = data['link']
            link = link[1:-1]
            url_list = link.split(', ') 
            product_type = data['product_type']
            # try:
            image_info = await func.get_images_size(url_list)
            # find product by product_type
            product = await db_client.find_one({'product_type': product_type}, collection_name=DB['COL_PRODUCTS'], filter_={'product_template_width': 1, 'product_template_height': 1})
            if not product:
                template = {
                    'product_template_width': 0,
                    'product_template_height': 0,
                }
            else:
                template = {
                    'product_template_width': product['product_template_width'],
                    'product_template_height': product['product_template_height'],
                }

            return JSONResponse({'message': "Get image info successfully!", 'images': image_info, 'template': template})
    except:
        print('Get image info failed! - ', traceback.format_exc())
        return JSONResponse({'message': "Get image info failed!"}, status_code=400)
        

async def get_link_des(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            data = await request.json()
            sku = data['sku']
            query = {
                'sku': sku,
            }
            link_des = await db_client.find_one(query, collection_name=DB['COL_REUSE_DES'], filter_={'link_des': 1})
            print('get_link_des: ', link_des['link_des'])
            if not link_des:
                return JSONResponse({'status': 400, 'message': "Get link des failed!", 'link_des': None})
            return JSONResponse({'status': 200, 'message': "Get link des successfully!", 'link_des': link_des['link_des']})
        except:
            return JSONResponse({'status': 400, 'message': f"Get link des failed! - {traceback.format_exc()}"}, status_code=400)

async def search_designer(request):
    db_client = request.state.db
    if request.method == "POST":
        query = {
            'groups': {'$in': ['designer', 'leader designer']},
        }
        try:
            users = await db_client.find_many(query, collection_name=DB['COL_USERS'], filter_={'username':1})
            users = [user['username'] for user in users]
            return JSONResponse({'message': "Get designer successfully!", 'designers': users})
        except:
            return JSONResponse({'message': "Get designer failed!"}, status_code=400)
        
async def assign_designer(request):
    user_name = request.state.user_data['username']
    try:
        db_client = request.state.db
        if request.method != "POST":
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        selected_values = data.get('selectedValues', [])
        designer = data.get('designer')

        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]
        product_ids = [str(selected_value['productId']) for selected_value in selected_values]
        line_item_ids = [str(selected_value['lineItemId']) for selected_value in selected_values]

        # Find the orders to be updated
        orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})

        for order in orders:
            for product in order['_original_data'].get('line_items', []):
                if str(product['product_id']) in product_ids and str(product['line_item_id']) in line_item_ids:
                    product['assign_designer'] = {
                        'designer': designer,
                        'created_at': convert_to_mongodb_time(time.time()),
                        'design_checked': False,
                        'checker': '',
                        'total_rejected': 0
                    }
                    product['fulfillment_status'] = 'Designing'
                    
                    link_des = product.get('link_des', [])
                    if 'link_des' in product:
                        product['link_des'] = link_des

                    if 'user_logs' in product and product['user_logs'] != None:
                        user_logs = product.get('user_logs', [])
                        user_logs.append(f"{user_name} - Assign design for {designer} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                        product['user_logs'] = user_logs
                    else:
                        user_logs = [f"{user_name} - Assign design for {designer} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                        product['user_logs'] = user_logs

            await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

        return JSONResponse({'status': 200, 'message': 'Assign designer successfully'})

    except Exception as e:
        return JSONResponse({'status': 400, 'message': f"Assign designer failed: {str(e)}"})

async def approve_designer(request):
    try:
        db_client = request.state.db
        user_name = request.state.user_data['username']
        if request.method != "POST":
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        product_id = data.get('product_id')
        order_id = data.get('order_id')
        line_item_id = data.get('line_item_id')

        # Find the order to be updated
        order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
        if not order:
            return JSONResponse({'status': 400, 'message': 'Order not found'})

        for product in order['_original_data'].get('line_items', []):
            if str(product['product_id']) == str(product_id) and str(product['line_item_id']) == str(line_item_id):
                assign_designer = product.get('assign_designer', {})
                assign_designer['designer'] = assign_designer.get('designer', '')
                assign_designer['finished_at'] = convert_to_mongodb_time(time.time())
                assign_designer['design_checked'] = True
                assign_designer['checker'] = user_name
                assign_designer['total_rejected'] = assign_designer.get('total_rejected', 0)

                product['assign_designer'] = assign_designer

                product['fulfillment_status'] = 'Ready'

                if 'user_logs' in product and product['user_logs'] != None:
                    user_logs = product.get('user_logs', [])
                    user_logs.append(f"{user_name} - Approve design at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    product['user_logs'] = user_logs
                else:
                    user_logs = [f"{user_name} - Approve design at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                    product['user_logs'] = user_logs

                await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
        return JSONResponse({'status': 200, 'message': 'Approve design successfully'})

    except Exception as e:
        print(f"Approve design failed: {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"Approve design failed: {traceback.format_exc()}"})

async def reject_designer(request):
    try:
        db_client = request.state.db
        user_name = request.state.user_data['username']

        if request.method != "POST":
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        product_id = data.get('product_id')
        order_id = data.get('order_id')
        line_item_id = data.get('line_item_id')
        note = data.get('note')

        # Find the order to be updated
        order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
        if not order:
            return JSONResponse({'status': 400, 'message': 'Order not found'})

        updated = False
        for product in order['_original_data']['line_items']:
            if str(product['product_id']) == str(product_id) and str(product['line_item_id']) == str(line_item_id):
                assign_designer = product.get('assign_designer', {})
                assign_designer['designer'] = assign_designer.get('designer', '')
                assign_designer['finished_at'] = convert_to_mongodb_time(time.time())
                assign_designer['design_checked'] = False
                assign_designer['checker'] = user_name
                assign_designer['total_rejected'] = assign_designer.get('total_rejected', 0) + 1


                product['assign_designer'] = assign_designer

                product['fulfillment_status'] = 'Rejected'
                product['design_note'] = note

                if 'user_logs' in product and product['user_logs'] != None:
                    user_logs = product.get('user_logs', [])
                    user_logs.append(f"Rejected design - {note} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    product['user_logs'] = user_logs
                else:
                    user_logs = [f"Rejected design  - {note} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                    product['user_logs'] = user_logs

                await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
                updated = True
                return JSONResponse({'status': 200, 'message': 'Rejected designer successfully'})

        if not updated:
            return JSONResponse({'status': 400, 'message': 'Product not found'})

    except Exception as e:
        print(f"Rejected designer failed: {traceback.format_exc()}")
        return JSONResponse({'status': 400, 'message': f"Rejected designer failed: {traceback.format_exc()}"})

async def bulk_edit(request):
    db_client = request.state.db
    user = request.state.user_data['username']
    if request.method == "POST":
        data = await request.json()
        selected_values = data.get('selectedValues', [])
        try: status = data.get('navs-top-status')
        except: status = None
        try: factory = data.get('navs-top-xuong')
        except: factory = None
        try: product_type = data.get('navs-top-type')
        except: product_type = None

        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]
        product_ids = [str(selected_value['productId']) for selected_value in selected_values]

        # Find the orders to be updated
        orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1, '_team': 1})
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        

        try:
            for order in orders:
                for product in order['_original_data'].get('line_items', []):
                    if str(product['product_id']) in product_ids:
                        if status != None:
                            product['fulfillment_status'] = status
                            if 'user_logs' in product and product['user_logs'] != None:
                                user_logs = product.get('user_logs', [])
                                user_logs.append(f"{user} - Bulk edit status to {status} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                                product['user_logs'] = user_logs
                                
                            else:
                                user_logs = [f"{user} - Bulk edit status to {status} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                                product['user_logs'] = user_logs
                            if status in ['Ready'] and product['link_des'] == '':
                                if order['_team'] != 'nb_team':
                                    pass
                                else:
                                    try:
                                        # find link des by sku in COL_REUSE_DES
                                        query = {
                                            'sku': product['sku'],
                                        }
                                        reuse_des = await db_client.find_one(query, collection_name=DB['COL_REUSE_DES'])
                                        if reuse_des is not None:
                                            product['link_des'] = reuse_des['link_des']
                                            product['assign_designer'] = {
                                                'designer': 'reuse_design',
                                                'design_checked': True,
                                                'created_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                            }
                                    except:
                                        pass
                        if factory != None:
                            product['factory'] = factory
                            if 'user_logs' in product and product['user_logs'] != None:
                                user_logs = product.get('user_logs', [])
                                user_logs.append(f"{user} - Bulk edit factory to {factory} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                                product['user_logs'] = user_logs
                            else:
                                user_logs = [f"{user} - Bulk edit factory to {factory} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                                product['user_logs'] = user_logs
                        if product_type != None:
                            product['product_type'] = product_type
                            if 'user_logs' in product and product['user_logs'] != None:
                                user_logs = product.get('user_logs', [])
                                user_logs.append(f"{user} - Bulk edit product type to {type} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                                product['user_logs'] = user_logs
                            else:
                                user_logs = [f"{user} - Bulk edit product type to {type} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                                product['user_logs'] = user_logs
                await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

            return JSONResponse({'status': 200, 'message': 'Bulk edit successfully'})
        except:
            return JSONResponse({'status': 400, 'message': f'Bulk edit failed - {traceback.format_exc()}'})
        

@authenticate
async def approve_invoice(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            data = await request.json()
            invoice_id = data.get('id')
            invoice = await db_client.find_one({'_id': ObjectId(invoice_id)}, collection_name=DB['COL_INVOICE'])
            if not invoice:
                return JSONResponse({'status': 400, 'message': 'Invoice not found'})
            invoice['status'] = 'paid'
            await db_client.update_one(invoice['_id'], invoice, collection_name=DB['COL_INVOICE'])

            # update team balance
            team = await db_client.find_one({'name': invoice['team']}, collection_name=DB['COL_TEAMS'], filter_={'balance': 1})
            if not team:
                return JSONResponse({'status': 400, 'message': 'Team not found'})
            
            # get rate
            if invoice['currency'] == 'VND':
                rate = await db_client.find_one({'currency': 'VND'}, collection_name=DB['COL_RATE'], filter_={'rate': 1})
                if not rate:
                    return JSONResponse({'status': 400, 'message': 'Rate not found'})
                else:
                    current_balance = team['balance'] + invoice['total'] / rate['rate']
            else:
                current_balance = team['balance'] + invoice['total']
            team['balance'] = float(round(current_balance, 2))

            await db_client.update_one(team['_id'], team, collection_name=DB['COL_TEAMS'])

            return JSONResponse({'status': 200, 'message': 'Approve invoice successfully'})
        except:
            return JSONResponse({'status': 400, 'message': f'Approve invoice failed - {traceback.format_exc()}'})
@authenticate
async def reject_invoice(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            data = await request.json()
            invoice_id = data.get('id')
            invoice = await db_client.find_one({'_id': ObjectId(invoice_id)}, collection_name=DB['COL_INVOICE'])
            if not invoice:
                return JSONResponse({'status': 400, 'message': 'Invoice not found'})
            invoice['status'] = 'rejected'
            await db_client.update_one(invoice['_id'], invoice, collection_name=DB['COL_INVOICE'])

            return JSONResponse({'status': 200, 'message': 'Reject invoice successfully'})
        except:
            return JSONResponse({'status': 400, 'message': f'Reject invoice failed - {traceback.format_exc()}'})
        
@authenticate
async def replace_items(request):
    db_client = request.state.db
    user = request.state.user_data['username']
    try:
        if request.method == "POST":
            data = await request.json()
            selected_values = data.get('selectedValues', [])
            reason = data.get('reason')

            order_ids = [str(selected_value['orderId']) for selected_value in selected_values]
            line_item_ids = [str(selected_value['lineItemId']) for selected_value in selected_values]

            # Find the orders to be updated
            orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_=None)

            if not orders:
                return JSONResponse({'status': 400, 'message': 'Orders not found'})
            
            try:
                # add new item to order 
                replace_orders = []
                for order in orders:
                    new_order = copy.deepcopy(order)
                    new_order.pop('_id', None)
                    new_order['_original_data']['order_id'] = 'RE-' + new_order['_original_data']['order_id']
                    new_order['_original_data']['line_items'] = []

                    for product in order['_original_data'].get('line_items', []):
                        if str(product['line_item_id']) in line_item_ids:
                            line_item = copy.deepcopy(product)
                            user_logs = [f"{user} Create new replace line item at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
                            line_item['user_logs'] = user_logs
                            line_item['fulfillment_order_id'] = 'RE-' + line_item['fulfillment_order_id']
                            line_item['line_item_id'] = line_item['fulfillment_order_id'] + '.' + str((len(new_order['_original_data']['line_items']) + 1))
                            line_item['fulfillment_status'] = ''
                            line_item['is_replaced'] = True
                            line_item['reason'] = reason
                            line_item['base_cost'] = 0
                            line_item['link_des'] = ''
                            line_item['factory'] = ''
                            try:
                                line_item['assign_designer']['checker'] = ''
                                line_item['assign_designer']['design_checked'] = False
                                line_item['assign_designer']['created_at'] = datetime.datetime.now()
                            except: pass
                            line_item.pop('ship_cost', None)

                            # add new line item to order
                            new_order['_original_data']['line_items'].append(line_item)
                    new_order.pop('shipping', None)

                    # insert new order
                    # remove _id
                    new_order['create_at'] = convert_to_mongodb_time(time.time())
                    new_order['_update_time'] = convert_to_mongodb_time(time.time())

                    replace_orders.append(new_order)

                await db_client.insert_many(documents=replace_orders, collection_name=DB['COL_FULFILLMENTS'])

                return JSONResponse({'status': 200, 'message': 'Replace items successfully'})
            except:
                print(f'Replace items failed - {traceback.format_exc()}')
                return JSONResponse({'status': 400, 'message': 'Replace items failed - ' + traceback.format_exc()})
    except:
        return JSONResponse({'status': 400, 'message': 'Replace items failed - ' + traceback.format_exc()})
    
        
async def kiotviet(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        selected_values = data.get('selectedValues', [])
        
        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]
        product_ids = [str(selected_value['productId']) for selected_value in selected_values]

        filter_ = {
            '_original_data.delivery_info': 1,
            '_original_data.order_id': 1,
            '_original_data.seller_id': 1,
        }

        # find orders
        orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})

        # create DataFrame
        df_data = []
        for order in orders:
            df_data.append({
                'Loại hàng': 'Hàng hóa',
                'Nhóm hàng(3 Cấp)': 'Phụ kiện Nam',
                'Mã hàng': order['_original_data']['order_id'],
                'Mã vạch': '',  # You can fill this with appropriate data
                'Tên hàng': order['_original_data']['seller_id'],
                'Thương hiệu': '',
                'Giá bán': 0,
                'Giá vốn': 0,
                'Tồn kho': 0,
                'KH đặt': 0,
                'Dự kiến hết hàng': '0 Ngày',
                'Tồn nhỏ nhất': 0,
                'Tồn lớn nhất': 999,
                'ĐVT': '',
                'Mã ĐVT Cơ bản': '',
                'Quy đổi': 1,
                'Thuộc tính': f"Tên: {order['_original_data']['delivery_info']['name']}",
                'Mã HH Liên quan': '',
                'Hình ảnh (url1,url2...)': '',
                'Trọng lượng': 0,
                'Đang kinh doanh': 1,
                'Được bán trực tiếp': 1,
                'Mô tả': 'Tên: ',
                'Mẫu ghi chú': order['_original_data']['delivery_info']['name'],
                'Vị trí': '',
                'Hàng thành phần': ''
            })
        df = pd.DataFrame(df_data)

        # Tạo tệp Excel bằng openpyxl trực tiếp
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active

        # Thêm dữ liệu vào tệp Excel
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Lưu tệp Excel vào BytesIO
        workbook.save(output)

        # in ra tệp Excel
        output.seek(0)

        excel_filename = "kiotviet.xlsx"
        response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
        
        return response

    except:
        return JSONResponse({'status': 400, 'message': 'Export failed - ' + traceback.format_exc()})


async def export_xuong(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        selected_values = data.get('selectedValues', [])
        
        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]
        product_ids = [str(selected_value['productId']) for selected_value in selected_values]

        filter_ = {
            'create_at': 1,
            '_original_data.line_items.fulfillment_order_id': 1,
            '_original_data.line_items.line_item_id': 1,
            '_original_data.line_items.product_type': 1,
            '_original_data.line_items.size': 1,
            '_original_data.line_items.color': 1,
            '_original_data.line_items.material': 1,
            '_original_data.line_items.personalization': 1,
            '_original_data.line_items.quantity': 1,
            '_original_data.line_items.num_holder': 1,
            '_original_data.line_items.note': 1,
            '_original_data.line_items.thumbnail': 1,
            '_original_data.line_items.link_des': 1,    
            '_original_data.line_items.other_option': 1,    
        }

        # find orders
        orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})

        df_data = []
        for order in orders:
            for item in order['_original_data']['line_items']:
                mockup = item['thumbnail']
                try:
                    if type(item['link_des']) is list:
                        link_des = ""
                        for i, link in enumerate(item['link_des']):
                            link_des += link
                            if i < len(item['link_des']) - 1:
                                link_des += '\r\n'

                    else:
                        link_des = item['link_des']
                except:
                    link_des = item['link_des']
                other_option = item.get('other_option', '').lower()
                pack = func.get_pack_in_item(other_option)
                
                quantity = int(item['quantity']) * pack
                
                if len(order['_original_data']['line_items']) > 1:
                    line_item_id = item['line_item_id'].replace('.', '-')
                else:
                    line_item_id = item['fulfillment_order_id']
            
                df_data.append({
                    'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                    'Etsy Order ID': line_item_id,
                    'Product type': item['product_type'],
                    'Size': item['size'],
                    'Color/ Material': f"{item['color']} / {item['material']}",
                    'Chữ trên charm': item['personalization'],
                    'Quantity': quantity,
                    'Số lượng charm': '',
                    'Number of Holder': item['num_holder'],
                    'Note': item['note'],
                    'FF status': 'Ready',
                    'Mockup': str(mockup),
                    'Design': str(link_des),
                })
        df = pd.DataFrame(df_data)

        # Tạo tệp Excel bằng openpyxl trực tiếp
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active

        # Thêm dữ liệu vào tệp Excel
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Lưu tệp Excel vào BytesIO
        workbook.save(output)

        # in ra tệp Excel
        output.seek(0)

        excel_filename = "export_xuong.xlsx"
        response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers['Content-Disposition'] = f'attachment; filename="{excel_filename}"'

        # Trả về response chứa tệp Excel
        return response

    except:
        return JSONResponse({'status': 400, 'message': 'Export failed - ' + traceback.format_exc()})
    

async def export_ornc(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        date = data['date']
        start_date = convert_to_mongodb_time(date.split(' - ')[0], 'start')
        end_date = convert_to_mongodb_time(date.split(' - ')[1], 'end')
        
        created_at = datetime.datetime.now()
        doc = {
            'query': {
                'create_at': {"$gte": start_date, "$lte": end_date}, 
                '_original_data.line_items.fulfillment_status': 'Ready',
                '_original_data.line_items.factory' :'',
                '_original_data.line_items.product_type': {'$in': ['ORNCS', 'ORNCM']},
            },
            'status': 'pending',
            'created_at': created_at,
        }
        # insert doc to db collection export_ornc
        await db_client.insert_one(doc, collection_name=DB['COL_EXPORT_ORNC'])
        return JSONResponse({'status': 200, 'message': 'Export ORNC successfully'})
        # return response

    except:
        return JSONResponse({'status': 400, 'message': 'Export failed - ' + traceback.format_exc()})

async def export_ship(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        selected_values = data.get('selectedValues', [])
        
        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]
        product_ids = [str(selected_value['productId']) for selected_value in selected_values]

        filter_ = {
            'create_at': 1,
            '_original_data.line_items.fulfillment_order_id': 1,
            '_original_data.line_items.product_type': 1,
            '_original_data.line_items.note': 1,
            '_original_data.line_items.quantity': 1,
            '_original_data.line_items.other_option': 1,
            '_original_data.seller_id': 1,
            '_original_data.delivery_info.name': 1,
            '_original_data.delivery_info.address1': 1,
            '_original_data.delivery_info.address2': 1,
            '_original_data.delivery_info.city': 1,
            '_original_data.delivery_info.state': 1,
            '_original_data.delivery_info.zip': 1,
            '_original_data.delivery_info.country': 1,
        }

        # find orders
        orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        # create DataFrame
        df_data = []
        for order in orders:
            for item in order['_original_data']['line_items']:
                other_option = item.get('other_option', '').lower()
                
                pack = func.get_pack_in_item(other_option)
                quantity = int(item['quantity']) * pack
                df_data.append({
                    'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                    'Order ID': item['fulfillment_order_id'],
                    'Seller ID': order['_original_data']['seller_id'],
                    'Product type': item['product_type'],
                    'Quantity': quantity,
                    'Name': order['_original_data']['delivery_info']['name'],
                    'Address 1': order['_original_data']['delivery_info']['address1'],
                    'Address 2': order['_original_data']['delivery_info']['address2'],
                    'City': order['_original_data']['delivery_info']['city'],
                    'State': order['_original_data']['delivery_info']['state'],
                    'Zip': order['_original_data']['delivery_info']['zip'],
                    'Country': order['_original_data']['delivery_info']['country'],
                    'Note': "Ship nhanh" if 'ship nhanh' in item['note'].lower() else '',
                })
        df = pd.DataFrame(df_data)

        # Tạo tệp Excel bằng openpyxl trực tiếp
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active

        # Thêm dữ liệu vào tệp Excel
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Lưu tệp Excel vào BytesIO
        workbook.save(output)

        # in ra tệp Excel
        output.seek(0)

        excel_filename = "export_ship.xlsx"
        response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
        
        return response

    except:
        return JSONResponse({'status': 400, 'message': 'Export failed - ' + traceback.format_exc()})

async def export_flast_ship(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        selected_values = data.get('selectedValues', [])
        
        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]

        filter_ = {
            '_original_data.line_items.fulfillment_order_id': 1,
            '_original_data.line_items.line_item_id': 1,
            '_original_data.line_items.product_type': 1,
            '_original_data.line_items.quantity': 1,
            '_original_data.line_items.other_option': 1,
            '_original_data.line_items.link_des': 1,
            '_original_data.line_items.thumbnail': 1,
            '_original_data.delivery_info.name': 1,
            '_original_data.delivery_info.address1': 1,
            '_original_data.delivery_info.address2': 1,
            '_original_data.delivery_info.city': 1,
            '_original_data.delivery_info.state': 1,
            '_original_data.delivery_info.zip': 1,
            '_original_data.delivery_info.country': 1,
            '_original_data.delivery_info.phone': 1,
        }

        # find orders
        orders = await db_client.find_many({'_original_data.order_id': {'$in': order_ids}}, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        # create DataFrame
        df_data = []
        for order in orders:
            for item in order['_original_data']['line_items']:
                other_option = item.get('other_option', '').lower()
                
                pack = func.get_pack_in_item(other_option)
                quantity = int(item['quantity']) * pack
                link_des = item['link_des']
                if type(link_des) is list:
                    link_des = link_des[0]
                if len(order['_original_data']['line_items']) > 1:
                    order_id = item['line_item_id']
                else:
                    order_id = item['fulfillment_order_id']
                if item['product_type'] == 'ORNCM':
                    variant_id = 10005
                elif item['product_type'] == 'ORNCS':
                    variant_id = 10006
                df_data.append({
                    'External ID': 'POD092',
                    'Order ID': order_id,
                    'Shipping method': 1,
                    'First name': order['_original_data']['delivery_info']['name'].split(' ')[0],
                    'Last name': order['_original_data']['delivery_info']['name'].split(' ')[-1],
                    'Email': '',
                    'Phone': order['_original_data']['delivery_info']['phone'],
                    'Country': order['_original_data']['delivery_info']['country'],
                    'Region': order['_original_data']['delivery_info']['state'],
                    'Address line 1': order['_original_data']['delivery_info']['address1'],
                    'Address line 2': order['_original_data']['delivery_info']['address2'],
                    'City': order['_original_data']['delivery_info']['city'],
                    'Zip': order['_original_data']['delivery_info']['zip'],
                    'quantity': quantity,
                    'Variant ID': variant_id,
                    'side1': link_des,
                    'side2': '',
                    'Mockup side 1': item['thumbnail'],
                    'Mockup side 2': '',
                })
        df = pd.DataFrame(df_data)

        # Tạo tệp CSV bằng pandas
        output = BytesIO()
        df.to_csv(output, index=False, quoting=csv.QUOTE_NONNUMERIC, encoding='utf-8')

        # in ra tệp CSV
        output.seek(0)

        csv_filename = "export_flast_ship.csv"
        response = StreamingResponse(output, media_type="text/csv")
        response.headers['Content-Disposition'] = f'attachment; filename="{csv_filename}"'

        return response

    except:
        return JSONResponse({'status': 400, 'message': 'Export failed - ' + traceback.format_exc()})

async def sync_tracking(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        selected_values = data.get('selectedValues', [])
        
        order_ids = [str(selected_value['orderId']) for selected_value in selected_values]

        query = {
            '_original_data.order_id': {'$in': order_ids},
            '_original_data.shipping': {'$exists': True},
            '_original_data.shipping.synced': False
        }

        filter_ = {
            '_original_data.shipping': 1,
        }

        # find orders
        orders = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        
        try:
            for order in orders:
                shipping = order['_original_data']['shipping']
                if shipping:
                    # update synced to waiting 
                    shipping['synced'] = 'waiting'
                    await db_client.update_one(order['_id'], {'_original_data.shipping': shipping}, collection_name=DB['COL_FULFILLMENTS'])
        except:
            return JSONResponse({'status': 400, 'message': 'Sync tracking failed - ' + traceback.format_exc()})
        
        return JSONResponse({'status': 200, 'message': 'Sync tracking successfully'})
    
    except:
        return JSONResponse({'status': 400, 'message': 'Sync tracking failed - ' + traceback.format_exc()})



async def get_ff_ngoai_product(request):
    if request.method == "POST":
        data = await request.json()
        products = await fetch_ff_ngoai_products(request, data)
        html = render_ff_ngoai_product(request, {'products': products})
        return HTMLResponse(html)

async def fetch_ff_ngoai_products(request, data: dict) -> list:
    db_client = request.state.db
    mongo_id = data.get('mongo_id')
    print_provider = await db_client.find_one(
        {'_id': ObjectId(mongo_id)},
        collection_name=DB['COL_FULFILL_NGOAI'],
        filter_={'products.id': 1, 'products.name': 1, 'products.brand': 1}
    )
    return print_provider.get('products', [])

async def get_ff_ngoai_variants(request):
    if request.method == "POST":
        data = await request.json()
        variants = await fetch_ff_ngoai_variants(request, data)
        return JSONResponse({'status': 200, 'message': 'Get variants successfully', 'data': variants})

async def fetch_ff_ngoai_variants(request, data: dict) -> dict:
    db_client = request.state.db
    mongo_id = data.get('mongo_id')
    product_id = data.get('product_id')
    print_provider = await db_client.find_one(
        {'_id': ObjectId(mongo_id)},
        collection_name=DB['COL_FULFILL_NGOAI'],
        filter_={'products.id': 1, 'products.title': 1, 'products.brand': 1, 'products.sizes': 1, 'products.colors': 1, 'products.placeholder': 1}
    )
    products = print_provider.get('products', [])
    for product in products:
       
        if int(product['id']) == int(product_id):
            sizes = product.get('sizes', [])
            colors = product.get('colors', [])
            placeholder = product.get('placeholder', [])
            return {'sizes': sizes, 'colors': colors, 'placeholder': placeholder}

async def get_ff_ngoai_sizes(request):
    if request.method == "POST":
        data = await request.json()
        sizes, placeholders = await fetch_ff_ngoai_sizes(data)
        html = render_ff_ngoai_sizes(request, {'sizes': sizes})
        data = {'html': html, 'placeholders': placeholders}
        return JSONResponse({'status': 200, 'message': 'Get sizes successfully', 'data': data})

async def fetch_ff_ngoai_sizes(data: dict) -> tuple:
    sizes = data.get('sizes', [])
    placeholders = data.get('placeholders', [])
    return sizes, placeholders

async def get_ff_ngoai_colors(request):
    if request.method == "POST":
        data = await request.json()
        colors = data.get('colors', [])
        html = render_ff_ngoai_colors(request, {'colors': colors})
        return HTMLResponse(html)

async def fulfill_pritify(request):
    db_client = request.state.db
    if request.method == "POST":
        
        data = await request.json()
        order_id = data.get('order_id','')
        order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={})
        order_ff_ngoai = await func.update_order_ff_ngoai(order,'printify')
        address_order = order['_original_data']['delivery_info']
        fast_shipping = data.get('fast_shipping',False)
        address_to = {
            "first_name" : address_order['name'].split(' ')[0],
            "last_name" : address_order['name'].split(' ')[-1],
            "address1" : address_order['address1'],
            "address2" : address_order['address2'],
            "region": "",
            "city" : address_order['city'],
            "zip" : address_order['zip'],
            "country" : address_order['country'],
        }
        print_provider_mongo_id = data.get('print_provider_mongo_id','')
        line_items = data.get('line_items',[])
        print_provider = await db_client.find_one({'_id': ObjectId(print_provider_mongo_id)}, collection_name=DB['COL_FULFILL_NGOAI'], filter_={'id':1,'products.id': 1,'products.variants': 1})
        for line_item in line_items:
            product_id = line_item.get('blueprint_id','')
            color = line_item.get('color','')
            size = line_item.get('size','')
            for product in print_provider['products']:
                if product['id'] == int(product_id):
                    for variant in product['variants']:
                        if variant['options']['color'] == color and variant['options']['size'] == size:
                            line_item['variant_id'] = variant['id']
                            break
                    break
            line_item['print_provider_id'] = print_provider['id']
            line_item.pop('color')
            line_item.pop('size')
        printify_order = await func.create_product_with_order(order_id,address_to,line_items,fast_shipping)
        if printify_order.get('id',False):
            result = {
                'id_printify':printify_order['id'],
                'order_id':order_id,
                'platform_ff': 'printify',
                'line_items':line_items,
                'status':'pending',
            }
            await db_client.insert_one(result, collection_name=DB['COL_FF_NGOAI'])
            await db_client.update_one(order['_id'], order_ff_ngoai, collection_name=DB['COL_FULFILLMENTS'])
            return JSONResponse({'status': 200, 'message': 'Fulfill successfully'})
        else:
            return JSONResponse({'status': 400, 'message': printify_order["errors"]['reason']})

async def fulfill_merchize(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        order_id = data.get('order_id','')
        order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={})
        # order_ff_ngoai = await func.update_order_ff_ngoai(order,'printify')
        address_order = order['_original_data']['delivery_info']
        ship_to = {
            "full_name" : address_order['name'],
            "address_1" : address_order['address1'],
            "address_2" : address_order['address2'],
            "state": address_order['state'],
            "city" : address_order['city'],
            "postcode" : address_order['zip'],
            "country" : address_order['country'],
            "email": "",
            "phone": ""

        }
        line_items = data.get('line_items',[])
        order_mechize = {
            "order_id":"esty "+ order_id,
            "identifier": "esty "+ order_id,
            "shipping_info": ship_to,
            "tax": "",
            "tags": [],
            "items": line_items,

        }
        data_merchize = await func.post_merchize_ff_ngoai(order_mechize)
        if (data_merchize.get('success', False)):
            # update status order
            for item in order['_original_data']['line_items']:
                item['fulfillment_status'] = 'Done recently'
                item['ff_provider_status'] = 'Fulfilled'
                item['provider'] = 'merchize'
                item['ff_provider_date'] = convert_to_mongodb_time(datetime.datetime.now())
                item['user_logs'].append(f"{request.state.user_data['username']} fullfiled by merchize at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

            return JSONResponse({'status': 200, 'message': 'Fulfill successfully'})
        else:
            return JSONResponse({'status': 400, 'message': data_merchize['message']})


async def fulfill_gearment(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        order_id = data.get('order_id','')
        order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={})
        # order_ff_ngoai = await func.update_order_ff_ngoai(order,'printify')
        line_items = data.get('line_items',[])
        for original_item in order['_original_data']['line_items']:
            for item in line_items:
                if original_item['line_item_id'] == item['line_item_id']:
                    original_item['design_link'] = item['design_link']
                    original_item['size'] = item['size']
                    original_item['color'] = item['color']
                    original_item['gearment_product_id'] = item['gearment_product_id']
                    original_item['product_name'] = item['product_name']
                    if 'design_link_back' in item:
                        original_item['design_link_back'] = item['design_link_back']
                    else:
                        original_item['design_link_back'] = ''    
                    original_item['quantity'] = item['quantity']
                    break
                    
        data_gearment = await func.post_fulfill_gearment(order)
        if (data_gearment.get('status') == 'success'):
                # update status order
                for item in order['_original_data']['line_items']:
                    item['fulfillment_status'] = 'Done recently'
                    item['ff_provider_status'] = 'Fulfilled'
                    item['provider'] = 'gearment'
                    item['ff_provider_date'] = convert_to_mongodb_time(datetime.datetime.now())
                    item['user_logs'].append(f"{request.state.user_data['username']} fullfiled by gearment at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

                return JSONResponse({'status': 200, 'message': 'Fulfill successfully'})
        else:
            return JSONResponse({'status': 400, 'message':'Gearment: ' +  data_gearment['message']})



async def import_trackings(request):
    # import data from csv file
    db_client = request.state.db
    try:
        if request.method == "POST":
            form = await request.form()
            file = form["file"]
            content_type = file.headers["content-type"]

            user = request.state.user_data
            # team_name = await db_client.find_one({"username": {"$in": [user['username']]}}, collection_name=DB['COL_TEAMS'])
            # if file is csv 

            if content_type == "text/csv":
                df = pd.read_csv(file.file)
            # if file is xlsx
            elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                df = pd.read_excel(file.file)
            else:
                return JSONResponse({"status": "error", "message": "File type not supported!"})
            df_json_string = df.to_json(orient='records')
            df_json = json.loads(df_json_string)

            # if df is empty
            if df.empty:
                return JSONResponse({"status": "error", "message": "File is empty!"})
            count = 0
            try:
                order_formated = func.format_tracking(df_json)
                insert_ = await asyncio.gather(*[func.insert_tracking(order, user, db_client) for order in order_formated])            
                return JSONResponse({"status": "success", "message": f"Import tracking successfully!"})
            except:
                return JSONResponse({"status": "error", "message": f"Import tracking failed: {traceback.format_exc()}"})
    except:
        return JSONResponse({"status": "error", "message": f"Import tracking failed: {traceback.format_exc()}"})
async def set_statistics(request):
    db_client = request.state.db
    try:
        if request.method == "POST":
            status = ['Ready','Confirming','Pending','Doing','Sai design','Design uploaded','In','Out','Half Done','Designing', '']

            # get all orders from db have status in status
            orders = await db_client.find_many_combo({'_original_data.line_items.fulfillment_status': {'$in': status}}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data.line_items.fulfillment_status': 1, '_original_data.shop_id': 1, '_team': 1})
            
            query = {
                '_original_data.line_items.fulfillment_status': 'Ready',
                '_original_data.line_items.assign_designer': {'$exists': True},
                '_original_data.line_items.assign_designer.created_at': {'$lte': datetime.datetime.now() - datetime.timedelta(days=7)},
            }
            orders_7 = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data.line_items.fulfillment_status': 1, '_original_data.shop_id': 1, '_team': 1})

            items = [item for order in orders for item in order['_original_data']['line_items']]

            nb_items = [item for order in orders for item in order['_original_data']['line_items'] if order['_team'] == 'nb_team']

            shops = await db_client.find_many({}, collection_name=DB['COL_SHOP'], filter_={'shop': 1, 'team': 1})
            # 
            team_tung = [shop['shop'] for shop in shops if shop['team'] == 'Tùng']
            team_vi = [shop['shop'] for shop in shops if shop['team'] == 'Vi']
            team_trung = [shop['shop'] for shop in shops if shop['team'] == 'Trung']
            team_thuan = [shop['shop'] for shop in shops if shop['team'] == 'Thuận']

            # get orders have status in status and team in each team
            items_trung = [item for order in orders for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_trung]
            items_vi = [item for order in orders for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_vi]
            items_tung = [item for order in orders for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_tung]
            items_thuan = [item for order in orders for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_thuan]
            items_7_trung = [item for order in orders_7 for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_trung]
            items_7_vi = [item for order in orders_7 for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_vi]
            items_7_tung = [item for order in orders_7 for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_tung]
            items_7_thuan = [item for order in orders_7 for item in order['_original_data']['line_items'] if order['_original_data']['shop_id'] in team_thuan]


            # get items have status in status and team in each team
            ready_trung = [item for item in items_trung if item['fulfillment_status'] == 'Ready']
            confirming_trung = [item for item in items_trung if item['fulfillment_status'] == 'Confirming']
            pending_trung = [item for item in items_trung if item['fulfillment_status'] == 'Pending']
            doing_trung = [item for item in items_trung if item['fulfillment_status'] == 'Doing']
            sai_design_trung = [item for item in items_trung if item['fulfillment_status'] == 'Sai design']
            design_uploaded_trung = [item for item in items_trung if item['fulfillment_status'] == 'Design uploaded']
            in_trung = [item for item in items_trung if item['fulfillment_status'] == 'In']
            out_trung = [item for item in items_trung if item['fulfillment_status'] == 'Out']
            half_done_trung = [item for item in items_trung if item['fulfillment_status'] == 'Half Done']
            designing_trung = [item for item in items_trung if item['fulfillment_status'] == 'Designing']
            ready_7_trung = [item for item in items_7_trung if item['fulfillment_status'] == 'Ready']
            none_trung = [item for item in items_trung if item['fulfillment_status'] == '']
            ready_tung = [item for item in items_tung if item['fulfillment_status'] == 'Ready']
            confirming_tung = [item for item in items_tung if item['fulfillment_status'] == 'Confirming']
            pending_tung = [item for item in items_tung if item['fulfillment_status'] == 'Pending']
            doing_tung = [item for item in items_tung if item['fulfillment_status'] == 'Doing']
            sai_design_tung = [item for item in items_tung if item['fulfillment_status'] == 'Sai design']
            design_uploaded_tung = [item for item in items_tung if item['fulfillment_status'] == 'Design uploaded']
            in_tung = [item for item in items_tung if item['fulfillment_status'] == 'In']
            out_tung = [item for item in items_tung if item['fulfillment_status'] == 'Out']
            half_done_tung = [item for item in items_tung if item['fulfillment_status'] == 'Half Done']
            designing_tung = [item for item in items_tung if item['fulfillment_status'] == 'Designing']
            ready_7_tung = [item for item in items_7_tung if item['fulfillment_status'] == 'Ready']
            none_tung = [item for item in items_tung if item['fulfillment_status'] == '']
            ready_vi = [item for item in items_vi if item['fulfillment_status'] == 'Ready']
            confirming_vi = [item for item in items_vi if item['fulfillment_status'] == 'Confirming']
            pending_vi = [item for item in items_vi if item['fulfillment_status'] == 'Pending']
            doing_vi = [item for item in items_vi if item['fulfillment_status'] == 'Doing']
            sai_design_vi = [item for item in items_vi if item['fulfillment_status'] == 'Sai design']
            design_uploaded_vi = [item for item in items_vi if item['fulfillment_status'] == 'Design uploaded']
            in_vi = [item for item in items_vi if item['fulfillment_status'] == 'In']
            out_vi = [item for item in items_vi if item['fulfillment_status'] == 'Out']
            half_done_vi = [item for item in items_vi if item['fulfillment_status'] == 'Half Done']
            designing_vi = [item for item in items_vi if item['fulfillment_status'] == 'Designing']
            ready_7_vi = [item for item in items_7_vi if item['fulfillment_status'] == 'Ready']
            none_vi = [item for item in items_vi if item['fulfillment_status'] == '']
            ready_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Ready']
            confirming_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Confirming']
            pending_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Pending']
            doing_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Doing']
            sai_design_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Sai design']
            design_uploaded_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Design uploaded']
            in_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'In']
            out_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Out']
            half_done_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Half Done']
            designing_thuan = [item for item in items_thuan if item['fulfillment_status'] == 'Designing']
            ready_7_thuan = [item for item in items_7_thuan if item['fulfillment_status'] == 'Ready']
            none_thuan = [item for item in items_thuan if item['fulfillment_status'] == '']


            # get total items have status in status
            total_ready = len([item for item in items if item['fulfillment_status'] == 'Ready'])
            total_confirming = len([item for item in items if item['fulfillment_status'] == 'Confirming'])
            total_pending = len([item for item in items if item['fulfillment_status'] == 'Pending'])
            total_doing = len([item for item in items if item['fulfillment_status'] == 'Doing'])
            total_sai_design = len([item for item in items if item['fulfillment_status'] == 'Sai design'])
            total_design_uploaded = len([item for item in items if item['fulfillment_status'] == 'Design uploaded'])
            total_in = len([item for item in items if item['fulfillment_status'] == 'In'])
            total_out = len([item for item in items if item['fulfillment_status'] == 'Out'])
            total_half_done = len([item for item in items if item['fulfillment_status'] == 'Half Done'])
            total_designing = len([item for item in items if item['fulfillment_status'] == 'Designing'])
            total_ready_7 = len([item for order in orders_7 for item in order['_original_data']['line_items']])
            total_none = len([item for item in nb_items if item['fulfillment_status'] == ''])

            data = {
                'total_ready': f"{total_ready} (Tùng: {len(ready_tung)}, Vi: {len(ready_vi)}, Trung: {len(ready_trung)}, Thuận: {len(ready_thuan)})",
                'total_confirming': f"{total_confirming} (Tùng: {len(confirming_tung)}, Vi: {len(confirming_vi)}, Trung: {len(confirming_trung)}, Thuận: {len(confirming_thuan)})",
                'total_pending': f"{total_pending} (Tùng: {len(pending_tung)}, Vi: {len(pending_vi)}, Trung: {len(pending_trung)}, Thuận: {len(pending_thuan)})",
                'total_doing': f"{total_doing} (Tùng: {len(doing_tung)}, Vi: {len(doing_vi)}, Trung: {len(doing_trung)}, Thuận: {len(doing_thuan)})",
                'total_sai_design': f"{total_sai_design} (Tùng: {len(sai_design_tung)}, Vi: {len(sai_design_vi)}, Trung: {len(sai_design_trung)}, Thuận: {len(sai_design_thuan)})",
                'total_design_uploaded': f"{total_design_uploaded} (Tùng: {len(design_uploaded_tung)}, Vi: {len(design_uploaded_vi)}, Trung: {len(design_uploaded_trung)}, Thuận: {len(design_uploaded_thuan)})",
                'total_in': f"{total_in} (Tùng: {len(in_tung)}, Vi: {len(in_vi)}, Trung: {len(in_trung)}, Thuận: {len(in_thuan)})",
                'total_out': f"{total_out} (Tùng: {len(out_tung)}, Vi: {len(out_vi)}, Trung: {len(out_trung)}, Thuận: {len(out_thuan)})",
                'total_half_done': f"{total_half_done} (Tùng: {len(half_done_tung)}, Vi: {len(half_done_vi)}, Trung: {len(half_done_trung)}, Thuận: {len(half_done_thuan)})",
                'total_designing': f"{total_designing} (Tùng: {len(designing_tung)}, Vi: {len(designing_vi)}, Trung: {len(designing_trung)}, Thuận: {len(designing_thuan)})",
                'total_ready_7': f"{total_ready_7} (Tùng: {len(ready_7_tung)}, Vi: {len(ready_7_vi)}, Trung: {len(ready_7_trung)}, Thuận: {len(ready_7_thuan)})",
                'total_none': f"{total_none} (Tùng: {len(none_tung)}, Vi: {len(none_vi)}, Trung: {len(none_trung)}, Thuận: {len(none_thuan)})"
            }

            return JSONResponse({'status': 200, 'message': 'Get statistics successfully', 'data': data})
    except:
        print(traceback.format_exc())
        return JSONResponse({'status': 400, 'message': 'Get statistics failed - ' + traceback.format_exc()})
    
async def export_data(request):
    try:
        db_client = request.state.db
        if request.method != 'POST':
            return JSONResponse({'status': 400, 'message': 'Invalid request method'})

        data = await request.json()
        date = data['date']
        start_date = convert_to_mongodb_time(date.split(' - ')[0], 'start')
        end_date = convert_to_mongodb_time(date.split(' - ')[1], 'end')

        team = data['team']
        type = data['type']

        if team == 'all':
            teams = await db_client.find_many({}, collection_name=DB['COL_TEAMS'], filter_={'name': 1})
            if not teams:
                return JSONResponse({'status': 400, 'message': 'Teams not found'})
            team = [team['name'] for team in teams if 'xuong' not in team['name']]
        else:
            team = [team]
        
        if type == 'finance':
            query = {
                '_team': {'$in': team},
                'create_at': {'$gte': start_date, '$lte': end_date}
            }
            filter_ = {
                'create_at': 1,
                '_original_data.line_items.fulfillment_order_id': 1,
                '_original_data.line_items.product_type': 1,
                '_original_data.line_items.quantity': 1,
                '_original_data.shop_id': 1,
                '_original_data.seller_id': 1,
                '_original_data.line_items.base_cost': 1,
                '_original_data.line_items.ship_cost': 1,
                '_original_data.shipping.shipping_price': 1,
                '_original_data.shipping.shipping_price_replace': 1,
                '_original_data.line_items.fulfillment_status': 1,
                '_original_data.line_items.note': 1,

            }
        elif type in ['designer', 'design_checker']:
            query = {
                '_team': {'$in': team},
                '_original_data.line_items.assign_designer': {'$exists': True},
                '_original_data.line_items.assign_designer.finished_at': {'$gte': start_date, '$lte': end_date},
                '_original_data.line_items.assign_designer.design_checked': True
            }
            filter_ = {
                'create_at': 1,
                '_original_data.line_items.fulfillment_order_id': 1,
                '_original_data.line_items.product_type': 1,
                '_original_data.line_items.design_type': 1,
                '_original_data.line_items.design_qty': 1,
                '_original_data.line_items.design_noted': 1,
                '_original_data.line_items.assign_designer': 1,
                '_original_data.seller_id': 1,
                '_original_data.line_items.design_type': 1,
            }
            design_types = await db_client.find_many({}, collection_name=DB['COL_DESIGN'], filter_={'name': 1, 'price': 1})
            design_types = {design_type['name']: float(design_type['price']) for design_type in design_types}
        elif type == 'tracking':
            query = {
                '_team': {'$in': team},
                'create_at': {'$gte': start_date, '$lte': end_date},
                '_original_data.shipping': {'$exists': True},
            }
            filter_ = {
                'create_at': 1,
                '_original_data.line_items.fulfillment_order_id': 1,
                '_original_data.line_items.product_type': 1,
                '_original_data.seller_id': 1,
                '_original_data.shipping': 1,
                '_original_data.delivery_info': 1,
            }
        elif type == 'shipping':
            query = {
                '_team': {'$in': team},
                'create_at': {'$gte': start_date, '$lte': end_date},
            }
            filter_ = {
                'create_at': 1,
                '_original_data.line_items.fulfillment_order_id': 1,
                '_original_data.line_items.product_type': 1,
                '_original_data.line_items.quantity': 1,
                '_original_data.line_items.note': 1,
                '_original_data.seller_id': 1,
                '_original_data.delivery_info': 1,
                '_original_data.line_items.other_option': 1,
            }
        elif type == 'thuy_cost':
            query = {
                '_original_data.line_items.factory': 'Thủy',
                'create_at': {'$gte': start_date, '$lte': end_date}
            }
            filter_ = {
                'create_at': 1,
                '_original_data.line_items.fulfillment_order_id': 1,
                '_original_data.line_items.line_item_id': 1,
                '_original_data.line_items.product_type': 1,
                '_original_data.line_items.quantity': 1,
                '_original_data.line_items.thuy_cost': 1,
                '_original_data.line_items.factory': 1,
                '_original_data.line_items.note': 1,
                '_original_data.line_items.design_note': 1,
            }
            

        # find orders
        orders = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        if not orders:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        
        if type == 'finance' and team == ['nb_team']:
            income_query = {
                'income.date': {'$gte': start_date, '$lte': end_date},
            }
            income_shops = await db_client.find_many_combo(income_query, collection_name=DB['COL_SHOP'], filter_={'income': 1, 'shop': 1, 'seller_id': 1})

            income_data = []  # Danh sách lưu trữ thông tin income

            if income_shops:
                for income_shop in income_shops:
                    for income in income_shop['income']:
                        if income['date'] >= start_date and income['date'] <= end_date:
                            income_data.append({
                                'Date': income['date'].strftime('%d/%m/%Y'),
                                'Shop': income_shop['shop'],
                                'Seller ID': income_shop['seller_id'],
                                'Income': income['total_income'],
                                'Currency': income['currency'],
                                'Note': income['note'],
                            })

        # create DataFrame
        df_data = []
        for order in orders:
            for item in order['_original_data']['line_items']:
                if type == 'finance':
                    shipping_cost = order['_original_data']['shipping']['shipping_price'] if 'shipping' in order['_original_data'] and order['_original_data']['shipping'] != {} else 0
                    provisional_shipping_price = item['ship_cost'] if 'ship_cost' in item else 0
                    item_shipping_cost = shipping_cost / len(order['_original_data']['line_items'])

                    shipping_price_replace = order['_original_data']['shipping']['shipping_price_replace'] if 'shipping' in order['_original_data'] and 'shipping_price_replace' in order['_original_data']['shipping'] else 0
                    shipping_price_replace = shipping_price_replace / len(order['_original_data']['line_items'])

                    df_data.append({
                        'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                        'Shop': order['_original_data']['shop_id'],
                        'Order ID': item['fulfillment_order_id'],
                        'Seller ID': order['_original_data']['seller_id'],
                        'Product type': item['product_type'],
                        'Quantity': item['quantity'],
                        'Base Cost': item['base_cost'] if 'base_cost' in item else 0,
                        'Provisional Shipping Cost': round(provisional_shipping_price, 2),
                        'Shipping Cost': round(item_shipping_cost, 2),
                        'Shipping Replace': round(shipping_price_replace, 2),
                        'Status': item['fulfillment_status'],
                        'Note': item['note'] if 'note' in item else '',
                    })
                elif type == 'designer':
                    created_at = item.get('assign_designer',{'created_at': ''}).get('created_at', '')
                    finished_at = item.get('assign_designer',{'finished_at':''}).get('finished_at', '')
                    designer =  item.get('assign_designer',{'designer':''}).get('designer', '')
                    
                    created_at = created_at.strftime('%d/%m/%Y') if created_at else ''
                    finished_at = finished_at.strftime('%d/%m/%Y') if finished_at else ''
                    try:
                        total_rejected = item['assign_designer']['total_rejected']
                    except:
                        total_rejected = 0
                    try:
                        design_qty = item.get('design_qty', '').replace(' ', '').replace('\n', '')
                        
                        if 'design_qty' in item and design_qty != '':
                            design_qtys = design_qty.upper().split('+')
                            price = 0
                            for des in design_qtys:
                                # check text 2 is number
                                try:
                                    string_2 = des[1:]
                                except:
                                    string_2 = ''
                                if not string_2.isdigit():
                                    try:
                                        qty = int(des[0])
                                        design_type = des[1:]
                                    except:
                                        qty = 1
                                        design_type = des
                                else:
                                    qty = int(des[:2])
                                    design_type = des[2:]
                                price += qty * design_types[design_type] if design_type in design_types else 0

                        else:
                            price = design_types[item['design_type']] if 'design_type' in item and item['design_type'] in design_types else 0 
                    except:
                        print(traceback.format_exc())
                        price = design_types[item['design_type']] if 'design_type' in item and item['design_type'] in design_types else 0
                    df_data.append({
                        'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                        'Order ID': item['fulfillment_order_id'],
                        'Seller ID': order['_original_data']['seller_id'],
                        'Product type': item['product_type'],
                        'Design type': item.get('design_type', ''),
                        'Design price': item.get('design_qty', ''),
                        'Design note': item.get('design_noted', ''),
                        'Assign date': created_at,
                        'Finish date': finished_at,
                        'Designer': designer,
                        'Total rejected': total_rejected,
                        'Price': price,
                    })
                elif type == 'design_checker':
                    if 'assign_designer' in item:
                        finished_at = item['assign_designer'].get('finished_at', '')
                        checker = item['assign_designer'].get('checker', '')
                    else:
                        finished_at = ''
                        checker = ''
                    finished_at = finished_at.strftime('%d/%m/%Y') if finished_at else ''
                    df_data.append({
                        'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                        'Order ID': item['fulfillment_order_id'],
                        'Product type': item['product_type'],
                        'Design type': item.get('design_type', ''),
                        'Design note': item.get('design_qty', ''),
                        'Checked date': finished_at,
                        'Checker': checker,
                    })
                elif type == 'tracking':
                    shipping_cost = order['_original_data']['shipping']['shipping_price'] if 'shipping' in order['_original_data'] else 0
                    item_shipping_cost = shipping_cost / len(order['_original_data']['line_items'])
                    df_data.append({
                        'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                        'Order ID': item['fulfillment_order_id'],
                        'Product type': item['product_type'],
                        'Seller ID': order['_original_data']['seller_id'],
                        'Tracking': order['_original_data']['shipping']['tracking_number'] if 'shipping' in order['_original_data'] else '',
                        'Carrier': order['_original_data']['shipping']['carrier'] if 'shipping' in order['_original_data'] else '',
                        'Shipping fee': round(item_shipping_cost, 2),
                        'Name': order['_original_data']['delivery_info']['name'],
                        'Address': order['_original_data']['delivery_info']['address1'],
                        'City': order['_original_data']['delivery_info']['city'],
                        'State': order['_original_data']['delivery_info']['state'],
                        'Zip': order['_original_data']['delivery_info']['zip'],
                        'Country': order['_original_data']['delivery_info']['country'],
                    })
                elif type == 'shipping':
                    other_option = item['other_option'] if 'other_option' in item else ''
                    
                    pack = func.get_pack_in_item(other_option)
                    quantity = int(item['quantity']) * pack
                    df_data.append({
                        'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                        'Order ID': item['fulfillment_order_id'],
                        'Seller ID': order['_original_data']['seller_id'],
                        'Product type': item['product_type'],
                        'Quantity': quantity,
                        'Name': order['_original_data']['delivery_info']['name'],
                        'Address': order['_original_data']['delivery_info']['address1'],
                        'City': order['_original_data']['delivery_info']['city'],
                        'State': order['_original_data']['delivery_info']['state'],
                        'Zip': order['_original_data']['delivery_info']['zip'],
                        'Country': order['_original_data']['delivery_info']['country'],
                        'Note': "Ship nhanh" if 'ship nhanh' in item['note'].lower() else '',
                    })
                elif type == 'thuy_cost':
                    if item['factory'] == 'Thủy':
                        df_data.append({
                            'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                            'Order ID': item['line_item_id'],
                            'Product type': item['product_type'],
                            'Quantity': item['quantity'],
                            'Thuy cost': item['thuy_cost'] if 'thuy_cost' in item else 0,
                            'Note': item['note'] if 'note' in item else '',
                            'Design note': item['design_note'] if 'design_note' in item else '',
                        })
                    

        df = pd.DataFrame(df_data)

        # Tạo tệp Excel bằng openpyxl trực tiếp
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active


        # Thêm dữ liệu vào tệp Excel
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        if type == 'finance' and team == ['nb_team']:
            new_df = pd.DataFrame(income_data)
            num_rows = new_df.shape[0]

            worksheet = workbook.create_sheet(title='Incomes')

            for row in dataframe_to_rows(new_df, index=False, header=True):
                worksheet.append(row)

        # Lưu tệp Excel vào BytesIO
        workbook.save(output)

        # in ra tệp Excel
        output.seek(0)

        start_date_str = start_date.strftime('%Y-%m-%d').replace('-', '')
        end_date_str = end_date.strftime('%Y-%m-%d').replace('-', '')

        excel_filename = f"export_{team}_{start_date_str}{end_date_str}.xlsx"
        response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers['Content-Disposition'] = f'attachment; filename="{excel_filename}"'
        return response

    except:
        print(traceback.format_exc())
        return JSONResponse({'status': 400, 'message': 'Export failed - ' + traceback.format_exc()})
    
async def update_rate(request):
    if request.method == "POST":
        try:
            db_client = request.state.db
            data = await request.json()
            rate = int(data.get('rate'))

            # update rate
            await db_client.update_one(ObjectId("64c6d3a49b03ffe1b59e554c"), {'rate': rate}, collection_name=DB['COL_RATE'])
            return JSONResponse({'status': 200, 'message': 'Update rate successfully'})
        except:
            return JSONResponse({'status': 400, 'message': 'Update rate failed - ' + traceback.format_exc()})
        
async def insert_income(request):
    try:
        db_client = request.state.db
        data = await request.json()
        _id = data.get('id')
        date = data.get('date')
        currency = data.get('currency')
        total_income = float(data.get('total_income'))
        note = data.get('note')

        date = convert_to_mongodb_time(date, 'start')

        # insert income
        shop = await db_client.find_one({'_id': ObjectId(_id)}, collection_name=DB['COL_SHOP'])

        if shop:
            # Check if 'income' key exists in shop and update it accordingly
            if 'income' in shop:
                income = shop['income']
                income.append({'date': date, 'currency': currency, 'total_income': total_income, 'note': note})
                shop['income'] = income
                update_ = await db_client.update_one(ObjectId(_id), shop, collection_name=DB['COL_SHOP'])
            else:
                income = [{'date': date, 'currency': currency, 'total_income': total_income, 'note': note}]
                shop['income'] = income
                update_ = await db_client.update_one(ObjectId(_id), shop, collection_name=DB['COL_SHOP'])
            # check if update successfully
            if update_:
                return JSONResponse({'status': 200, 'message': 'Insert income successfully'})
        else:
            raise Exception('Shop not found')
        
    except:
        return JSONResponse({'status': 400, 'message': 'Insert income failed - ' + traceback.format_exc()})
    
async def get_income_history(request):
    try:
        db_client = request.state.db
        _id = request.query_params.get('id')

        # get income history
        shop = await db_client.find_one({'_id': ObjectId(_id)}, collection_name=DB['COL_SHOP'])
        if shop:
            income = shop.get('income', [])
            for item in income:
                item['date'] = item['date'].strftime('%d/%m/%Y')
            return JSONResponse({'status': 200, 'message': 'Get income history successfully', 'incomeHistory': income})
        else:
            raise Exception('Shop not found')
    except:
        return JSONResponse({'status': 400, 'message': 'Get income history failed - ' + traceback.format_exc()})

async def get_barcode_data(request):
    try:
        data = await request.json()
        selected_values = data.get('selectedValues', [])

        # Create a PDF buffer using BytesIO
        pdf_buffer = BytesIO()

        # Create a new PDF with ReportLab
        c = canvas.Canvas(pdf_buffer, pagesize=letter)
        width, height = letter

        # List to store barcode data
        barcode_data_list = []

        # Loop through selected values
        num = 1
        for item in selected_values:
            order_id = str(item.get('orderId'))
            name = item.get('name')
            seller_id = item.get('seller_id')
            totalBarcode = int(item.get('totalBarcode'))

            # Get barcode value
            barcode_param = order_id

            # Create barcode image
            barcode_image = code128.image(barcode_param, height=100)

            # Create empty image for barcode + text
            top_bott_margin = 70
            l_r_margin = 10
            new_height = barcode_image.height + (2 * top_bott_margin)
            new_width = barcode_image.width + (2 * l_r_margin)
            new_image = Image.new('RGB', (new_width, new_height), (255, 255, 255))

            # put barcode on new image
            barcode_y = 100
            new_image.paste(barcode_image, (0, barcode_y))

            # object to draw text
            draw = ImageDraw.Draw(new_image)

            # Define custom text size and font
            h1_size = 28
            h2_size = 28
            h3_size = 16
            footer_size = 21

            # Define custom text
            product_type = seller_id + ' - ' + name
            center_product_type = (barcode_image.width / 2) - len(product_type) * 5
            center_barcode_value = (barcode_image.width / 2) - len(barcode_param) * 8

            # Draw text on picture
            draw.text((center_product_type, (h1_size + h2_size + h3_size)), product_type, fill=(0, 0, 0))
            draw.text((center_barcode_value, (new_height - footer_size - 15)), barcode_param, fill=(0, 0, 0))

            # Convert the image to RGB mode (if it's not already)
            new_image = new_image.convert("RGB")

            # Save the image as a temporary BytesIO object
            image_buffer = BytesIO()
            new_image.save(image_buffer, format="PNG")

            # Seek to the beginning of the buffer before saving it to the PDF
            image_buffer.seek(0)

            # Embed the image into the PDF
            pdf_image = ImageReader(image_buffer)
            c.drawImage(pdf_image, 100, height - 200, width=200, height=100)

            # Add a new page to the PDF
            c.showPage()

            # Append barcode data for HTML
            barcode_data_list.append({
                "pageNumber": num,
                "orderId": order_id,
                "barcodeSvg": func.barcode_image_to_svg(barcode_image, product_type, barcode_param, totalBarcode),
            })
            num += 1

        # Save the complete PDF to the buffer
        c.save()

        # Convert the PDF buffer to base64
        pdf_base64 = base64.b64encode(pdf_buffer.getvalue()).decode()

        # Create a dictionary for the response
        response_data = {
            "pageReady": True,
            "barcodeData": barcode_data_list,
            "pdfData": pdf_base64,
        }

        # Trả về response JSON
        return JSONResponse(content=response_data)

    except:
        return JSONResponse({'status': 400, 'message': 'Generate barcode failed' + traceback.format_exc()})
    
async def analytics(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "POST":
        try:
            data = await request.json()
            start_date = data.get('startDate')
            end_date = data.get('endDate')

            start_date = convert_to_mongodb_time(start_date, 'start')
            end_date = convert_to_mongodb_time(end_date, 'end')

            query = {
                "_team": user['team'],
                "create_at": {'$gte': start_date, '$lte': end_date},
                "_original_data.line_items.fulfillment_status": {"$nin": ["Cancel", "Refund"]},
                "_original_data.order_id": {"$not": {"$regex": "RE"}}
            }

            filter_ = {
                "_original_data": 1,
                "create_at": 1,
            }
            orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")

            # format data
            analytics_data = func.format_analytics_data(orders)

            count_sales, count_items, shop_sales, seller_sales = await func.get_count_sales(db_client, orders)

            # sort count_items by value
            count_items = dict(sorted(count_items.items(), key=lambda item: item[1], reverse=True))
            count_sales = dict(sorted(count_sales.items(), key=lambda item: item[1], reverse=True))
            seller_sales = dict(sorted(seller_sales.items(), key=lambda x: x[1], reverse=True))
            shop_sales = dict(sorted(shop_sales.items(), key=lambda x: x[1], reverse=True))

            context = {
                "result": analytics_data,
                "count_sales": count_sales,
                "count_items": count_items,
                "shop_sales": shop_sales,
                "seller_sales": seller_sales,
            }
            
            
            if user['team'] == 'nb_team':
                html = render_analytics(request, context)
            else:
                html = render_other_analytics(request, context)

            return HTMLResponse(html)
        except:
            return JSONResponse({'status': 400, 'message': 'Get analytics failed - ' + traceback.format_exc()})

async def analytics_private_team(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "POST":
        try:
            data = await request.json()
            date = data.get('date')
            start_date = date.split(' - ')[0]
            end_date = date.split(' - ')[1]

            start_date = convert_to_mongodb_time(start_date, 'start')
            end_date = convert_to_mongodb_time(end_date, 'end')
            
            query = {
                "_team": user['team'],
                "create_at": {'$gte': start_date, '$lte': end_date},
                "_original_data.line_items.fulfillment_status": {"$nin": ["Cancel", "Refund"]},
                "_original_data.order_id": {"$not": {"$regex": "RE"}}
            }

            private = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_name": 1, 'seller_id': 1, "team": 1})
            sellers = {}
            for item in private:
                if item['seller_name'] not in sellers:
                    sellers[item['seller_name']] = item['seller_id']
                    

            team = data.get('team')
            if team != '':
                if not private:
                    return JSONResponse({'status': 400, 'message': 'Private not found'})
                if team == 'all':
                    private = [item['seller_id'] for item in private]
                    private = list(set(private))
                    query['_original_data.seller_id'] = {"$in": private}
                else:
                    private = [item['seller_id'] for item in private if item['team'] == team]
                    private = list(set(private))
                    query['_original_data.seller_id'] = {"$in": private}
            
            seller = data.get('seller')
            if seller != '' and seller != 'all':
                query['_original_data.seller_id'] = sellers[seller]

            shop = data.get('shop')
            if shop != '' and shop != 'all':
                query['_original_data.shop_id'] = shop

            filter_ = {
                "_original_data": 1,
                "create_at": 1,
            }
            orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")

            # format data
            analytics_data = func.format_analytics_data(orders)

            count_sales, count_items, shop_sales, seller_sales = await func.get_count_sales(db_client, orders)

            # sort count_items by value
            count_items = dict(sorted(count_items.items(), key=lambda item: item[1], reverse=True))
            count_sales = dict(sorted(count_sales.items(), key=lambda item: item[1], reverse=True))
            seller_sales = dict(sorted(seller_sales.items(), key=lambda x: x[1], reverse=True))
            shop_sales = dict(sorted(shop_sales.items(), key=lambda x: x[1], reverse=True))

            context = {
                "result": analytics_data,
                "count_sales": count_sales,
                "count_items": count_items,
                "shop_sales": shop_sales,
                "seller_sales": seller_sales,
            }
            
            html = render_private_analytics(request, context)

            return HTMLResponse(html)
        except:
            return JSONResponse({'status': 400, 'message': 'Get analytics failed - ' + traceback.format_exc()})
    
async def designer_analytics(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            data = await request.json()
            start_date = data.get('startDate')
            end_date = data.get('endDate')

            start_date = convert_to_time(start_date)
            end_date = convert_to_time(end_date)

            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)

            query = {
                "_original_data.line_items.assign_designer.updated_at": {'$gte': start_date, '$lte': end_date},
                "_team": "nb_team"
            }
            filter_ = {
                "_original_data.line_items.assign_designer": 1,
                "_original_data.line_items.product_type": 1,
                "_original_data.line_items.fulfillment_order_id": 1,
            }
            orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
            analytics_data = func.format_analytics_designer(orders, start_date, end_date)

            context = {
                "meta_title": "Analytics",
                "records": analytics_data,
            }
            

            html = render_designer_analytics(request, context)

            return HTMLResponse(html)

        except:
            return JSONResponse({'status': 400, 'message': 'Get analytics failed - ' + traceback.format_exc()})
async def design_checked_analytics(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            data = await request.json()
            start_date = data.get('startDate')
            end_date = data.get('endDate')

            start_date = convert_to_mongodb_time(start_date, 'start')
            end_date = convert_to_mongodb_time(end_date, 'end')

            query = {
                "_original_data.line_items.assign_designer.design_checked": True,
                "_original_data.line_items.assign_designer.finished_at": {'$gte': start_date, '$lte': end_date},
            }
            filter_ = {
                "_original_data.line_items.seller": 1,
                "_original_data.line_items.assign_designer": 1,
            }
            orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
            analytics_data = func.format_analytics_design_checked(orders, start_date, end_date)

            context = {
                "meta_title": "Analytics",
                "records": analytics_data,
            }
            

            html = render_designer_analytics(request, context)

            return HTMLResponse(html)

        except:
            return JSONResponse({'status': 400, 'message': 'Get analytics failed - ' + traceback.format_exc()})

    
async def process_order(request):
    try:
        db_client = request.state.db
        data = await request.json()
        input_value = data.get('input')
        order_type = data.get('type')

        def create_result_item(order_id, status, link=''):
            return {
                'text': input_value,
                'order_id': order_id,
                'status': f" - {status}",
                'label_link': link
            }

        if '-' in str(input_value) and 'RE-' not in str(input_value):
            query = {
                '_original_data.line_items.line_item_id': str(input_value).replace('-', '.'),
            }
        elif '-' in str(input_value) and 'RE-' in str(input_value) and '.' not in str(input_value):
            text_split = str(input_value).split('-')
            try:
                prefix = text_split[0] + '-' + text_split[1] + '.' + text_split[2]
                query = {
                    '_original_data.line_items.line_item_id': prefix,
                }
            except:
                prefix = text_split[0] + '-' + text_split[1]
                query = {
                    '_original_data.order_id': prefix,
                }
            
        elif 'RE-' in str(input_value) and '.' in str(input_value):
            query = {
                '_original_data.line_items.line_item_id': str(input_value),
            }
        else:
            query = {
                '_original_data.order_id': str(input_value),
            }

        if order_type == 'order':
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})

            if not order:
                result = [create_result_item('Order not found', '')]
                return JSONResponse({'status': 400, 'message': 'Order not found', 'result': result})

            line_items = order['_original_data']['line_items']
            tracking_label = order['_original_data']['shipping']['tracking_label'] if 'shipping' in order['_original_data'] and 'tracking_label' in order['_original_data']['shipping'] else ''
            if len(line_items) > 1:
                result = [create_result_item(item['line_item_id'], item['fulfillment_status'], tracking_label) for item in line_items]
            else:
                result = [create_result_item(input_value, line_items[0]['fulfillment_status'], tracking_label)]
        elif order_type == 'tracking':
            tracking_number = input_value
            orders = await db_client.find_many_combo({'_original_data.shipping.tracking_number': str(tracking_number)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})

            if orders == []:
                result = [create_result_item('Order not found', '')]
                return JSONResponse({'status': 400, 'message': 'Order not found', 'result': result})

            line_items = []
            for order in orders:
                line_items.extend(order['_original_data']['line_items'])

            if len(line_items) > 1:
                result = [create_result_item(item['line_item_id'], item['fulfillment_status']) for item in line_items]
            else:
                result = [create_result_item(line_items[0]['fulfillment_order_id'], line_items[0]['fulfillment_status'])]

            
        else:
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1, 'create_at': 1})
            if not order:
                result = [{
                    'text': input_value,
                    'order_id': 'Order not found',
                    'link_des': [],
                    'mockup': '',
                    'size': '',
                    'personalization': '',
                    'note': '',
                    'quantity': '',
                    'date': '',
                    'no_holder': '',
                    'material': '',
                    'color': '',
                    'option': ''
                }]
                return JSONResponse({'status': 400, 'message': 'Order not found', 'result': result})

            line_items = order['_original_data']['line_items']
            if len(line_items) > 1:
                if '-' in str(input_value):
                    for item in line_items:
                        if str(input_value).replace('-', '.') == item['line_item_id']:
                            other_option = item.get('other_option', '').lower()
                            
                            pack = func.get_pack_in_item(other_option)
                            quantity = int(item['quantity']) * pack
                            if type(item['link_des']) == list:
                                link_des = item['link_des']
                            else:
                                link_des = [item['link_des']]

                            result = [{
                                'text': input_value,
                                'order_id': item.get('line_item_id', ''),
                                'link_des': link_des,
                                'mockup': item.get('thumbnail', ''),
                                'size': item.get('size', ''),
                                'personalization': item.get('personalization', ''),
                                'note': item.get('note', ''),
                                'quantity': quantity,
                                'date': order['create_at'].strftime('%d/%m/%Y'),
                                'no_holder': item.get('num_holder', ''),
                                'material': item.get('material', ''),
                                'color': item.get('color', ''),
                                'option': item.get('other_option', ''),
                            }]
                else:
                    result = []
                    for item in line_items:
                        other_option = item.get('other_option', '').lower()
                        
                        pack = func.get_pack_in_item(other_option)
                        quantity = int(item['quantity']) * pack
                        if type(item['link_des']) == list:
                            link_des = item['link_des']
                        else:
                            link_des = [item['link_des']]

                        result.append({
                            'text': input_value,
                            'order_id': item.get('line_item_id', ''),
                            'link_des': link_des,
                            'mockup': item.get('thumbnail', ''),
                            'size': item.get('size', ''),
                            'personalization': item.get('personalization', ''),
                            'note': item.get('note', ''),
                            'quantity': quantity,
                            'date': order['create_at'].strftime('%d/%m/%Y'),
                            'no_holder': item.get('num_holder', ''),
                            'material': item.get('material', ''),
                            'option': item.get('other_option', ''),
                        })

                
            else:
                if type(line_items[0]['link_des']) == list:
                    link_des = line_items[0]['link_des']
                else:
                    link_des = [line_items[0]['link_des']]
                try:
                    other_option = line_items[0]['other_option'].lower()
                except:
                    other_option = ''
                
                pack = func.get_pack_in_item(other_option)      
                    
                quantity = int(line_items[0]['quantity']) * pack
                result = [{
                    'text': input_value,
                    'order_id': line_items[0].get('line_item_id', ''),
                    'link_des': link_des,
                    'mockup': line_items[0].get('thumbnail', ''),
                    'size': line_items[0]['size'],
                    'personalization': line_items[0].get('personalization', ''),
                    'note': line_items[0].get('note', ''),
                    'quantity': quantity,
                    'date': order['create_at'].strftime('%d/%m/%Y'),
                    'no_holder': line_items[0].get('num_holder', ''),
                    'material': line_items[0].get('material', ''),
                    'color': line_items[0].get('color', ''),
                    'option': line_items[0].get('other_option', '')
                }]

        return JSONResponse({'status': 200, 'message': 'Get orders successfully', 'result': result})
    except:
        return JSONResponse({'status': 400, 'message': 'Get orders failed - ' + traceback.format_exc()})

async def export_scan_data(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        order_id = data.get('orderIds')
        order_ids = [id for id in order_id if 'Order not found' not in id]

        tracking_number = data.get('trackingNumbers')
        tracking_numbers = [id for id in tracking_number if 'Order not found' not in id]

        if order_ids != []:
            query = {
                '_original_data.order_id': {'$in': order_ids},
            }
        else:
            query = {
                '_original_data.shipping.tracking_number': {'$in': tracking_numbers},
                '_original_data.line_items[0].fulfillment_status': {'$nin': ['Done', 'Cancel']},
            }
        
        # get orders
        filter_ = {
            '_original_data': 1,
            'create_at': 1,
        }
        orders = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)

        if orders == []:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        
        # create DataFrame
        df_data = []
        if order_ids != []:
            for order in orders:
                for item in order['_original_data']['line_items']:
                    other_option = item.get('other_option', '').lower()
                    
                    pack = func.get_pack_in_item(other_option)
                    quantity = int(item['quantity']) * pack

                    df_data.append({
                        'Order Date': order['create_at'].strftime('%d/%m/%Y'),
                        'Order ID': item['fulfillment_order_id'],
                        'Seller ID': order['_original_data']['seller_id'],
                        'Product type': item['product_type'],
                        'Quantity': quantity,
                        'Name': order['_original_data']['delivery_info']['name'],
                        'Address': order['_original_data']['delivery_info']['address1'],
                        'City': order['_original_data']['delivery_info']['city'],
                        'State': order['_original_data']['delivery_info']['state'],
                        'Zip': order['_original_data']['delivery_info']['zip'],
                        'Country': order['_original_data']['delivery_info']['country'],
                        'Note': "Ship nhanh" if 'ship nhanh' in item['note'].lower() else '',
                    })
        else:
            for order in orders:
                df_data.append({
                    'Order ID': order['_original_data']['order_id'],
                    'Tracking No': order['_original_data']['shipping']['tracking_number'] if 'shipping' in order['_original_data'] else '',
                    'Carrier': order['_original_data']['shipping']['carrier'] if 'shipping' in order['_original_data'] else '',
                })

        df = pd.DataFrame(df_data)

        # Tạo tệp Excel bằng openpyxl trực tiếp
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active

        # Thêm dữ liệu vào tệp Excel
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Lưu tệp Excel vào BytesIO
        workbook.save(output)
        
        # in ra tệp Excel
        output.seek(0)

        excel_filename = f"scan_data.xlsx"

        response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers['Content-Disposition'] = f'attachment; filename="{excel_filename}"'

        return response

async def export_scan_kiot(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        order_id = data.get('orderIds')
        order_ids = [id for id in order_id if 'Order not found' not in id]

        tracking_number = data.get('trackingNumbers')
        tracking_numbers = [id for id in tracking_number if 'Order not found' not in id]

        if order_ids != []:
            query = {
                '_original_data.order_id': {'$in': order_ids},
            }
        else:
            query = {
                '_original_data.shipping.tracking_number': {'$in': tracking_numbers},
                '_original_data.line_items[0].fulfillment_status': {'$nin': ['Done', 'Cancel']},
            }
        
        # get orders
        filter_ = {
            '_original_data': 1,
            'create_at': 1,
        }
        orders = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)

        if orders == []:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        
        # create DataFrame
        df_data = []
        for order in orders:
            df_data.append({
                'Loại hàng': 'Hàng hóa',
                'Nhóm hàng(3 Cấp)': 'Phụ kiện Nam',
                'Mã hàng': order['_original_data']['order_id'],
                'Mã vạch': '',  # You can fill this with appropriate data
                'Tên hàng': order['_original_data']['seller_id'],
                'Thương hiệu': '',
                'Giá bán': 0,
                'Giá vốn': 0,
                'Tồn kho': 0,
                'KH đặt': 0,
                'Dự kiến hết hàng': '0 Ngày',
                'Tồn nhỏ nhất': 0,
                'Tồn lớn nhất': 999,
                'ĐVT': '',
                'Mã ĐVT Cơ bản': '',
                'Quy đổi': 1,
                'Thuộc tính': f"Tên: {order['_original_data']['delivery_info']['name']}",
                'Mã HH Liên quan': '',
                'Hình ảnh (url1,url2...)': '',
                'Trọng lượng': 0,
                'Đang kinh doanh': 1,
                'Được bán trực tiếp': 1,
                'Mô tả': 'Tên: ',
                'Mẫu ghi chú': order['_original_data']['delivery_info']['name'],
                'Vị trí': '',
                'Hàng thành phần': ''
            })

        df = pd.DataFrame(df_data)

        # Tạo tệp Excel bằng openpyxl trực tiếp
        output = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active

        # Thêm dữ liệu vào tệp Excel
        for row in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(row)

        # Lưu tệp Excel vào BytesIO
        workbook.save(output)
        
        # in ra tệp Excel
        output.seek(0)

        excel_filename = f"scan_kiot.xlsx"

        response = StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers['Content-Disposition'] = f'attachment; filename="{excel_filename}"'

        return response

async def scan_change_order_status(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        order_ids_sent = data.get('orderIds')

        order_ids = []
        line_item_ids = []
        for order_id in order_ids_sent:
            if 'Order not found' not in order_id:
                if '-' in str(order_id) and 'RE-' not in str(order_id):
                    line_item_id = str(order_id).replace('-', '.')
                else:
                    line_item_id = str(order_id) + '.1'
                
                order_ids.append(str(order_id))
                line_item_ids.append(line_item_id)

        tracking_number = data.get('trackingNumbers')
        tracking_numbers = [id for id in tracking_number if 'Order not found' not in id]

        query = {
            '_original_data.order_id': {'$in': order_ids},
            '_original_data.line_items.line_item_id': {'$in': line_item_ids},
        }
        
        # get orders
        filter_ = {
            '_original_data': 1,
            '_team': 1,
        }
        
        is_tracking = False
        orders = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
        
        if orders == []:
            query = {
                '_original_data.shipping.tracking_number': {'$in': tracking_numbers},
                '_original_data.line_items.fulfillment_status': {'$nin': ['Done', 'Cancel', 'Refund']},
            }
            orders = await db_client.find_many(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_)
            is_tracking = True
        
        if orders == []:
            return JSONResponse({'status': 400, 'message': 'Orders not found'})
        
        try:
            # update orders
            for order in orders:
                if is_tracking == False:
                    for line_item in order['_original_data']['line_items']:
                        if line_item['line_item_id'] in line_item_ids or line_item['fulfillment_order_id'] in order_ids:
                            user_logs = line_item.get('user_logs', [])
                            user_log = f"{request.state.user_data['username']} changed status in scan from {line_item['fulfillment_status']} -> {data['status']} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                            user_logs.append(user_log)
                            line_item['fulfillment_status'] = data['status']
                            
                            if data['status'] in ['Ready'] and line_item['base_cost'] == 0:
                                # get product basecost
                                filter_ = {
                                    'base_cost': 1,
                                    'us_shipping_1st': 1,
                                    'us_shipping_additional': 1,
                                    'ww_shipping_1st': 1,
                                    'ww_shipping_additional': 1,
                                }

                                product = await db_client.find_one({'product_type': line_item['product_type']}, collection_name=DB['COL_PRODUCTS'], filter_=filter_)
                                # update line_item basecost
                                
                                try:
                                    if product is not None:
                                        other_option = line_item.get('other_option', '').lower()
                                        pack = func.get_pack_in_item(other_option)
                                        
                                        quantity = int(line_item['quantity']) * pack
                                        basecost = round(float(product['base_cost']) * quantity, 2)
                                        line_item['base_cost'] = basecost
                                        basecost_log = f"Basecost: {line_item['base_cost']} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                                        user_logs.append(basecost_log)
                                        # update line_item ship_cost
                                        if 'ship_cost' not in line_item or line_item['ship_cost'] == 0:
                                            # Provisional shipping fee
                                            ship_cost = 0
                                            if order['_original_data']['delivery_info']['country'].lower() in ['us', 'united states']:
                                                ship_cost = product['us_shipping_1st']
                                                if quantity > 1:
                                                    ship_cost += round(product['us_shipping_additional'] * (quantity - 1), 2)
                                            else:
                                                ship_cost = product['ww_shipping_1st']
                                                if quantity > 1:
                                                    ship_cost += round(product['ww_shipping_additional'] * (quantity - 1), 2)
                                            line_item['ship_cost'] = ship_cost
                                            ship_cost_log = f"Provisional shipping fee {line_item['ship_cost']} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                                            user_logs.append(ship_cost_log)

                                except:
                                    return JSONResponse({'status': 400, 'message': f"Product {line_item['product_type']} doesn't have basecost!"})

                                # update order basecost
                                if order['_team'] != 'nb_team':
                                    query = {
                                        'name': order['_team'],
                                    }
                                    team = await db_client.find_one(query, collection_name=DB['COL_TEAMS'])
                                    if team is None:
                                        return JSONResponse({'status': 400, 'message': f"Team with name {order['_team']} not found!"})
                                    else:
                                        if 'balance' not in team:
                                            team['balance'] = 0
                                        current_balance = team['balance']
                                        balance = round(float(current_balance) - basecost - ship_cost, 2)
                                        # if balance <= 0:
                                        #     return JSONResponse({'status': 400, 'message': f"Team {order['_team']} doesn't have enough money to pay for this order!"})
                                        # else:
                                        team['balance'] = balance
                                        await db_client.update_one(team['_id'], team, collection_name=DB['COL_TEAMS'])
                            
                            line_item['user_logs'] = user_logs
                else:
                    for line_item in order['_original_data']['line_items']:
                        user_logs = line_item.get('user_logs', [])
                        user_log = f"{request.state.user_data['username']} changed status in scan from {line_item['fulfillment_status']} -> {data['status']} at {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        user_logs.append(user_log)
                        line_item['fulfillment_status'] = data['status']
                        
                        if data['status'] in ['Ready'] and line_item['base_cost'] == 0:
                            # get product basecost
                            filter_ = {
                                'base_cost': 1,
                                'us_shipping_1st': 1,
                                'us_shipping_additional': 1,
                                'ww_shipping_1st': 1,
                                'ww_shipping_additional': 1,
                            }
                            product = await db_client.find_one({'product_type': line_item['product_type']}, collection_name=DB['COL_PRODUCTS'], filter_=filter_)
                            # update line_item basecost
                            try:
                                if product is not None:
                                    other_option = line_item.get('other_option', '').lower()
                                    pack = func.get_pack_in_item(other_option)
                                    
                                    quantity = int(line_item['quantity']) * pack
                                    basecost = round(float(product['base_cost']) * quantity, 2)
                                    line_item['base_cost'] = basecost
                                    basecost_log = f"Basecost: {line_item['base_cost']} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                                    user_logs.append(basecost_log)

                                    # update line_item ship_cost
                                    if 'ship_cost' not in line_item or line_item['ship_cost'] == 0:
                                        # Provisional shipping fee
                                        ship_cost = 0
                                        if order['_original_data']['delivery_info']['country'].lower() in ['us', 'united states']:
                                            ship_cost = product['us_shipping_1st']
                                            if quantity > 1:
                                                ship_cost += round(product['us_shipping_additional'] * (quantity - 1), 2)
                                        else:
                                            ship_cost = product['ww_shipping_1st']
                                            if quantity > 1:
                                                ship_cost += round(product['ww_shipping_additional'] * (quantity - 1), 2)
                                        line_item['ship_cost'] = ship_cost
                                        ship_cost_log = f"Provisional shipping fee {line_item['ship_cost']} - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                                        user_logs.append(ship_cost_log)
                            
                            except:
                                return JSONResponse({'status': 400, 'message': f"Product {line_item['product_type']} doesn't have basecost!"})

                            # update order basecost
                            if order['_team'] != 'nb_team':
                                query = {
                                    'name': order['_team'],
                                }
                                team = await db_client.find_one(query, collection_name=DB['COL_TEAMS'])
                                if team is None:
                                    return JSONResponse({'status': 400, 'message': f"Team with name {order['_team']} not found!"})
                                else:
                                    if 'balance' not in team:
                                        team['balance'] = 0
                                    current_balance = team['balance']
                                    balance = round(float(current_balance) - basecost - ship_cost, 2)
                                    # if balance <= 0:
                                    #     return JSONResponse({'status': 400, 'message': f"Team {order['_team']} doesn't have enough money to pay for this order!"})
                                    # else:
                                    team['balance'] = balance
                                    await db_client.update_one(team['_id'], team, collection_name=DB['COL_TEAMS'])
                        
                        line_item['user_logs'] = user_logs
                
                await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])

            return JSONResponse({'status': 200, 'message': 'Change status successfully'})
        except:
            return JSONResponse({'status': 400, 'message': 'Change status failed - ' + traceback.format_exc()})
    
async def upload_clipart(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.body()
        data = json.loads(data)
        list_fonts = []
        exist_option = []
        custom_layer = {}
        list_background = []
        links = google_storage.upload_to_bucket_personalize(data['google_files'])
        for file in data['results']:
            file['link'] = links[file['type']][file['file_name']]
            if file['type'] == 'font':
                list_fonts.append(file)
            elif file['type'] == 'background':
                list_background.append(file)
            elif file['type'] == 'option' and file['name_custom'] in exist_option:
                custom_layer[file['name_custom']].append({'name':file['name'],'path':file['link']})
            elif file['type'] == 'option' and file['name_custom'] not in exist_option:
                exist_option.append(file['name_custom'])
                custom_layer[file['name_custom']] = [{'name':file['name'],'path':file['link']}]
            elif file['type'] == 'mockup':
                mockup = file['link']
        
        doc = {
            'name': data['name'],
            'custom_files' :custom_layer,
            'list_fonts':list_fonts,
            'list_background':list_background,
            'mockup':mockup
        }
        await db_client.insert_one(doc, collection_name=DB['COL_PERSONALIZE'])
        return JSONResponse({'status': 200, 'message': 'Upload clipart successfully'})
        # process zip data

async def save_pdf(request):
    db_client = request.state.db
    if request.method == "POST":
        data = await request.json()
        psd_mongo_id = data.get('psd_mongo_id','')
        if psd_mongo_id != '':
        # update psd
            psd = await db_client.find_one({'_id': ObjectId(psd_mongo_id)}, collection_name=DB['COL_PSD'], filter_=None)
            psd['process_pdf'] = True
            await db_client.update_one(psd['_id'], psd, collection_name=DB['COL_PSD'])
            pdf_find = await db_client.find_one({'psd_mongo_id': psd_mongo_id}, collection_name=DB['COL_PDF'], filter_=None)
            if pdf_find:
                await db_client.update_one(pdf_find['_id'],data, collection_name=DB['COL_PDF'])
                return JSONResponse({'status': 200, 'message': 'Update Pdf successfully'})
            await db_client.insert_one(data, collection_name=DB['COL_PDF'])    
        else:
            await db_client.insert_one(data, collection_name=DB['COL_PDF'])
        return JSONResponse({'status': 200, 'message': 'Save Pdf successfully'})
    
async def get_order_history(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            data = await request.json()
        except:
            return JSONResponse({'status': 400, 'message': 'Form data not found'})    
        order_id = data.get('orderId','')
        line_item_id = data.get('lineItemId','')

        if order_id != '':
            # get order history
            order = await db_client.find_one({'_original_data.order_id': str(order_id)}, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})

            if not order:
                return JSONResponse({'status': 400, 'message': 'Order not found'})
            
            line_items = order['_original_data']['line_items']

            result = []
            for item in line_items:
                if item['line_item_id'] == str(line_item_id):
                    try:
                        user_logs = []
                        for logs in item['user_logs']:
                            if not isinstance(logs, list) and logs not in user_logs:
                                user_logs.append(logs)

                    except:
                        print(traceback.format_exc())
                    result = list(user_logs) # Loại bỏ các giá trị trùng lặp và chuyển thành danh sách
                    result.reverse()

            return JSONResponse({'status': 200, 'message': 'Get order history successfully', 'result': result})
        else:
            return JSONResponse({'status': 400, 'message': 'Order not found'})


async def edit_order_note(request):
    db_client = request.state.db
    if request.method == "POST":
        try:
            form_data = await request.form()
        except:
            return JSONResponse({'status': 400, 'message': 'Form data not found'})    
        order_id = form_data['order_id']
        line_item_id = form_data['line_item_id']
        note = form_data['note']
        query = {
            '_original_data.order_id': order_id,
        }
        order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'], filter_={'_original_data': 1})
        if not order:
            return JSONResponse({'status': 400, 'message': 'Order not found'})
        item = next((line_item for line_item in order['_original_data'].get('line_items', []) if line_item['line_item_id'] == line_item_id), None)
        if not item:
            return JSONResponse({'status': 400, 'message': 'Line item not found'})
        index = order['_original_data']['line_items'].index(item)
        order['_original_data']['line_items'][index]['note'] = note
        await db_client.update_one(order['_id'], order, collection_name=DB['COL_FULFILLMENTS'])
        
        return JSONResponse({'status': 200, 'message': 'Edit order note images successfully'})