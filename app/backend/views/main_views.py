from starlette.responses import PlainTextResponse, FileResponse, RedirectResponse, JSONResponse
from starlette.responses import StreamingResponse
import time, datetime
from bson import ObjectId
from resources.db import convert_to_mongodb_time, convert_to_time
from ..modules.users import authenticate,db_get_users_not_in_team,db_get_users_by_ids
from ..modules.func import format_show_orders, insert_order_db, build_query_filter_orders
from ..modules import func, read_pdf
from ..modules.data_model import Products, UserLog, Warehouse, Teams, Design
from ..resources import check_token, templates, get_page_data, make_page_navi
from ..settings import DB, THEME, list_status, factories
from io import BytesIO
import base64
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import asyncio
import traceback
from datetime import datetime, timedelta
import urllib



async def favicon(request):
    return FileResponse(f"./backend/themes/{THEME}/assets/img/favicon.ico")

#
# Home
#
@authenticate
async def home_page(request):
    db_client = request.state.db

    # start = datetime.now() - datetime.timedelta(days=7)
    # end = datetime.now()

    # default_prodct = ['Wooden Bar KC']

    # start_datetime = convert_to_mongodb_time(start, "start")
    # end_datetime = convert_to_mongodb_time(end, "end")
   
    # query = [
    #         {"$match": {"data_date.date": {"$gte": start_datetime, "$lte": end_datetime}, "name": {"$in": default_prodct}}},
    #         {"$project": {"name": 1, "data_date.date": 1, "data_date.quantity": 1}},
    #         {"$unwind": "$data_date"},
    #         {"$match": {"data_date.date": {"$gte": start_datetime, "$lte": end_datetime}, "name": {"$in": default_prodct}}},
    #         {"$group": {"_id": {"name": "$name", "date": "$data_date.date"}, "quantity": {"$sum": "$data_date.quantity"}}},
    #         {"$project": {"name": "$_id.name", "date": "$_id.date", "quantity": 1, "_id": 0}}
    #     ]
    #     # Find the products 
    # products = await db_client.query_aggregate(query, collection_name=DB['COL_WAREHOUSE'])

    # result = format_data(products)
    # data = {
    #     "products": result,
    # }

    # get team of current user
    user = request.state.user_data
    team = user['team']
    try:
        team = await db_client.find_one({"name": team}, collection_name=DB['COL_TEAMS'])
        current_balance = team['balance']
    except:
        current_balance = 0

    data = {
        "meta_title": "Home",
        "current_balance": current_balance,
    }


    return templates.TemplateResponse('index.html', {'request': request, 'data': data})

#
# Warehouse data
#
@authenticate
async def list_products(request):
    # Query
    db_client = request.state.db
    page_number = 1
    skip_ = (page_number - 1) * 100
    limit_ = 100
    # Get page
    page_number, limit_, skip_ = await get_page_data(request, limit_=limit_)
    
    # Query
    db_client = request.state.db
    query = {}

    filter_ = {
        "category": 1,
        "name": 1,
        "sku": 1,
        "img": 1,
        "quantity": 1,
        "min": 1,
        "material": 1,
        "color": 1,
        "data_date": 1,
        "description": 1,
    }
    products = await db_client.find_many_combo(query, collection_name=DB['COL_WAREHOUSE'], filter_=filter_, limit_=limit_, sort_="sku", skip_=skip_)
        
    data = {
        "meta_title": "Warehouse",
        "products": products,
    }

    return templates.TemplateResponse('factory/list_products.html', {'request': request, 'data': data})

async def list_export_ornc(request):
    # Query
    db_client = request.state.db
    page_number = 1
    skip_ = (page_number - 1) * 100
    limit_ = 100
    # Get page
    page_number, limit_, skip_ = await get_page_data(request, limit_=limit_)
    
    # Query
    db_client = request.state.db
    query = {}

    filter_ = {
       
    }
    exports = await db_client.find_many_combo(query, collection_name=DB['COL_EXPORT_ORNC'], filter_=filter_, limit_=limit_, sort_="sku", skip_=skip_)
    for export in exports:
        export['created_at'] = export['created_at'].strftime("%Y-%m-%d")
        if 'updated_at' in export:
            export['updated_at'] = export['updated_at'].strftime("%Y-%m-%d")
    data = {
        "exports": exports,
    }

    return templates.TemplateResponse('fulfillment/list_export_ornc.html', {'request': request, 'data': data})




@authenticate
async def out_of_stock(request):
    # Query
    db_client = request.state.db
    page_number = 1
    skip_ = (page_number - 1) * 100
    limit_ = 100
    # Get page
    page_number, limit_, skip_ = await get_page_data(request, limit_=limit_)
    
    # Query
    db_client = request.state.db
    query = {}

    filter_ = {
        "category": 1,
        "name": 1,
        "sku": 1,
        "img": 1,
        "quantity": 1,
        "min": 1,
        "material": 1,
        "color": 1,
        "data_date": 1,
        "description": 1,
    }
    products = await db_client.find_many_combo(query, collection_name=DB['COL_WAREHOUSE'], filter_=filter_, limit_=limit_, sort_="_id", skip_=skip_)

    # list all product if quantity <= min
    products = [product for product in products if product['min'] >= product['quantity']]
        
    data = {
        "meta_title": "Warehouse",
        "products": products,
    }

    return templates.TemplateResponse('factory/list_products.html', {'request': request, 'data': data})
@authenticate
async def edit_product(request):
    data = {}
    current_url = str(request.url)
    user = request.state.user_data
    db_client = request.state.db
    if request.method == "GET":
        _id = request.query_params.get('id')
        query = {"_id": ObjectId(_id)}
        product = await db_client.find_one(query, collection_name=DB['COL_WAREHOUSE'])
        data = {
            "meta_title": "Warehouse",
            "product": product,
        }
        return templates.TemplateResponse('factory/edit_product.html', {'request': request, 'data': data})
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        img = form_data['img'].file
        encoded_string = base64.b64encode(img.read()).decode('utf-8')

        time_now = convert_to_mongodb_time(time.time())
        today = datetime.now()

        product = await db_client.find_one({'_id': ObjectId(form_data['_id'])}, collection_name=DB['COL_WAREHOUSE'])

        new_quantity = int(form_data['quantity'])
        new_min = int(form_data['min'])
        ok = False
        for i in range(len(product['data_date'])):
            if product['data_date'][i]['date'].strftime("%Y-%m-%d") == today.strftime("%Y-%m-%d"):
                product['data_date'][i]['quantity'] = new_quantity
                ok = True
                break
        if not ok:
            product['data_date'].append({'date': today, 'quantity': new_quantity})

        if encoded_string == "":
            encoded_string = product['img']
        document = Warehouse(
            category = form_data['cate'],
            name = form_data['name'],
            sku = form_data['sku'],
            img = encoded_string,
            quantity = new_quantity,
            min = new_min,
            material = form_data['material'],
            color = form_data['color'],
            data_date = product['data_date'],
            description = form_data['description'],
            updated_at = time_now,
            user_logs = [f'{user["username"]} - {time_now} - insert new product {form_data["name"]}']
        ).dict()
        
        update_ = await db_client.update_one(form_data['_id'], document, collection_name=DB['COL_WAREHOUSE'])
        return RedirectResponse(url=current_url, status_code=302)
@authenticate
async def add_product(request):
    data = {
        "meta_title": "Warehouse",
    }
    user = request.state.user_data
    current_url = str(request.url)  
    if request.method == "GET":
        return templates.TemplateResponse('factory/add_product.html', {'request': request, 'data': data})
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        
        img = form_data['img'].file
        encoded_string = base64.b64encode(img.read()).decode('utf-8')
        time_now = convert_to_mongodb_time(time.time())
        today = datetime.now()

        document = Warehouse(
            category = form_data['cate'],
            name = form_data['name'],
            sku = form_data['sku'],
            img = encoded_string,
            quantity = int(form_data['quantity']),
            min = int(form_data['min']),
            material = form_data['material'],
            color = form_data['color'],
            data_date = [
                {
                    'date': today,
                    'quantity': int(form_data['quantity'])
                }
            ],
            description = form_data['description'],
            updated_at = time_now,
            user_logs = [f'{user["username"]} - {time_now} - insert new product {form_data["name"]}']
        ).dict()
        update_ = await db_client.insert_one(document, collection_name=DB['COL_WAREHOUSE'])
        return RedirectResponse(url=current_url, status_code=302)

#
# Fulfillment products
#
@authenticate
async def ff_list_products(request):
    # Query
    db_client = request.state.db
    page_number = 1
    skip_ = (page_number - 1) * 200
    limit_ = 200
    # Get page
    page_number, limit_, skip_ = await get_page_data(request, limit_=limit_)
    
    # Query
    query = {}
    filter_ = {
        "name": 1,
        "sku": 1,
        "xuong": 1,
        "description": 1,
        "box_size": 1,
        "product_type": 1,
        "quantity": 1,
        "product_size": 1,
        "product_template_width": 1,
        "product_template_height": 1,
        "base_cost": 1,
        "thuy_cost": 1,
        "us_shipping_additional": 1,
        "us_shipping_1st": 1,
        "ww_shipping_1st": 1,
        "ww_shipping_additional": 1,
    }
    products = await db_client.find_many_combo_sort_lasted(query, collection_name=DB['COL_PRODUCTS'], filter_=filter_, limit_=limit_, sort_="sku", skip_=skip_)
        
    data = {
        "meta_title": "Products",
        "products": products,
    }

    return templates.TemplateResponse('fulfillment/list_products.html', {'request': request, 'data': data})
@authenticate
async def ff_edit_product(request):
    data = {}
    current_url = str(request.url)
    user = request.state.user_data
    db_client = request.state.db
    if request.method == "GET":
        _id = request.query_params.get('id')
        query = {"_id": ObjectId(_id)}
        product = await db_client.find_one(query, collection_name=DB['COL_PRODUCTS'])
        data = {
            "meta_title": "Products",
            "product": product,
            "factories": factories,
        }
        return templates.TemplateResponse('fulfillment/edit_product.html', {'request': request, 'data': data})
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        img = form_data['img'].file
        encoded_string = base64.b64encode(img.read()).decode('utf-8')
        time_now = convert_to_mongodb_time(time.time())

        product = await db_client.find_one({'_id': ObjectId(form_data['_id'])}, collection_name=DB['COL_PRODUCTS'])

        if encoded_string == "":
            encoded_string = product['img']

        print(form_data['us_shipping_1st'], form_data['us_shipping_additional'], form_data['ww_shipping_1st'], form_data['ww_shipping_additional'])

        document = Products(
            name = form_data['name'],
            sku = form_data['sku'],
            xuong = form_data['xuong'],
            img = encoded_string,
            description = form_data['description'],
            base_cost = float(form_data['base_cost']),
            thuy_cost = float(form_data['thuy_cost']),
            us_shipping_1st = float(form_data['us_shipping_1st']),
            us_shipping_additional = float(form_data['us_shipping_additional']),
            ww_shipping_1st = float(form_data['ww_shipping_1st']),
            ww_shipping_additional = float(form_data['ww_shipping_additional']),
            quantity = int(form_data['quantity']),
            product_template_width = int(form_data['width']),
            product_template_height = int(form_data['height']),
            product_size = form_data['size'],
            box_size = form_data['box_size'],
            product_type = form_data['type'],
            created_at = time_now,
            updated_at = time_now,
            user_logs = [f'{user["username"]} - {time_now} - update product {form_data["name"]}'],
        ).dict()
        update_ = await db_client.update_one(form_data['_id'], document, collection_name=DB['COL_PRODUCTS'])
        return RedirectResponse(url=current_url, status_code=302)
@authenticate
async def ff_add_product(request):
    data = {
        "meta_title": "Products",
        "factories": factories,
    }
    user = request.state.user_data
    current_url = str(request.url)  
    if request.method == "GET":
        return templates.TemplateResponse('fulfillment/add_product.html', {'request': request, 'data': data})
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        img = form_data['img'].file
        encoded_string = base64.b64encode(img.read()).decode('utf-8')
        time_now = convert_to_mongodb_time(time.time())
        document = Products(
            name = form_data['name'],
            sku = form_data['sku'],
            xuong = form_data['xuong'],
            img = encoded_string,
            description = form_data['description'],
            base_cost = float(form_data['base_cost']),
            thuy_cost = float(form_data['thuy_cost']),
            us_shipping_1st = float(form_data['us_shipping_1st']),
            us_shipping_additional = float(form_data['us_shipping_additional']),
            ww_shipping_1st = float(form_data['ww_shipping_1st']),
            ww_shipping_additional = float(form_data['ww_shipping_additional']),
            quantity = int(form_data['quantity']),
            product_template_width = int(form_data['width']),
            product_template_height = int(form_data['height']),
            box_size = form_data['box_size'],
            product_size = form_data['size'],
            product_type = form_data['type'],
            created_at = time_now,
            updated_at = time_now,
            user_logs = [f'{user["username"]} - {time_now} - insert new product {form_data["name"]}'],
        ).dict()
        update_ = await db_client.insert_one(document, collection_name=DB['COL_PRODUCTS'])


        return RedirectResponse(url=current_url, status_code=302)

#
# Fulfillment products
#
@authenticate
async def design_type(request):
    # Query
    db_client = request.state.db
    page_number = 1
    skip_ = (page_number - 1) * 200
    limit_ = 200
    # Get page
    page_number, limit_, skip_ = await get_page_data(request, limit_=limit_)
    
    # Query
    query = {}
    filter_ = {
        "name": 1,
        "description": 1,
        "price": 1,
    }
    design_types = await db_client.find_many_combo_sort_lasted(query, collection_name=DB['COL_DESIGN'], filter_=filter_, limit_=limit_, sort_="name", skip_=skip_)
        
    data = {
        "meta_title": "Design type",
        "design_types": design_types,
    }

    return templates.TemplateResponse('design_type/list_design_types.html', {'request': request, 'data': data})
@authenticate
async def edit_design_type(request):
    data = {}
    current_url = str(request.url)
    user = request.state.user_data
    db_client = request.state.db
    if request.method == "GET":
        _id = request.query_params.get('id')
        query = {"_id": ObjectId(_id)}
        design_type = await db_client.find_one(query, collection_name=DB['COL_DESIGN'])
        data = {
            "meta_title": "Design type",
            "design_type": design_type,
        }
        return templates.TemplateResponse('design_type/edit_design_type.html', {'request': request, 'data': data})
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        print(form_data)
        time_now = convert_to_mongodb_time(time.time())

        design_type = await db_client.find_one({'_id': ObjectId(form_data['_id'])}, collection_name=DB['COL_DESIGN'])


        document = Design(
            name = form_data['name'],
            description = form_data['description'],
            price = float(form_data['price']),
            user_logs = [f'{user["username"]} - {time_now} - update design type {form_data["name"]}'],
        ).dict()
        update_ = await db_client.update_one(form_data['_id'], document, collection_name=DB['COL_DESIGN'])
        return RedirectResponse(url=current_url, status_code=302)
@authenticate
async def add_design_type(request):
    data = {
        "meta_title": "Design type",
    }
    user = request.state.user_data
    current_url = str(request.url)  
    if request.method == "GET":
        return templates.TemplateResponse('design_type/add_design_type.html', {'request': request, 'data': data})
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        time_now = convert_to_mongodb_time(time.time())
        document = Design(
            name = form_data['name'],
            description = form_data['description'],
            price = float(form_data['price']),
            user_logs = [f'{user["username"]} - {time_now} - insert new design type {form_data["name"]}'],
        ).dict()
        update_ = await db_client.insert_one(document, collection_name=DB['COL_DESIGN'])

        return RedirectResponse(url=current_url, status_code=302)



#
# Fulfillment orders
#
def render_content_orders(request, context):
    # Rend with context
    template = templates.get_template('api_render/api_orders.html')
    html_content = template.render(request=request, context=context)
    return html_content

@authenticate
async def orders(request):
    db_client = request.state.db
    user = request.state.user_data
    team = user['team']
    
    try:
        table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
    except:
        table_column = {}
    
    # Process filter parameters
    params = await func.process_filter_params(request)

    # build query
    query = build_query_filter_orders(params)
    if team in ['nb_team']:
        query.update({
            "_original_data.line_items.wrong": {"$ne": "sku_wrong"}
        })


    # Get page
    page_number, limit_, skip_ = await get_page_data(request)

    # make page navi
    count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
    page_navi = make_page_navi(request, count_data, limit_, page_number, 'orders')

    # Process date range
    date_range = await func.process_date_range(request)
    
    # Retrieve orders
    orders = await func.retrieve_orders(db_client, query, page_number, limit_, skip_)

    
    # Format orders
    formatted_orders = await format_show_orders(orders)

    # Extract options for filters
    all_teams, shops = await func.extract_filter_options(formatted_orders)

    # find shops
    db_shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"shop": 1})
    shops += [shop['shop'] for shop in db_shops if shop['shop'] not in shops]

    # Get designer
    user_db = await db_client.find_many_combo({"team": team}, collection_name=DB['COL_USERS'], filter_={"username": 1, "groups": 1})
    list_designer = [user['username'] for user in user_db if 'designer' in user['groups']]

    # get all shops
    accounts = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_id": 1})
    sellers = []
    unique_sellers = []

    for account in accounts:
        if account['seller_id'] not in unique_sellers:
            unique_sellers.append(account['seller_id'])
            sellers.append(account['seller_id'])
    sellers = sorted(sellers)
    # Prepare data for response

    p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
    nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
    nb_product_types = sorted(set(nb_product_types))

    d_types = await db_client.find_many_combo({}, collection_name=DB['COL_DESIGN'], filter_={"name": 1})
    design_types = [d_type['name'] for d_type in d_types if d_type['name'] != '']
    design_types = sorted(set(design_types))

    if team in ['nb_team']:
        all_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1})
        all_teams = [team['name'] for team in all_teams if 'xuong' not in team['name']]

    if 'mod' in user['groups']:
        status_list = ["None", "Pending", "Ready", "Designing", "Design uploaded", "Sai design", "Rejected", "Support check", "Cancel", "Refund"]
    elif 'admin' in user['groups'] or 'dev' in user['groups']:
        status_list = list_status + ['Cancel', 'Refund']
    else:
        status_list = list_status
    
    if params['factory'] == ['', None] or params['status'] == ['None']:
        factory = ['None']
    else:
        factory = params['factory']
    if params['status'] == ['', None] or params['status'] == ['None']:
        status = ['None']
    else:
        status = params['status']
    if  'start_date' in params and params['start_date'] != None:
          date = params['start_date'].strftime("%Y/%m/%d") + ' - ' + params['end_date'].strftime("%Y/%m/%d")
    else:
        date = ''      
    paramurl = {
        'date' : date,
        'status': ','.join(status),
        'factory': ','.join(factory) ,
        'designer':  ','.join(params.get('designer',[])),
        'type':   ','.join(params.get('product_type',[])),
        'shop':   ','.join(params.get('shop',[])),
        'team': ','.join(params.get('filter_team',[])),
        'seller': ','.join(params.get('seller',[])) ,
        'shipping_day': '',
    }

    param_url = urllib.parse.urlencode(paramurl).replace('%5B%27%27%5D','').replace('%5B%5D','')
    if team in ['nb_team', 'vmm_team']:
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "nb_product_types": nb_product_types,
            "design_types": design_types,
            "table_column": table_column.get('table_column', []),
            "teams": {'options': all_teams, 'selected': params['filter_team']},
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "shops": {'options': shops, 'selected': params['shop']},
            "status": {"options": status_list, "selected": status},
            "type": {"options": nb_product_types, "selected": None},
            "factories": {"options": factories, "selected": factory},
            "designer": {'options': list_designer, 'selected': params['designer']},
            "seller": {'options': sellers, 'selected': params['seller']},
            "search": params['search'],
            "shipping_day": params['shipping_day'],
            "have_tracking": params.get('have_tracking', None),
            "tracking_synced": params.get('tracking_synced', None),
            "date_range": date_range,
            "page_navi": page_navi,
            'param_url':param_url,
        }
        template = 'fulfillment/list_orders.html'
    else:
        status_list = ["Ready", "Pending", "Half Done", "Doing", "Done recently", "In", "Out", "Sai design"]
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "table_column": table_column.get('table_column', []),
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "status": {"options": status_list, "selected": params['status']},
            "factories": {"options": factories, "selected": params['factory']},
            "search": params['search'],
            "date_range": date_range,
            "page_navi": page_navi,
        }
        template = 'other_team/list_orders.html'

    return templates.TemplateResponse(template, {'request': request, 'data': data})
@authenticate
async def ff_ngoai(request):
    db_client = request.state.db
    user = request.state.user_data
    team = user['team']
    if request.method == "GET":
        # Process filter parameters
        params = await func.process_filter_params(request)

        # build query
        query = build_query_filter_orders(params, ff_ngoai=True)

        query.update({
            "_original_data.line_items.wrong": {"$ne": "sku_wrong"},
            "_team": "nb_team"
        })
        
        # Get page
        page_number, limit_, skip_ = await get_page_data(request)

        # make page navi
        count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
        page_navi = make_page_navi(request, count_data, limit_, page_number, 'orders')

        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None, limit_=limit_, skip_=skip_, sort_="_id")

        formatted_orders = await format_show_orders(orders)

        # Extract options for filters
        all_teams, shops = await func.extract_filter_options(formatted_orders)

        # find shops
        db_shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"shop": 1})
        shops += [shop['shop'] for shop in db_shops if shop['shop'] not in shops]

        # Get designer
        user_db = await db_client.find_many_combo({"team": team}, collection_name=DB['COL_USERS'], filter_={"username": 1, "groups": 1})
        list_designer = [user['username'] for user in user_db if 'designer' in user['groups']]

        try:
            table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
        except:
            table_column = {} 
        
        # Process date range
        date_range = await func.process_date_range(request)

         # get all shops
        accounts = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_id": 1})
        sellers = []
        unique_sellers = []

        for account in accounts:
            if account['seller_id'] not in unique_sellers:
                unique_sellers.append(account['seller_id'])
                sellers.append(account['seller_id'])
        sellers = sorted(sellers)

        p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
        nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
        nb_product_types = sorted(set(nb_product_types))

        d_types = await db_client.find_many_combo({}, collection_name=DB['COL_DESIGN'], filter_={"name": 1})
        design_types = [d_type['name'] for d_type in d_types if d_type['name'] != '']
        design_types = sorted(set(design_types))

        all_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1})
        all_teams = [team['name'] for team in all_teams if 'xuong' not in team['name']]

        if params['factory'] == ['', None] or params['status'] == ['None']:
            factory = ['None']
        else:
            factory = params['factory']
        if 'start_date' in params and params['start_date'] != None:
            date = params['start_date'].strftime("%Y/%m/%d") + ' - ' + params['end_date'].strftime("%Y/%m/%d")
        else:
            date = ''
        paramurl = {
            'date' : date,
            'factory': ','.join(factory) ,
            'designer':  ','.join(params.get('designer',[])),
            'type':   ','.join(params.get('product_type',[])),
            'shop':   ','.join(params.get('shop',[])),
            'team': ','.join(params.get('filter_team',[])),
            'seller': ','.join(params.get('seller',[])) ,
            'shipping_day': '',
        }

        param_url = urllib.parse.urlencode(paramurl).replace('%5B%27%27%5D','').replace('%5B%5D','')
            
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "nb_product_types": nb_product_types,
            "design_types": design_types,
            "table_column": table_column.get('table_column', []),
            "teams": {'options': all_teams, 'selected': params['filter_team']},
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "shops": {'options': shops, 'selected': params['shop']},
            "status": {"options": ['FF Ngoai'], "selected": "FF Ngoai"},
            "type": {"options": nb_product_types, "selected": None},
            "factories": {"options": factories, "selected": factory},
            "designer": {'options': list_designer, 'selected': params['designer']},
            "seller": {'options': sellers, 'selected': params['seller']},
            "search": params['search'],
            "shipping_day": params['shipping_day'],
            "have_tracking": params.get('have_tracking', None),
            "tracking_synced": params.get('tracking_synced', None),
            "date_range": date_range,
            "page_navi": page_navi,
            "is_ff_ngoai": True,
            "param_url": param_url,
        }
        return templates.TemplateResponse('fulfillment/list_orders.html', {'request': request, 'data': data})  
@authenticate
async def wrong_sku(request):
    db_client = request.state.db
    user = request.state.user_data
    team = 'nb_team'
    if request.method == "GET":
        # Process filter parameters
        params = await func.process_filter_params(request)

        # build query
        query = build_query_filter_orders(params, ff_ngoai=None)

        query.update({
            "_original_data.line_items.wrong": "sku_wrong",
            "_original_data.line_items.product_type": {"$ne": "AF"}
        })

        # Get page
        page_number, limit_, skip_ = await get_page_data(request)

        # make page navi
        count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
        page_navi = make_page_navi(request, count_data, limit_, page_number, 'orders')

        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None, limit_=limit_, skip_=skip_, sort_="create_at")

        formatted_orders = await format_show_orders(orders)

        # Extract options for filters
        all_teams, shops = await func.extract_filter_options(formatted_orders)

        # find shops
        db_shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"shop": 1})
        shops += [shop['shop'] for shop in db_shops if shop['shop'] not in shops]

        # Get designer
        user_db = await db_client.find_many_combo({"team": team}, collection_name=DB['COL_USERS'], filter_={"username": 1, "groups": 1})
        list_designer = [user['username'] for user in user_db if 'designer' in user['groups']]

        try:
            table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
        except:
            table_column = {} 
        
        # Process date range
        date_range = await func.process_date_range(request)

         # get all shops
        accounts = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_id": 1})
        sellers = []
        unique_sellers = []

        for account in accounts:
            if account['seller_id'] not in unique_sellers:
                unique_sellers.append(account['seller_id'])
                sellers.append(account['seller_id'])
        sellers = sorted(sellers)

        p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
        nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
        nb_product_types = sorted(set(nb_product_types))

        d_types = await db_client.find_many_combo({}, collection_name=DB['COL_DESIGN'], filter_={"name": 1})
        design_types = [d_type['name'] for d_type in d_types if d_type['name'] != '']
        design_types = sorted(set(design_types))

        all_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1})
        all_teams = [team['name'] for team in all_teams if 'xuong' not in team['name']]
        print(params['status'])
        if  params['factory'] == ['', None] or params['status'] == ['None']:
            factory = ['None']
        else:
            factory = params['factory']
        if params['status'] == ['', None] or params['status'] == ['None']:
            status = ['None']
        else:
            status = params['status']
        if  'start_date' in params and params['start_date'] != None:
          date = params['start_date'].strftime("%Y/%m/%d") + ' - ' + params['end_date'].strftime("%Y/%m/%d")
        else:
            date = ''      
        
        paramurl = {
            'date' : date,
            'status': ','.join(status),
            'factory': ','.join(factory) ,
            'designer':  ','.join(params.get('designer',[])),
            'type':   ','.join(params.get('product_type',[])),
            'shop':   ','.join(params.get('shop',[])),
            'team': ','.join(params.get('filter_team',[])),
            'seller': ','.join(params.get('seller',[])) ,
            'shipping_day': '',
        }

        param_url = urllib.parse.urlencode(paramurl).replace('%5B%27%27%5D','').replace('%5B%5D','')
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "nb_product_types": nb_product_types,
            "design_types": design_types,
            "table_column": table_column.get('table_column', []),
            "teams": {'options': all_teams, 'selected': params['filter_team']},
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "shops": {'options': shops, 'selected': params['shop']},
            "status": {"options": list_status, "selected": status},
            "factories": {"options": factories, "selected": factory},
            "type": {"options": nb_product_types, "selected": None},
            "designer": {'options': list_designer, 'selected': params['designer']},
            "seller": {'options': sellers, 'selected': params['seller']},
            "search": params['search'],
            "shipping_day": params['shipping_day'],
            "have_tracking": params.get('have_tracking', None),
            "tracking_synced": params.get('tracking_synced', None),
            "date_range": date_range,
            "page_navi": page_navi,
            'param_url':param_url,
        }
        return templates.TemplateResponse('fulfillment/list_orders.html', {'request': request, 'data': data})  
@authenticate
async def wrong_product_type(request):
    db_client = request.state.db
    user = request.state.user_data
    team = 'nb_team'
    if request.method == "GET":
        # Process filter parameters
        params = await func.process_filter_params(request)

        # build query
        query = build_query_filter_orders(params, ff_ngoai=None)

        product_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
        product_types = [product_type['product_type'] for product_type in product_types if product_type['product_type'] != '']

        query.update({
            "_original_data.line_items": {
                "$elemMatch": {
                    "product_type": {
                        "$nin": product_types
                    }
                }
            }
        })
        print(query)
        # Get page
        page_number, limit_, skip_ = await get_page_data(request)

        # make page navi
        count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
        page_navi = make_page_navi(request, count_data, limit_, page_number, 'orders')

        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None, limit_=limit_, skip_=skip_, sort_="create_at")

        formatted_orders = await format_show_orders(orders)

        # Extract options for filters
        all_teams, shops = await func.extract_filter_options(formatted_orders)

        # find shops
        db_shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"shop": 1})
        shops += [shop['shop'] for shop in db_shops if shop['shop'] not in shops]

        # Get designer
        user_db = await db_client.find_many_combo({"team": team}, collection_name=DB['COL_USERS'], filter_={"username": 1, "groups": 1})
        list_designer = [user['username'] for user in user_db if 'designer' in user['groups']]

        try:
            table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
        except:
            table_column = {} 
        
        # Process date range
        date_range = await func.process_date_range(request)

         # get all shops
        accounts = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_id": 1})
        sellers = []
        unique_sellers = []

        for account in accounts:
            if account['seller_id'] not in unique_sellers:
                unique_sellers.append(account['seller_id'])
                sellers.append(account['seller_id'])
        sellers = sorted(sellers)

        p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
        nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
        nb_product_types = sorted(set(nb_product_types))

        d_types = await db_client.find_many_combo({}, collection_name=DB['COL_DESIGN'], filter_={"name": 1})
        design_types = [d_type['name'] for d_type in d_types if d_type['name'] != '']
        design_types = sorted(set(design_types))


        all_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1})
        all_teams = [team['name'] for team in all_teams if 'xuong' not in team['name']]

        if params['factory'] == ['', None]:
            factory = ['None']
        else:
            factory = params['factory']
        if params['status'] == ['', None]:
            status = ['None']
        else:
            status = params['status']
        if  'start_date' in params and params['start_date'] != None:
          date = params['start_date'].strftime("%Y/%m/%d") + ' - ' + params['end_date'].strftime("%Y/%m/%d")
        else:
            date = ''      
        paramurl = {
            'date' : date,
            'status': ','.join(params.get('status',[])),
            'factory': ','.join(params.get('factory',[])) ,
            'designer':  ','.join(params.get('designer',[])),
            'type':   ','.join(params.get('product_type',[])),
            'shop':   ','.join(params.get('shop',[])),
            'team': ','.join(params.get('filter_team',[])),
            'seller': ','.join(params.get('seller',[])) ,
            'shipping_day': '',
        }

        param_url = urllib.parse.urlencode(paramurl).replace('%5B%27%27%5D','').replace('%5B%5D','')
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "nb_product_types": nb_product_types,
            "design_types": design_types,
            "table_column": table_column.get('table_column', []),
            "teams": {'options': all_teams, 'selected': params['filter_team']},
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "shops": {'options': shops, 'selected': params['shop']},
            "status": {"options": list_status, "selected": status},
            "factories": {"options": factories, "selected": factory},
            "type": {"options": nb_product_types, "selected": None},
            "designer": {'options': list_designer, 'selected': params['designer']},
            "seller": {'options': sellers, 'selected': params['seller']},
            "search": params['search'],
            "shipping_day": params['shipping_day'],
            "have_tracking": params.get('have_tracking', None),
            "tracking_synced": params.get('tracking_synced', None),
            "date_range": date_range,
            "page_navi": page_navi,
            'param_url':param_url,
        }
        return templates.TemplateResponse('fulfillment/list_orders.html', {'request': request, 'data': data})  
@authenticate
async def check_design(request):
    db_client = request.state.db
    user = request.state.user_data
    try:
        seller_id = user['special_id']
    except:
        seller_id = ''
    team = user['team']
    
    try:
        table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
    except:
        table_column = {}
    
    # Process filter parameters
    params = await func.process_filter_params(request)

    # build query
    query = build_query_filter_orders(params)
    query.update({
        "_original_data.line_items.wrong": {"$ne": "sku_wrong"},
        "_original_data.line_items.fulfillment_status": {"$in": ["Design uploaded", "Sai design"]},
        "_original_data.seller_id": seller_id
    })
    print(query)
    # Get page
    page_number, limit_, skip_ = await get_page_data(request)

    # make page navi
    count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
    page_navi = make_page_navi(request, count_data, limit_, page_number, 'check_design')

    # Process date range
    date_range = await func.process_date_range(request)
    
    # Retrieve orders
    orders = await func.retrieve_orders(db_client, query, page_number, limit_, skip_)

    
    # Format orders
    formatted_orders = await format_show_orders(orders)

    # Extract options for filters
    all_teams, shops = await func.extract_filter_options(formatted_orders)

    # find shops
    db_shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"shop": 1})
    shops += [shop['shop'] for shop in db_shops if shop['shop'] not in shops]

    # Get designer
    user_db = await db_client.find_many_combo({"team": team}, collection_name=DB['COL_USERS'], filter_={"username": 1, "groups": 1})
    list_designer = [user['username'] for user in user_db if 'designer' in user['groups']]

    # get all shops
    accounts = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_id": 1})
    sellers = []
    unique_sellers = []

    for account in accounts:
        if account['seller_id'] not in unique_sellers:
            unique_sellers.append(account['seller_id'])
            sellers.append(account['seller_id'])
    sellers = sorted(sellers)
    # Prepare data for response

    p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
    nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
    nb_product_types = sorted(set(nb_product_types))

    d_types = await db_client.find_many_combo({}, collection_name=DB['COL_DESIGN'], filter_={"name": 1})
    design_types = [d_type['name'] for d_type in d_types if d_type['name'] != '']
    design_types = sorted(set(design_types))


    if team in ['nb_team']:
        all_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1})
        all_teams = [team['name'] for team in all_teams if 'xuong' not in team['name']]

    status_list = ["Design uploaded", "Sai design"]
    
    if params['factory'] == ['', None] or params['status'] == ['None']:
        factory = ['None']
    else:
        factory = params['factory']
    if params['status'] == ['', None] or params['status'] == ['None']:
        status = ['None']
    else:
        status = params['status']
    if  'start_date' in params and params['start_date'] != None:
          date = params['start_date'].strftime("%Y/%m/%d") + ' - ' + params['end_date'].strftime("%Y/%m/%d")
    else:
        date = ''      
        
    paramurl = {
        'date' : date,
        'status': ','.join(status),
        'factory': ','.join(factory) ,
        'designer':  ','.join(params.get('designer',[])),
        'type':   ','.join(params.get('product_type',[])),
        'shop':   ','.join(params.get('shop',[])),
        'team': ','.join(params.get('filter_team',[])),
        'seller': ','.join(params.get('seller',[])) ,
        'shipping_day': '',
    }

    param_url = urllib.parse.urlencode(paramurl).replace('%5B%27%27%5D','').replace('%5B%5D','')
    if team in ['nb_team', 'vmm_team']:
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "nb_product_types": nb_product_types,
            "design_types": design_types,
            "table_column": table_column.get('table_column', []),
            "teams": {'options': all_teams, 'selected': params['filter_team']},
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "shops": {'options': shops, 'selected': params['shop']},
            "status": {"options": status_list, "selected": status},
            "type": {"options": nb_product_types, "selected": None},
            "factories": {"options": factories, "selected": factory},
            "designer": {'options': list_designer, 'selected': params['designer']},
            "seller": {'options': sellers, 'selected': params['seller']},
            "search": params['search'],
            "shipping_day": params['shipping_day'],
            "have_tracking": params.get('have_tracking', None),
            "tracking_synced": params.get('tracking_synced', None),
            "date_range": date_range,
            "page_navi": page_navi,
            'param_url':param_url,
        }
        template = 'fulfillment/list_orders.html'
    else:
        status_list = ["Ready", "Pending", "Half Done", "Doing", "Done recently", "In", "Out", "Sai design"]
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "table_column": table_column.get('table_column', []),
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "status": {"options": status_list, "selected": params['status']},
            "factories": {"options": factories, "selected": params['factory']},
            "search": params['search'],
            "date_range": date_range,
            "page_navi": page_navi,
        }
        template = 'other_team/list_orders.html'

    return templates.TemplateResponse(template, {'request': request, 'data': data})
@authenticate
async def order_assigned(request):
    db_client = request.state.db
    user = request.state.user_data
    user_name = user['username']
    team = user['team']
    if request.method == "GET":
        # Process filter parameters
        params = await func.process_filter_params(request)

        # build query
        query = build_query_filter_orders(params)
        
        query.update({
            "_original_data.line_items.assign_designer.designer": user['username'],
            "_original_data.line_items.assign_designer.design_checked": False,
            "_original_data.line_items.wrong": {"$ne": "sku_wrong"}
        })

        # Get page
        page_number, limit_, skip_ = await get_page_data(request)

        # make page navi
        count_data = await db_client.count_(query=query, collection_name=DB['COL_FULFILLMENTS'])
        page_navi = make_page_navi(request, count_data, limit_, page_number, 'orders')

        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None, limit_=limit_, skip_=skip_, sort_="_id")

        formatted_orders = await format_show_orders(orders)

        # Extract options for filters
        all_teams, shops = await func.extract_filter_options(formatted_orders)

        # find shops
        db_shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"shop": 1})
        shops += [shop['shop'] for shop in db_shops if shop['shop'] not in shops]

        try:
            table_column = await db_client.find_one({"username": user['user_id']}, collection_name=DB['COL_USERS'], filter_={"table_column": 1})
        except:
            table_column = {} 
        
         # Process date range
        date_range = await func.process_date_range(request)

        if 'designer' in user['groups'] or 'mod' in user['groups']:
            status_list = ["None", "Pending", "Ready", "Designing", "Design uploaded", "Sai design", "Rejected", "Support check", "Cancel", "Refund"]
        elif 'admin' in user['groups'] or 'dev' in user['groups']:
            status_list = list_status + ['Cancel', 'Refund']
        else:
            status_list = list_status
        # get all shops
        accounts = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_id": 1})
        sellers = []
        unique_sellers = []

        for account in accounts:
            if account['seller_id'] not in unique_sellers:
                unique_sellers.append(account['seller_id'])
                sellers.append(account['seller_id'])
        sellers = sorted(sellers)
        
        p_types = await db_client.find_many_combo({}, collection_name=DB['COL_PRODUCTS'], filter_={"product_type": 1})
        nb_product_types = [p_type['product_type'] for p_type in p_types if p_type['product_type'] != '']
        nb_product_types = sorted(set(nb_product_types))

        d_types = await db_client.find_many_combo({}, collection_name=DB['COL_DESIGN'], filter_={"name": 1})
        design_types = [d_type['name'] for d_type in d_types if d_type['name'] != '']
        design_types = sorted(set(design_types))


        all_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1})
        all_teams = [team['name'] for team in all_teams if 'xuong' not in team['name']]
        data = {
            "meta_title": "Orders",
            "orders": formatted_orders,
            "user": user,
            "nb_product_types": nb_product_types,
            "design_types": design_types,
            "table_column": table_column.get('table_column', []),
            "teams": {'options': all_teams, 'selected': params['filter_team']},
            "product_types": {'options': nb_product_types, 'selected': params['product_type']},
            "shops": {'options': shops, 'selected': params['shop']},
            "status": {"options": status_list, "selected": params['status']},
            "type": {"options": nb_product_types, "selected": None},
            "factories": {"options": factories, "selected": params['factory']},
            "designer": {'options': [user_name], 'selected': params['designer']},
            "seller": {'options': sellers, 'selected': params['seller']},
            "search": params['search'],
            "shipping_day": params['shipping_day'],
            "have_tracking": params.get('have_tracking', None),
            "tracking_synced": params.get('tracking_synced', None),
            "date_range": date_range,
            "page_navi": page_navi,
        }
        return templates.TemplateResponse('fulfillment/list_orders.html', {'request': request, 'data': data})
@authenticate
async def edit_order(request):
    try:
        db_client = request.state.db
        user = request.state.user_data
        team = user['team']

        if request.method == "GET":
            order_id = request.query_params.get('external_id')
            line_item_id = request.query_params.get('line_item_id')
            
            query = {"_original_data.order_id": str(order_id)}

            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])
            product = next((line_item for line_item in order['_original_data'].get('line_items', []) if line_item['line_item_id'] == line_item_id), None)
            
            data = {
                "meta_title": "Orders",
                "order": order,
                "product": product,
            }
            if team == 'nb_team':
                return templates.TemplateResponse('fulfillment/edit_order.html', {'request': request, 'data': data})
            else:
                return templates.TemplateResponse('other_team/edit_order.html', {'request': request, 'data': data})

        elif request.method == "POST":
            form_data = await request.form()
            _id = form_data['_id']
            line_item_id = form_data['line_item_id']
            quantity = int(form_data['quantity'])
            note = form_data['note']
            sku = form_data['sku']
            seller = form_data['seller']
            base_cost = float(form_data['base_cost'])
            link_des = form_data['link_des']
            product_type = form_data['product_type']
            size = form_data['size']
            material = form_data['material']
            color = form_data['color']
            personalization = form_data['personalization']
            num_color_kit = form_data['num_color_kit']
            num_holder = form_data['num_holder']

            if '[' in link_des:
                link_des = [url.strip() for url in link_des.replace("[", "").replace("]", "").replace("\\", "").replace("'", "").replace('"', "").split(",")]
            
            if link_des == ['']:
                link_des = []

            query = {"_id": ObjectId(_id)}
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])
            
            sku_format = sku.split("-")
            

            for line_item in order['_original_data'].get('line_items', []):
                if line_item['line_item_id'] == line_item_id:
                    sku_format = sku.split("-")
                    if order['_team'] == 'nb_team':
                        print(len(sku_format) in [5, 6], sku_format[1].lower() == 'et', 'd' not in sku_format[4])
                        if len(sku_format) in [5, 6] and sku_format[1].lower() == 'et' and 'd' not in sku_format[4]:
                            wrong = ''
                        else:
                            wrong = 'sku_wrong'
                    else:
                        wrong = ''
                    if product_type in ['AF']:
                        wrong = ''
                    line_item.update({
                        'note': note,
                        'sku': sku,
                        'seller': seller,
                        'base_cost': base_cost,
                        'quantity': quantity,
                        'link_des': link_des,
                        'product_type': product_type,
                        'size': size,
                        'material': material,
                        'color': color,
                        'wrong': wrong,
                        'personalization': personalization,
                        'num_color_kit': num_color_kit,
                        'num_holder': num_holder,
                        'wrong': wrong,
                    })

                    user_logs = line_item.get('user_logs', [])
                    user_logs.append(f"{request.state.user_data['username']} edit data to {line_item} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    line_item['user_logs'] = user_logs
                    break
            order['_original_data']['seller_id'] = seller
            await db_client.update_one(_id, order, collection_name=DB['COL_FULFILLMENTS'])
            return RedirectResponse(url='/orders', status_code=302)
    
    except Exception as e:
        return JSONResponse({'status': 400, 'message': f"Edit order failed: {str(e)}"})      
@authenticate
async def edit_address(request):
    try:
        db_client = request.state.db
        user = request.state.user_data
        team = user['team']

        if request.method == "GET":
            order_id = request.query_params.get('external_id')
            
            query = {"_original_data.order_id": str(order_id)}

            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'], filter_={"_original_data.delivery_info": 1, "_original_data.order_id": 1})
            data = {
                "meta_title": "Orders",
                "order": order,
            }
            if team == 'nb_team':
                return templates.TemplateResponse('fulfillment/edit_address.html', {'request': request, 'data': data})
            else:
                return templates.TemplateResponse('other_team/edit_address.html', {'request': request, 'data': data})

        elif request.method == "POST":
            form_data = await request.form()

            _id = form_data['_id']
            query = {"_id": ObjectId(_id)}
            order = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])

            order['_original_data']['delivery_info']['name'] = form_data['fullname']
            order['_original_data']['delivery_info']['phone'] = form_data['phone']
            order['_original_data']['delivery_info']['address1'] = form_data['address1']
            order['_original_data']['delivery_info']['address2'] = form_data['address2']
            order['_original_data']['delivery_info']['city'] = form_data['city']
            order['_original_data']['delivery_info']['state'] = form_data['state']
            order['_original_data']['delivery_info']['zip'] = form_data['zipcode']
            order['_original_data']['delivery_info']['country'] = form_data['country']
    
            await db_client.update_one(_id, order, collection_name=DB['COL_FULFILLMENTS'])

            return RedirectResponse(url='/orders', status_code=302)
    
    except Exception as e:
        return JSONResponse({'status': 400, 'message': f"Edit order address failed: {str(e)}"})    


async def fulfillment_page(request):
    db_client = request.state.db
    order_id = request.query_params.get('order_id')
    provider = request.query_params.get('provider')

    query = {
        "_original_data.order_id" : order_id
    }
    orders = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'], filter_=None)
    pack = 1
    for item in orders['_original_data']['line_items']:
        other_option = item.get('other_option', '').lower()
        if 'pack' in other_option:
            for number in range(1, 21):
                pack_str = f'pack {number}'
                str_pack = f'{number} pack'
                if pack_str in other_option or str_pack in other_option and(f'pack {number}0' not in other_option and f'{number}0 pack'  not in other_option):
                    pack = number
                    break        
            
        quantity = int(item['quantity']) * pack
        item['quantity'] = quantity
    delivery_info = orders['_original_data']['delivery_info']
    if (provider == 'printify'):
        print_providers = await db_client.find_many_combo({}, collection_name=DB['COL_FULFILL_NGOAI'], filter_={'name':1, '_id':1})
    elif(provider == 'merchize'):
        print_providers = await db_client.find_many_combo({'id': 1000000}, collection_name=DB['COL_FULFILL_NGOAI'], filter_={'name':1, '_id':1,'products.id':1,'products.name':1})
    elif(provider == 'gearment'):
        print_providers = await db_client.find_many_combo({'id': 1312832832}, collection_name=DB['COL_FULFILL_NGOAI'], filter_={'name':1, '_id':1,'products.id':1,'products.name':1})
    data = {
        'delivery_info' : delivery_info,
        'line_items' : orders['_original_data']['line_items'],
        'print_providers' : print_providers,
        'provider' : provider,
        'team' : orders['_team'],
        'order_date': orders['create_at'].strftime("%Y/%m/%d"),
        'shop': orders['_original_data']['shop_id'],
    }
    if provider == 'printify':
        return templates.TemplateResponse('ff_ngoai/fulfillment_printify.html', {'request': request, 'data': data})
    elif provider == 'merchize':
        return templates.TemplateResponse('ff_ngoai/fulfillment_merchize.html', {'request': request, 'data': data})
    elif provider == 'gearment':
        return templates.TemplateResponse('ff_ngoai/fulfillment_gearment.html', {'request': request, 'data': data})

# Teams
async def get_users_team(team):
    if team.get('users', False):
        ids  = [ObjectId(user) for user in team['users']]
        team['users'] = await db_get_users_by_ids( ids)
    return team    
@authenticate
async def teams(request):
    # Query
    db_client = request.state.db
    limit_ = 100
    # Get page
    page_number, limit_, skip_ = await get_page_data(request, limit_=limit_)
    skip_ = (page_number - 1) * 100
    
    # Query
    db_client = request.state.db
    query = {}
    filter_ = {
        "name": 1,
        "users": 1,
    }
    teams = await db_client.find_many_combo(query, collection_name=DB['COL_TEAMS'], filter_=filter_, limit_=limit_, sort_="_id", skip_=skip_)
    teams = await asyncio.gather(*(get_users_team(team) for team in teams))
    data = {
        "meta_title": "Teams",
        "teams": teams,
    }

    return templates.TemplateResponse('teams/list_teams.html', {'request': request, 'data': data})
async def edit_users_team(db_client, user, old_user_in_team, name):
    if user not in old_user_in_team:
        await db_client.update_one(user, {"team": None}, collection_name=DB['COL_USERS'])
    else:
        await db_client.update_one(user, {"team": name}, collection_name=DB['COL_USERS'])
@authenticate
async def edit_team(request):
    data = {}
    db_client = request.state.db
    if request.method == "GET":
        _id = request.query_params.get('id')
        query = {"_id": ObjectId(_id)}
        users_not_in_team = await db_get_users_not_in_team()
        users = [{'name': user['username'], '_id': user['_id']}  for user in users_not_in_team]
        team = await db_client.find_one(query, collection_name=DB['COL_TEAMS'])
        if team.get('users', False):
            ids  = [ObjectId(user) for user in team['users']]
            team['users'] = await db_get_users_by_ids( ids)
        # col users
        data = {
            "meta_title": "Teams",
            "_id": _id,
            "team": team,
            "users": users,
        }

        return templates.TemplateResponse('teams/edit_team.html', {'request': request, 'data': data})
    
    elif request.method == "POST":
        form_data = await request.form()
        user_update = form_data.getlist('users')
        name = form_data['name']
        _id = form_data['_id']

        # get old user in team
        find_team = await db_client.find_one({"name": name}, collection_name=DB['COL_TEAMS'], filter_={"users": 1, "balance": 1})
        old_user_in_team = find_team['users']

        update_user = await asyncio.gather(*(edit_users_team(db_client, user, old_user_in_team, name) for user in user_update))                

        document = Teams(
            name=name,
            users=user_update,
            balance=find_team['balance']
        ).dict()
        update_ = await db_client.update_one(_id, document, collection_name=DB['COL_TEAMS'])

        
        return RedirectResponse(url='/teams', status_code=302)
async def update_user_teams(db_client,form_data,user):
    await db_client.update_one(user, {"team": form_data['name']}, collection_name=DB['COL_USERS'])
@authenticate
async def add_team(request):
    if request.method == "GET":
        users_not_in_team = await db_get_users_not_in_team()
        users = [{'name':user['username'],'_id':user['_id']}  for user in users_not_in_team]
        return templates.TemplateResponse('teams/add_team.html', {'request': request, 'data': {"meta_title": "Teams", 'users': users}})
    
    elif request.method == "POST":
        db_client = request.state.db
        form_data = await request.form()
        users = form_data.getlist('users')
        update_user = await asyncio.gather(*(update_user_teams(db_client,form_data,user) for user in users))
     
        document = Teams(
            name = form_data['name'], 
            users = users,
            balance = 0,
        ).dict()
        update_ = await db_client.insert_one(document, collection_name=DB['COL_TEAMS'])
        return RedirectResponse(url='/teams', status_code=302)    
@authenticate
async def wallet(request):      
    db_client = request.state.db
    user_id = request.state.user_data['_id']
    team_of_user = await db_client.find_one({"users": str(user_id)}, collection_name=DB['COL_TEAMS'])
    if request.method == "GET":
        if team_of_user == None:
            return templates.TemplateResponse('wallet/wallet.html', {'request': request, 'data': {}})
        query = {"team": team_of_user['name']}
        invoice = await db_client.find_many_combo(query, collection_name=DB['COL_INVOICE'], filter_=None, sort_="_id")
        data = {
            "meta_title": "Wallet",
            "team": team_of_user,
            "invoice": invoice,
        }
        return templates.TemplateResponse('wallet/wallet.html', {'request': request, 'data': data})
async def check_topup(request):
    # Query
    db_client = request.state.db
    
    # Query
    query = {
    }

    filter_ = {
        "platform": 1,
        "img": 1,
        "team": 1,
        "total": 1,
        "transaction": 1,
        "currency": 1,
        "created_at": 1,
        "status": 1,
    }
    invoices = await db_client.find_many_combo(query, collection_name=DB['COL_INVOICE'], filter_=filter_, limit_=0, sort_="_id")

    # Get rate
    rate = await db_client.find_one({"currency": "VND"}, collection_name=DB['COL_RATE'], filter_={"rate": 1})

        
    data = {
        "meta_title": "Check topup",
        "invoices": invoices,
        "rate": rate['rate']
    }

    return templates.TemplateResponse('wallet/check_topup.html', {'request': request, 'data': data})


# Shops
@authenticate
async def shops(request):
    db_client = request.state.db
    if request.method == "GET":
        shops = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_=None, sort_="_id")
        
        data = {
            "meta_title": "Shops",
            "shops": shops,
        }
        return templates.TemplateResponse('shops/list_shops.html', {'request': request, 'data': data})

async def edit_shop(request):
    db_client = request.state.db
    if request.method == "GET":
        _id = request.query_params.get('id')
        query = {"_id": ObjectId(_id)}
        filter_ = {
            'shop': 1,
            'seller_id': 1,
            'seller_name': 1,
            'proxy': 1,
            'team': 1,
        }
        shop = await db_client.find_one(query, collection_name=DB['COL_SHOP'], filter_=filter_)
        data = {
            "meta_title": "Shops",
            "_id": _id,
            "shop": shop,
        }
        return templates.TemplateResponse('shops/edit_shop.html', {'request': request, 'data': data})
    elif request.method == "POST":
        form_data = await request.form()
        _id = form_data['_id']
        shop = form_data['shop']
        seller_id = form_data['seller_id']
        seller_name = form_data['seller_name']
        proxy = form_data['proxy']
        team = form_data['team']

        document = {
            "shop": shop,
            "seller_id": seller_id,
            "seller_name": seller_name,
            "proxy": proxy,
            "team": team
        }
        update_ = await db_client.update_one(ObjectId(_id), document, collection_name=DB['COL_SHOP'])
        return RedirectResponse(url='/shops', status_code=302)

async def add_shop(request):
    db_client = request.state.db
    if request.method == "GET":
        return templates.TemplateResponse('shops/add_shop.html', {'request': request, 'data': {"meta_title": "Shops"}})
    elif request.method == "POST":
        form_data = await request.form()
        shop = form_data['shop']
        seller_id = form_data['seller_id']
        seller_name = form_data['seller_name']
        proxy = form_data['proxy']
        team = form_data['team']

        document = {
            "shop": shop,
            "seller_id": seller_id,
            "seller_name": seller_name,
            "proxy": proxy,
            "team": team,
        }
        update_ = await db_client.insert_one(document, collection_name=DB['COL_SHOP'])
        return RedirectResponse(url='/shops', status_code=302)


# import
@authenticate
async def import_orders(request):
    # import data from csv file
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        query = {}
        filter_ = {
            "name": 1,
        }
        if user['team'] == 'nb_team':
            teams = await db_client.find_many_combo(query, collection_name=DB['COL_TEAMS'], filter_=filter_, sort_="_id")
        else:
            teams = [{
                'name': user['team'],
            }]
        data = {
            "meta_title": "Import orders",
            "teams": teams,
        }
        return templates.TemplateResponse('fulfillment/import_orders.html', {'request': request, 'data': data})
    elif request.method == "POST":
        form = await request.form()
        file = form["file"]
        content_type = file.headers["content-type"]

        user = request.state.user_data
        # team_name = await db_client.find_one({"username": {"$in": [user['username']]}}, collection_name=DB['COL_TEAMS'])
        # if file is csv 
        team = form['team']
        try:
            if content_type == "text/csv":
                df = pd.read_csv(file.file, encoding='utf-8')
            # if file is xlsx
            elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                df = pd.read_excel(file.file)
            else:
                return PlainTextResponse("File type is not support!", status_code=400)
        except:
            return JSONResponse({"status": "error", "message": f"Import orders error - {traceback.format_exc()}!"})
        
        df_json_string = df.to_json(orient='records', force_ascii=False)
        df_json = json.loads(df_json_string)
        # if df is empty
        if df.empty:
            return PlainTextResponse("File is empty!", status_code=400)
        
        count = 0
        try:
            insert_results = []

            for order in df_json:
                if order['Order id'] is not None:
                    insert_result = await insert_order_db(order, user, db_client, team)
                    insert_results.append(insert_result)

            success_inserts = sum(1 for result in insert_results if result is True)
            failed_inserts = [error for error in insert_results if error is not True]

            if all(result for result in insert_results) and success_inserts == len(insert_results):
                return JSONResponse({"status": "success", "message": f"Import {success_inserts} items successfully!"})
            else:
                failed_count = len(failed_inserts)
                return JSONResponse({"status": "error", "message": f"Import {success_inserts} items successfully! -- Import {failed_count} orders error! - {failed_inserts}"})

        except:
            return JSONResponse({"status": "error", "message": f"Import orders error - {traceback.format_exc()}!"})

# @authenticate
# async def import_labels(request):
#     # import data from csv file
#     db_client = request.state.db
#     user = request.state.user_data
#     if request.method == "GET":
#         query = {}
#         filter_ = {
#             "name": 1,
#         }
#         if user['team'] == 'nb_team':
#             teams = await db_client.find_many_combo(query, collection_name=DB['COL_TEAMS'], filter_=filter_, sort_="_id")
#         else:
#             teams = [{
#                 'name': user['team'],
#             }]
#         data = {
#             "meta_title": "Import labels",
#             "teams": teams,
#         }
#         return templates.TemplateResponse('import/import_labels.html', {'request': request, 'data': data})
#     elif request.method == "POST":
#         form = await request.form()
#         file = form["file"]
#         content_type = file.headers["content-type"]

#         user = request.state.user_data

#         # if file not pdf
#         # if content_type != "application/pdf":
#         #     return PlainTextResponse("File type is not support!", status_code=400)

#         # for pdf file in folder 12123

        
#         # if file is pdf
#         try:
#             t0 = time.time()
#             pdf_file = await file.read()

#             # Bắt đầu một coroutine mới trong một event loop riêng biệt
#             pages_processed = asyncio.create_task(read_pdf.process_pdf(pdf_file, db_client))

#             # Chờ cho đến khi coroutine hoàn thành
#             await pages_processed

#             return JSONResponse({"status": "success", "message": f"Import labels started successfully!: {pages_processed.result()} - {time.time() - t0} seconds"})

#         except:
#             return JSONResponse({"status": "error", "message": f"Import labels error - {traceback.format_exc()}!"})
@authenticate
async def import_labels(request):
    # import data from csv file
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        query = {}
        filter_ = {
            "name": 1,
        }
        if user['team'] == 'nb_team':
            teams = await db_client.find_many_combo(query, collection_name=DB['COL_TEAMS'], filter_=filter_, sort_="_id")
        else:
            teams = [{
                'name': user['team'],
            }]
        data = {
            "meta_title": "Import labels",
            "teams": teams,
        }
        return templates.TemplateResponse('import/import_labels.html', {'request': request, 'data': data})
    elif request.method == "POST":
        form = await request.form()
        file = form["file"]
        content_type = file.headers["content-type"]

        user = request.state.user_data

        # if file not pdf
        if content_type != "application/pdf":
            return PlainTextResponse("File type is not support!", status_code=400)
        
        # if file is pdf
        try:
            t0 = time.time()
            pdf_file = await file.read()
            ok = await read_pdf.read_pdf(pdf_file, db_client)
            print(f"Import labels finished successfully!: {time.time() - t0} seconds")
            if ok == True:
                return JSONResponse({"status": "success", "message": f"Import labels successfully!"})
            else:
                return JSONResponse({"status": "error", "message": f"Import labels error!"})
            
        except:
            return JSONResponse({"status": "error", "message": f"Import labels error - {traceback.format_exc()}!"})

async def bulk_edit_by_csv(request):
    # import data from csv file
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        query = {}
        filter_ = {
            "name": 1,
        }
        if user['team'] == 'nb_team':
            teams = await db_client.find_many_combo(query, collection_name=DB['COL_TEAMS'], filter_=filter_, sort_="_id")
        else:
            teams = [{
                'name': user['team'],
            }]
        data = {
            "meta_title": "Import orders",
            "teams": teams,
        }
        return templates.TemplateResponse('fulfillment/bulk_edit_by_csv.html', {'request': request, 'data': data})
    
    elif request.method == "POST":
        form = await request.form()
        file = form["file"]
        content_type = file.headers["content-type"]

        user = request.state.user_data
        # team_name = await db_client.find_one({"username": {"$in": [user['username']]}}, collection_name=DB['COL_TEAMS'])
        # if file is csv 
        try:
            if content_type == "text/csv":
                df = pd.read_csv(file.file, encoding='utf-8')
            # if file is xlsx
            elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                df = pd.read_excel(file.file)
            else:
                return PlainTextResponse("File type is not support!", status_code=400)
        except:
            return JSONResponse({"status": "error", "message": f"Import orders error - {traceback.format_exc()}!"})
        
        df_json_string = df.to_json(orient='records', force_ascii=False)
        df_json = json.loads(df_json_string)
        orders_process = [    dict(zip((key.lower().replace(' ', '_') for key in  order.keys() ),(order.values())))  for order in df_json]
        # if df is empty
        if orders_process == []:
            return PlainTextResponse("File is empty!", status_code=400)
        
        for order in orders_process:
            order_id_ = str(order['order_id'])
            status = order['status']
            factory = order['factory'] if 'factory' in order else None
            if '-' in order_id_ and 'RE-' not in   order_id_:
                order_id = order_id_.split('-')[0]
                query = {"_original_data.order_id": order_id}
            elif 'RE-' in   order_id_:
                order_id ='RE-'+ order_id_.split('-')[1]
                query = {"_original_data.order_id": order_id}

            else:
                query = {"_original_data.order_id": order_id_}    
            data = await db_client.find_one(query, collection_name=DB['COL_FULFILLMENTS'])
            if data:
                if '-' in order_id_.replace('RE-',''):
                    for item in data['_original_data']['line_items']:
                        if order_id_.lower() == item['line_item_id'].lower().replace('.','-'):
                            if status and status != '':
                                item['fulfillment_status'] = status
                            if factory and factory != '':
                                item['factory'] = factory
                            break   
                
                elif status and status != '':
                    data['_original_data']['line_items'][0]['fulfillment_status'] = status    
                    if factory and factory != '':
                        data['_original_data']['line_items'][0]['factory'] = factory
                a = await db_client.update_one(data['_id'], data, collection_name=DB['COL_FULFILLMENTS'])       
                print(a)
        return JSONResponse({"status": "success", "message": f"Import {len(orders_process)} items successfully!"})
        


# export
@authenticate
async def export_data(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        query = {}
        filter_ = {
            "name": 1,
        }
        if user['team'] == 'nb_team':
            teams = await db_client.find_many_combo(query, collection_name=DB['COL_TEAMS'], filter_=filter_, sort_="_id")
        else:
            teams = [{
                'name': user['team'],
            }]
        data = {
            "meta_title": "Export data",
            "teams": teams,
        }
        return templates.TemplateResponse('export/export_data.html', {'request': request, 'data': data})
        

# template
async def download_csv(request):
    csv_data = "Shop name,Order date,Order id,Thumbnail,SKU,Product type,Style,Size,Color,Material,Personalized,Num color kit,Num holder,Link des,Quantity,Note,Buyer name,Address 1,Address 2,City,State,Zip,Country\n"

    # Tạo generator để gửi dữ liệu theo từng phần
    async def generate():
        yield csv_data.encode("utf-8")

    # Thiết lập phản hồi để trả về tệp tin CSV
    response = StreamingResponse(generate(), media_type="text/csv")
    response.headers["Content-Disposition"] = 'attachment; filename="template.csv"'
    return response
async def download_xlsx(request):
    # Tạo workbook và sheet
    workbook = Workbook()
    sheet = workbook.active

    # Thiết lập tiêu đề
    headers = [
        "Shop name", "Order date", "Order id", "Thumbnail", "SKU", "Product type", "Style", "Size", 
        "Color", "Material", "Personalized",'Num color kit','Num holder', "Link des", "Quantity", "Note", "Buyer name",
        "Address 1", "Address 2", "City", 
        "State", "Zip", "Country"
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Ghi workbook vào buffer tạm thời
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Thiết lập phản hồi để trả về tệp tin XLSX
    response = StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = 'attachment; filename="template.xlsx"'
    return response

async def download_tracking_csv(request):
    csv_data = "Order id,Tracking number,Carrier,Shipping price,Currency\n"

    # Tạo generator để gửi dữ liệu theo từng phần
    async def generate():
        yield csv_data.encode("utf-8")

    # Thiết lập phản hồi để trả về tệp tin CSV
    response = StreamingResponse(generate(), media_type="text/csv")
    response.headers["Content-Disposition"] = 'attachment; filename="template_tracking.csv"'
    return response
async def download_tracking_xlsx(request):
    # Tạo workbook và sheet
    workbook = Workbook()
    sheet = workbook.active

    # Thiết lập tiêu đề
    headers = [
        "Order id", "Tracking number", "Carrier", "Shipping price", "Currency"
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Ghi workbook vào buffer tạm thời
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Thiết lập phản hồi để trả về tệp tin XLSX
    response = StreamingResponse(iter([output.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = 'attachment; filename="template_tracking.xlsx"'
    return response


# Analytics
@authenticate
async def nb_team_analytics(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        # get date range 30 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        query = {
            "_team": "nb_team",
            "create_at": {'$gte': start_date, '$lte': end_date},
            "_original_data.line_items.fulfillment_status": {"$nin": ["Cancel", "Refund"]},
            "_original_data.order_id": {"$not": {"$regex": "RE"}}
        }
        filter_ = {
            "_original_data": 1,
            "create_at": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_data(orders)

        count_sales, count_items, shop_sales, seller_sales = await func.get_count_sales(db_client, orders)

        # sort count_items by value
        count_items = dict(sorted(count_items.items(), key=lambda item: item[1], reverse=True))
        count_sales = dict(sorted(count_sales.items(), key=lambda item: item[1], reverse=True))
        seller_sales = dict(sorted(seller_sales.items(), key=lambda x: x[1], reverse=True))
        shop_sales = dict(sorted(shop_sales.items(), key=lambda x: x[1], reverse=True))

    
        data = {
            "meta_title": "Analytics",
            "result": analytics_data,
            "count_sales": count_sales,
            "count_items": count_items,
            "shop_sales": shop_sales,
            "seller_sales": seller_sales,
        }
        return templates.TemplateResponse('analytics/nb_team_analytics.html', {'request': request, 'data': data})

    elif request.method == "POST":
        form_data = await request.form()
        start_date = form_data['start_date']
        end_date = form_data['end_date']

        query = {
            "_team": "nb_team",
            "create_at": {'$gte': start_date, '$lte': end_date},
        }
        filter_ = {
            "_original_data": 1,
            "create_at": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_data(orders)


        data = {
            "meta_title": "Analytics",
            "result": analytics_data,
        }
        return templates.TemplateResponse('analytics/nb_team_analytics.html', {'request': request, 'data': data})

@authenticate
async def analytics_private_team(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        # get date range 30 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        query = {
            "_team": "nb_team",
            "create_at": {'$gte': start_date, '$lte': end_date},
            "_original_data.line_items.fulfillment_status": {"$nin": ["Cancel", "Refund"]},
            "_original_data.order_id": {"$not": {"$regex": "RE"}}
        }
        filter_ = {
            "_original_data": 1,
            "create_at": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_data(orders)

        count_sales, count_items, shop_sales, seller_sales = await func.get_count_sales(db_client, orders)

        # sort count_items by value
        count_items = dict(sorted(count_items.items(), key=lambda item: item[1], reverse=True))
        count_sales = dict(sorted(count_sales.items(), key=lambda item: item[1], reverse=True))
        seller_sales = dict(sorted(seller_sales.items(), key=lambda x: x[1], reverse=True))
        shop_sales = dict(sorted(shop_sales.items(), key=lambda x: x[1], reverse=True))

        private = await db_client.find_many_combo({}, collection_name=DB['COL_SHOP'], filter_={"seller_name": 1, "shop": 1, "team": 1})
        private_team = []
        private_seller = []
        private_shop = []

        for item in private:
            if item['team'] not in private_team:
                private_team.append(item['team'])
            if item['seller_name'] not in private_seller:
                private_seller.append(item['seller_name'])
            if item['shop'] not in private_shop:
                private_shop.append(item['shop'])

        data = {
            "meta_title": "Analytics",
            "result": analytics_data,
            "count_sales": count_sales,
            "count_items": count_items,
            "shop_sales": shop_sales,
            "seller_sales": seller_sales,
            "private_team": private_team,
            "private_seller": private_seller,
            "private_shop": private_shop,
        }
        return templates.TemplateResponse('analytics/analytics_private_team.html', {'request': request, 'data': data})

    elif request.method == "POST":
        form_data = await request.form()
        start_date = form_data['start_date']
        end_date = form_data['end_date']

        query = {
            "_team": "nb_team",
            "create_at": {'$gte': start_date, '$lte': end_date},
        }
        filter_ = {
            "_original_data": 1,
            "create_at": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_data(orders)


        data = {
            "meta_title": "Analytics",
            "result": analytics_data,
        }
        return templates.TemplateResponse('analytics/analytics_private_team.html', {'request': request, 'data': data})


@authenticate 
async def other_team_analytics(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        # get date range 30 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        team = user['team']

        query = {
            "_team": team,
            "create_at": {'$gte': start_date, '$lte': end_date},
            "_original_data.line_items.fulfillment_status": {"$nin": ["Cancel", "Refund"]},
            "_original_data.order_id": {"$not": {"$regex": "RE"}}
        }
        filter_ = {
            "_original_data": 1,
            "create_at": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_data(orders)

        count_sales, count_items, shop_sales, seller_sales = await func.get_count_sales(db_client, orders)

        # sort count_items by value
        count_items = dict(sorted(count_items.items(), key=lambda item: item[1], reverse=True))
        count_sales = dict(sorted(count_sales.items(), key=lambda item: item[1], reverse=True))
        seller_sales = dict(sorted(seller_sales.items(), key=lambda x: x[1], reverse=True))
        shop_sales = dict(sorted(shop_sales.items(), key=lambda x: x[1], reverse=True))

        db_teams = await db_client.find_many_combo({}, collection_name=DB['COL_TEAMS'], filter_={"name": 1, "_id": 1})
        
        data = {
            "meta_title": "Analytics",
            "result": analytics_data,
            "count_sales": count_sales,
            "count_items": count_items,
            "shop_sales": shop_sales,
            "seller_sales": seller_sales,
        }
        return templates.TemplateResponse('analytics/other_team_analytics.html', {'request': request, 'data': data})

    elif request.method == "POST":
        form_data = await request.form()
        start_date = form_data['start_date']
        end_date = form_data['end_date']

        query = {
            "_team": team,
            "create_at": {'$gte': start_date, '$lte': end_date},
        }
        filter_ = {
            "_original_data": 1,
            "create_at": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_data(orders)


        data = {
            "meta_title": "Analytics",
            "result": analytics_data,
        }
        return templates.TemplateResponse('analytics/other_team_analytics.html', {'request': request, 'data': data})

@authenticate
async def designer_analytics(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        # get date range 30 days
        end_date = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        start_date = end_date - timedelta(days=30)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

        # convert to mongodb time
        start_date = convert_to_time(start_date, 'start')
        end_date = convert_to_time(end_date, 'end')

        query = {
            "_original_data.line_items.assign_designer.updated_at": {'$gte': start_date, '$lte': end_date},
            "_team": "nb_team",
        }
        filter_ = {
            "_original_data.line_items.assign_designer": 1,
            "_original_data.line_items.product_type": 1,
            "_original_data.line_items.fulfillment_order_id": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_designer(orders, start_date, end_date)

        data = {
            "meta_title": "Analytics",
            "records": analytics_data,
        }
        return templates.TemplateResponse('analytics/designer_analytics.html', {'request': request, 'data': data})
    
@authenticate
async def design_checked_analytics(request):
    db_client = request.state.db
    user = request.state.user_data
    if request.method == "GET":
        # get date range 30 days
        end_date = datetime.now()
        start_date = end_date - timedelta(days=30)

        start_date = convert_to_mongodb_time(start_date, 'start')
        end_date = convert_to_mongodb_time(end_date, 'end')

        query = {
            "_original_data.line_items.assign_designer.design_checked": True,
            "_original_data.line_items.assign_designer.finished_at": {'$gte': start_date, '$lte': end_date},
        }
        filter_ = {
            "_original_data.line_items.seller": 1,
            "_original_data.line_items.assign_designer": 1,
        }
        orders = await db_client.find_many_combo(query, collection_name=DB['COL_FULFILLMENTS'], filter_=filter_, sort_="_id")
        analytics_data = func.format_analytics_design_checked(orders, start_date, end_date)
        print(analytics_data)

        data = {
            "meta_title": "Analytics",
            "records": analytics_data,
        }
        return templates.TemplateResponse('analytics/design_checked_analytics.html', {'request': request, 'data': data})

# scan orders
async def scan_orders(request):
    data = {
        "meta_title": "Scan orders",
        "status_list": list_status
    }

    return templates.TemplateResponse('scan_orders/scan_orders.html', {'request': request, 'data': data})

async def personalize(request):
    db_client = request.state.db
    if request.method == "GET":
        query = {}
        filter_ = {
            "name": 1,
            "mockup":1,
            "_id":1
        }
        cliparts = await db_client.find_many_combo(query, collection_name=DB['COL_PERSONALIZE'], filter_=filter_, sort_="_id")
        return templates.TemplateResponse('/personalize/personalize_list.html', {'request': request, 'data': {'cliparts':cliparts}})

async def list_psd(request):
    db_client = request.state.db
    if request.method == "GET":
        query = {}
        filter_ = {
            "psd.width": 1,
            "psd.height":1,
            "_id":1,
            "psd.canvas_link":1,
        }
        psds = await db_client.find_many_combo(query, collection_name=DB['COL_PSD'], filter_=filter_, sort_="_id")
        print(psds)
        return templates.TemplateResponse('/personalize/list_psd.html', {'request': request, 'data': {'psds':psds}})

async def process_pdf(request):
    db_client = request.state.db
    if request.method == "GET":
       
        return templates.TemplateResponse('/personalize/process_pdf.html', {'request': request, 'data': {}})


async def customize_design(request):
    db_client = request.state.db
    if request.method == "GET":
        psd_mongo_id = request.query_params.get('_id')
        psd_query = {
            "_id": ObjectId(psd_mongo_id)
        }
        filter_ = None 
        
        psd = await db_client.find_one(psd_query, collection_name=DB['COL_PSD'], filter_=filter_)
        pdf_query = {
            "psd_mongo_id": psd_mongo_id
        }
        pdf = await db_client.find_one(pdf_query, collection_name=DB['COL_PDF'], filter_=filter_)
        clipart_query = {
            "_id": ObjectId(psd['clipart_id'])
        }
        clipart = await db_client.find_one(clipart_query, collection_name=DB['COL_PERSONALIZE'], filter_=filter_)
        psd.pop('_id')
        clipart.pop('_id')
        pdf.pop('_id')

        return templates.TemplateResponse('/personalize/customize_design.html', {'request': request, 'data': {'psd':psd, 'pdf':pdf, 'clipart':clipart}})


async def png2pdf(request):
    db_client = request.state.db
    if request.method == "GET":
        mongo_id = request.query_params.get('_id')
        if mongo_id:
            query = {
                "_id": ObjectId(mongo_id)
            }
            pdf = await db_client.find_one(query, collection_name=DB['COL_PDF'], filter_=None)
        else:
            pdf = None    
        return  templates.TemplateResponse('/personalize/png2pdf.html', {'request': request, 'data': {'pdf':pdf}})


async def list_pdf(request):
    db_client = request.state.db
    if request.method == "GET":
        query = {'type':'png_gen'}
        filter_ = {
            "pdf.width": 1,
            "pdf.height":1,
            "pdf.custom":1,
            "pdf.tolerance":1,
            'preview':1,
        }
        pdfs = await db_client.find_many_combo(query, collection_name=DB['COL_PDF'], filter_=filter_, sort_="_id")
        data =[]
        for doc in pdfs:
            doc['_id'] = str(doc['_id']) # This does the trick!
            data.append(doc)
        # print(pdfs)
        return templates.TemplateResponse('/personalize/list_pdf_normal.html', {'request': request, 'data': {'pdfs':data}})
